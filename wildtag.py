"""
wildtag.py
==========
wildtag.ai - camera trap image pipeline

  STEP 1  Enrich CSV
          Adds image_id  (hash of absolute_path + relative_path + DateTimeOriginal)
          Adds detection_id (hash of image_id + bbox coordinates)
          Saves as <original>_with_ids.csv next to the original.

  STEP 2  Sort, resize, annotate
          For each detection row:
            • Resizes image for faster human review
            • Draws bounding box + confidence score onto the image
            • Copies annotated image to  validation/<species>/<detection_id>.jpg
            • Writes validation.csv into each species folder

Run with GUI (recommended):
    python wildtag.py

Run headless:
    python wildtag.py --no-gui --csv "S:/path/to/results.csv"
"""

import os, sys, csv, re, hashlib, argparse, threading, json
import subprocess as _subprocess
from pathlib import Path
from collections import defaultdict

# On Windows a windowless (pythonw) app still pops a console window for each
# subprocess it launches (nvidia-smi, pip, relaunch, model runner) unless we
# suppress it. CREATE_NO_WINDOW does that; it's 0 (no-op) on other platforms.
_NO_WINDOW = getattr(_subprocess, "CREATE_NO_WINDOW", 0)

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    sys.exit("Pillow is not installed.\n\nRun:  pip install Pillow  then try again.")

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

# ── persist settings next to the script ───────────────────────────────────────
SETTINGS_FILE = Path(__file__).parent / "wildtag_settings.json"

def load_settings():
    if SETTINGS_FILE.exists():
        try:
            s = json.loads(SETTINGS_FILE.read_text())
            # Migrate: old gpu_upgrade_asked blocked re-checking — remove it
            if "gpu_upgrade_asked" in s:
                del s["gpu_upgrade_asked"]
                SETTINGS_FILE.write_text(json.dumps(s, indent=2))
            return s
        except Exception:
            pass
    return {}

def save_settings(d: dict):
    try:
        SETTINGS_FILE.write_text(json.dumps(d, indent=2))
    except Exception:
        pass

# ── brand colours ──────────────────────────────────────────────────────────────
# ── Colour palettes ───────────────────────────────────────────────────────────

THEMES = {
    "light": {
        "forest":      "#2D7A45",
        "leaf":        "#4CAF72",
        "mist":        "#C8D8C0",
        "canopy":      "#1A2E1E",
        "undergrowth": "#0D1A10",
        "frost":       "#F0F7F2",
        "white":       "#FFFFFF",
        "card":        "#F7FBF8",
        "border":      "#D6E8DA",
        "text":        "#1A2E1E",
        "text_muted":  "#6B8F72",
        "log_bg":      "#F2F8F4",
        "ok":          "#2D7A45",
        "skip":        "#B07D1A",
        "error":       "#C0392B",
        "warn":        "#C0392B",
        "head":        "#1A6B8A",
        "sidebar":     "#FFFFFF",
        "sidebar_fg":  "#1A2E1E",
        "sidebar_sel": "#E8F4EC",
    },
    "dark": {
        "forest":      "#4CAF72",
        "leaf":        "#74C69D",
        "mist":        "#2D4A35",
        "canopy":      "#74C69D",
        "undergrowth": "#1B4332",
        "frost":       "#1A2420",
        "white":       "#1E2B22",
        "card":        "#243029",
        "border":      "#2D4A35",
        "text":        "#D8EEE0",
        "text_muted":  "#7BAF8A",
        "log_bg":      "#161E18",
        "ok":          "#4CAF72",
        "skip":        "#D4A017",
        "error":       "#E05252",
        "warn":        "#E05252",
        "head":        "#5BB8D4",
        "sidebar":     "#141E17",
        "sidebar_fg":  "#D8EEE0",
        "sidebar_sel": "#1B4332",
    },
}

C = dict(THEMES["light"])  # mutable — swapped on theme change

# ══════════════════════════════════════════════════════════════════════════════
#  PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def _sha256_short(text, length=12):
    return hashlib.sha256(text.encode()).hexdigest()[:length]

def make_image_id(absolute_path, relative_path, datetime_original):
    raw = f"{absolute_path.strip()}|{relative_path.strip()}|{datetime_original.strip()}"
    return f"img_{_sha256_short(raw)}"

def make_detection_id(image_id, bbox_left, bbox_top, bbox_right, bbox_bottom):
    raw = f"{image_id}|{bbox_left}|{bbox_top}|{bbox_right}|{bbox_bottom}"
    return f"det_{_sha256_short(raw)}"

def sanitise_label(label):
    label = label.strip().lower()
    label = re.sub(r'[<>:"/\\|?*]', '_', label)
    label = re.sub(r'\s+', '_', label)
    return label or "unknown"


# ── Full class lists for each model ──────────────────────────────────────────
# Used to create all species folders and populate the correction dialog.
# Order: unidentifiable first, then species A-Z, then groups A-Z.

# DeepFaune v1.4 — 38 classes in training order
_DEEPFAUNE_SPECIES = [
    "badger", "beaver", "bison", "cat", "chamois", "dog", "fallow_deer",
    "fox", "genet", "goat", "hedgehog", "ibex", "lynx", "marmot", "moose",
    "mouflon", "otter", "raccoon", "raccoon_dog", "red_deer", "reindeer",
    "roe_deer", "sheep", "squirrel", "wild_boar", "wolf", "wolverine",
]
_DEEPFAUNE_GROUPS = [
    "bird", "cow", "equid", "lagomorph", "micromammal", "mustelid",
]
_DEEPFAUNE_SPECIAL = [
    "golden_jackal", "human", "porcupine", "muskrat", "vehicle",
]
# Order: [Unidentifiable, Empty] → species A-Z → groups A-Z → special A-Z
DEEPFAUNE_ALL_LABELS = (
    ["unidentifiable", "empty"]
    + sorted(_DEEPFAUNE_SPECIES)
    + sorted(_DEEPFAUNE_GROUPS)
    + sorted(_DEEPFAUNE_SPECIAL)
)

# Model ID → full label list mapping
# None means "use observed labels only, ordered dynamically"
MODEL_CLASS_LISTS = {
    "deepfaune-v1.4": DEEPFAUNE_ALL_LABELS,
    "speciesnet-global": None,
}

# SpeciesNet taxonomy group labels — used to sort into species vs group
SPECIESNET_GROUP_SUFFIXES = (
    "_family", "_order", "_class", "_genus", "_species",
)
SPECIESNET_SPECIAL = {
    "blank", "empty", "human", "vehicle", "low_confidence",
    "unclassified", "no_cv_result", "animal", "mammal",
    "bird", "vertebrate", "undefined", "unknown",
}

def get_all_labels_for_model(classifier_id: str) -> list:
    """Return the full ordered class list for a given model ID.
    Returns None for models where we use observed labels only."""
    return MODEL_CLASS_LISTS.get(classifier_id, [])


def fmt_model_size(mb):
    """Human-readable model size from a size-in-MB registry field."""
    try:
        mb = float(mb)
    except (TypeError, ValueError):
        return ""
    return f"{mb/1024:.1f} GB" if mb >= 1024 else f"{int(mb)} MB"


def model_install_state(meta, is_ready=None, cache_bundle_present=None,
                        model_dir=None):
    """Return 'installed' or 'available' for a registry entry, using the real
    downloader checks. Cache-bundle models (SpeciesNet) count as installed only
    when their extracted cache is present; other models when the ready.json
    marker exists, or - for a model pre-bundled without a marker - when the
    weights file is already on disk. Unknown / erroring cases report
    'available' so the UI offers a safe, idempotent download."""
    try:
        cb = meta.get("cache_bundle")
        if cb:
            if cache_bundle_present:
                return "installed" if cache_bundle_present(cb) else "available"
            return "available"
        if is_ready and is_ready(meta["id"]):
            return "installed"
        wf = meta.get("weights_file")
        if wf and model_dir and (Path(model_dir(meta["id"])) / wf).exists():
            return "installed"
    except Exception:
        pass
    return "available"

def build_image_path(absolute_path, relative_path):
    base = Path(absolute_path.strip())
    rel  = Path(relative_path.strip().replace("\\", os.sep).replace("/", os.sep))
    return base / rel

def _to_int(val):
    try:    return int(float(val))
    except: return None

def load_csv(csv_path):
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096); f.seek(0)
        delim  = "\t" if "\t" in sample.split("\n")[0] else ","
        reader = csv.DictReader(f, delimiter=delim)
        reader.fieldnames = [h.strip() for h in reader.fieldnames]
        rows = list(reader)
    return rows, list(reader.fieldnames)


def load_json_as_rows(json_path, log):
    """
    Parse an AddaxAI image_recognition_file.json directly into the same
    row structure that load_csv() produces, then write a results.csv
    alongside the JSON so the rest of the pipeline is unchanged.

    Returns (rows, fieldnames, csv_path) where csv_path is the newly
    written CSV file.
    """
    log("── Parsing AddaxAI JSON ─────────────────────────────────────────────", "head")

    json_path = Path(json_path)
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    # Label maps
    det_cats = data.get("detection_categories", {})
    cls_cats = data.get("classification_categories", {})

    # The absolute folder containing the images is the folder that holds the JSON
    absolute_path = str(json_path.parent)

    CSV_FIELDS = [
        "absolute_path","relative_path","data_type","label","confidence",
        "human_verified",
        "bbox_left","bbox_top","bbox_right","bbox_bottom","bbox_normalised",
        "file_height","file_width",
        "DateTimeOriginal","DateTime","DateTimeDigitized",
        "Latitude","Longitude","GPSLink","Altitude",
        "Make","Model","Flash","ExifOffset","ResolutionUnit",
        "YCbCrPositioning","XResolution","YResolution","ExifVersion",
        "ComponentsConfiguration","FlashPixVersion","ColorSpace",
        "ExifImageWidth","ISOSpeedRatings","ExifImageHeight",
        "ExposureMode","WhiteBalance","SceneCaptureType","ExposureTime",
        "Software","Sharpness","Saturation","ReferenceBlackWhite",
    ]

    # EXIF tag IDs we want, mapped to CSV column names
    EXIF_TAGS = {
        36867: "DateTimeOriginal",
        36868: "DateTimeDigitized",
        306:   "DateTime",
        271:   "Make",
        272:   "Model",
        37385: "Flash",
        34665: "ExifOffset",
        296:   "ResolutionUnit",
        531:   "YCbCrPositioning",
        282:   "XResolution",
        283:   "YResolution",
        36864: "ExifVersion",
        37121: "ComponentsConfiguration",
        40960: "FlashPixVersion",
        40961: "ColorSpace",
        40962: "ExifImageWidth",
        34855: "ISOSpeedRatings",
        40963: "ExifImageHeight",
        41986: "ExposureMode",
        41987: "WhiteBalance",
        41990: "SceneCaptureType",
        33434: "ExposureTime",
        305:   "Software",
        41994: "Sharpness",
        41993: "Saturation",
        532:   "ReferenceBlackWhite",
    }

    def extract_exif(img_path: Path) -> dict:
        """Read EXIF tags from an image file. Returns dict of column: value."""
        result = {col: "NA" for col in EXIF_TAGS.values()}
        result.update({"Latitude": "NA", "Longitude": "NA",
                       "GPSLink": "NA", "Altitude": "NA",
                       "file_width": "NA", "file_height": "NA"})
        try:
            from PIL import Image as PilImage
            from PIL.ExifTags import TAGS, GPSTAGS
            with PilImage.open(img_path) as img:
                result["file_width"]  = str(img.width)
                result["file_height"] = str(img.height)
                exif_data = img._getexif()
                if not exif_data:
                    return result
                for tag_id, value in exif_data.items():
                    if tag_id in EXIF_TAGS:
                        result[EXIF_TAGS[tag_id]] = str(value)
                # GPS
                gps_tag_id = 34853
                if gps_tag_id in exif_data:
                    gps = exif_data[gps_tag_id]
                    def to_deg(vals):
                        d, m, s = vals
                        return float(d) + float(m)/60 + float(s)/3600
                    try:
                        lat = to_deg(gps[2])
                        if gps[1] == "S": lat = -lat
                        lon = to_deg(gps[4])
                        if gps[3] == "W": lon = -lon
                        result["Latitude"]  = str(round(lat, 6))
                        result["Longitude"] = str(round(lon, 6))
                        result["GPSLink"]   = f"https://maps.google.com/?q={lat},{lon}"
                        if 6 in gps:
                            result["Altitude"] = str(float(gps[6]))
                    except Exception:
                        pass
        except Exception:
            pass
        return result

    # Cache EXIF per image path so we only read each file once
    exif_cache: dict = {}

    rows = []
    for image in data.get("images", []):
        relative_path = image.get("file", "")
        detections    = image.get("detections", [])
        if not detections:
            continue

        # Build full path and read EXIF once per image
        img_full_path = Path(absolute_path) / Path(relative_path.replace("\\", os.sep))
        if relative_path not in exif_cache:
            exif_cache[relative_path] = extract_exif(img_full_path)
        exif = exif_cache[relative_path]

        for det in detections:
            det_cat_id  = det.get("category", "")
            det_label   = det_cats.get(det_cat_id, "unknown")
            det_conf    = det.get("conf", 0.0)

            classifications = det.get("classifications", [])
            if classifications:
                top_cls_idx, top_cls_conf = classifications[0]
                label = cls_cats.get(str(top_cls_idx), det_label)
                conf  = top_cls_conf
            else:
                label = det_label
                conf  = det_conf

            # bbox in AddaxAI JSON is [x, y, w, h] normalised 0-1
            bbox = det.get("bbox", [0, 0, 0, 0])
            if len(bbox) == 4:
                bx, by, bw, bh = bbox
                bbox_left   = bx
                bbox_top    = by
                bbox_right  = bx + bw
                bbox_bottom = by + bh
                normalised  = True
            else:
                bbox_left = bbox_top = bbox_right = bbox_bottom = 0
                normalised = False

            rows.append({
                "absolute_path":   absolute_path,
                "relative_path":   relative_path,
                "data_type":       "img",
                "label":           label,
                "confidence":      str(round(conf, 5)),
                "human_verified":  "FALSE",
                "bbox_left":       str(bbox_left),
                "bbox_top":        str(bbox_top),
                "bbox_right":      str(bbox_right),
                "bbox_bottom":     str(bbox_bottom),
                "bbox_normalised": "1" if normalised else "0",
                "file_height":     exif["file_height"],
                "file_width":      exif["file_width"],
                "DateTimeOriginal":  exif["DateTimeOriginal"],
                "DateTime":          exif["DateTime"],
                "DateTimeDigitized": exif["DateTimeDigitized"],
                "Latitude":          exif["Latitude"],
                "Longitude":         exif["Longitude"],
                "GPSLink":           exif["GPSLink"],
                "Altitude":          exif["Altitude"],
                "Make":              exif["Make"],
                "Model":             exif["Model"],
                "Flash":             exif["Flash"],
                "ExifOffset":        exif["ExifOffset"],
                "ResolutionUnit":    exif["ResolutionUnit"],
                "YCbCrPositioning":  exif["YCbCrPositioning"],
                "XResolution":       exif["XResolution"],
                "YResolution":       exif["YResolution"],
                "ExifVersion":       exif["ExifVersion"],
                "ComponentsConfiguration": exif["ComponentsConfiguration"],
                "FlashPixVersion":   exif["FlashPixVersion"],
                "ColorSpace":        exif["ColorSpace"],
                "ExifImageWidth":    exif["ExifImageWidth"],
                "ISOSpeedRatings":   exif["ISOSpeedRatings"],
                "ExifImageHeight":   exif["ExifImageHeight"],
                "ExposureMode":      exif["ExposureMode"],
                "WhiteBalance":      exif["WhiteBalance"],
                "SceneCaptureType":  exif["SceneCaptureType"],
                "ExposureTime":      exif["ExposureTime"],
                "Software":          exif["Software"],
                "Sharpness":         exif["Sharpness"],
                "Saturation":        exif["Saturation"],
                "ReferenceBlackWhite": exif["ReferenceBlackWhite"],
            })

    log(f"  {len(rows):,} detections parsed from JSON.", "ok")

    # Write a results.csv alongside the JSON so the rest of the pipeline
    # can treat it identically to an AddaxAI-exported CSV
    csv_path = json_path.parent / "results.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    log(f"  Written: {csv_path.name}", "ok")
    return rows, CSV_FIELDS, csv_path


def load_input(path, log):
    """
    Accept either a .json (AddaxAI output) or a .csv (AddaxAI export).
    Returns (rows, fieldnames, csv_path) in both cases.
    """
    path = Path(path)
    if path.suffix.lower() == ".json":
        return load_json_as_rows(path, log)
    else:
        rows, fields = load_csv(path)
        return rows, fields, path

# ── Step 1 ────────────────────────────────────────────────────────────────────

def enrich_csv(input_path, log):
    log("── Step 1: Generating image and detection IDs ───────────────────────", "head")
    rows, fieldnames, csv_path = load_input(input_path, log)

    required = {"absolute_path","relative_path","bbox_left","bbox_top","bbox_right","bbox_bottom"}
    missing  = required - set(fieldnames)
    if missing:
        raise ValueError(f"Input is missing columns: {missing}\nFound: {fieldnames}")

    new_cols  = [c for c in ("image_id","detection_id") if c not in fieldnames]
    out_fields = new_cols + fieldnames
    enriched_path = csv_path.parent / (csv_path.stem + "_with_ids.csv")

    with open(enriched_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=out_fields)
        writer.writeheader()
        for row in rows:
            img_id = make_image_id(
                row.get("absolute_path",""), row.get("relative_path",""),
                row.get("DateTimeOriginal",""))
            det_id = make_detection_id(
                img_id, row.get("bbox_left",""), row.get("bbox_top",""),
                row.get("bbox_right",""), row.get("bbox_bottom",""))
            row["image_id"] = img_id
            row["detection_id"] = det_id
            writer.writerow({k: row.get(k,"") for k in out_fields})

    log(f"  {len(rows):,} rows processed.", "ok")
    log(f"  Saved: {enriched_path.name}", "ok")
    return enriched_path

# ── Step 2 ────────────────────────────────────────────────────────────────────

def pad_bbox(x0, y0, x1, y1, img_w, img_h, frac=0.12, min_px=None):
    """Expand a bbox by a buffer, proportional to the box size with a
    minimum in pixels, clamped to the image edges. Tight boxes crowd the
    animal and make small detections hard to see under the outline, so the
    minimum margin scales with the image when no explicit value is given."""
    if min_px is None:
        min_px = max(8, img_w // 130)
    px = max(min_px, int((x1 - x0) * frac))
    py = max(min_px, int((y1 - y0) * frac))
    return (max(0, x0 - px), max(0, y0 - py),
            min(img_w, x1 + px), min(img_h, y1 + py))

def draw_bbox_on_image(img, bbox, label, conf, scale):
    draw = ImageDraw.Draw(img)
    x0,y0 = int(bbox["left"]*scale), int(bbox["top"]*scale)
    x1,y1 = int(bbox["right"]*scale), int(bbox["bottom"]*scale)
    x0,y0,x1,y1 = pad_bbox(x0, y0, x1, y1, img.width, img.height)
    draw.rectangle([x0,y0,x1,y1], outline="#00FF00",
                   width=max(2, min(6, img.width // 350)))
    text = f"{label}  {conf:.2f}"
    # Font size proportional to image width — readable after thumbnail
    fsz  = max(32, img.width // 25)
    try:    font = ImageFont.truetype("arial.ttf", fsz)
    except: font = ImageFont.load_default(size=fsz)
    # Place the label above the box if there is room, below it otherwise,
    # so it never sits on top of the detection itself
    _, mt, _, mb = draw.textbbox((0,0), text, font=font)
    th = mb - mt
    label_y = y0 - th if y0 - th >= 0 else y1
    tb = draw.textbbox((x0,label_y), text, font=font)
    draw.rectangle(tb, fill="#00FF00")
    draw.text((x0,label_y), text, fill="#000000", font=font)
    return img

def resize_draw_save(src, dst, quality, max_long_edge, bbox, label, conf,
                     bbox_normalised=False, sibling_boxes=None):
    dst.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(src) as img:
        img = img.convert("RGB")
        w,h = img.size; scale = 1.0
        if max_long_edge:
            le = max(w,h)
            if le > max_long_edge:
                scale = max_long_edge/le
                img   = img.resize((int(w*scale),int(h*scale)), Image.LANCZOS)

        # If bbox coords are normalised (0-1), convert to pixels using original dims
        if bbox_normalised and all(v is not None for v in bbox.values()):
            bbox = {
                "left":   int(bbox["left"]   * w),
                "top":    int(bbox["top"]    * h),
                "right":  int(bbox["right"]  * w),
                "bottom": int(bbox["bottom"] * h),
            }

        # Draw sibling detections (other detections in this same source
        # image) permanently, before the green focal box so they sit
        # behind it. Padded thin amber boxes, no label. Doing this once
        # at sort time means the Validate gallery never needs to load the
        # master CSV or redraw them.
        if sibling_boxes:
            sdraw = ImageDraw.Draw(img)
            for sb, s_norm, s_label in sibling_boxes:
                if not all(v is not None for v in sb.values()):
                    continue
                if s_norm:
                    b = {
                        "left":   int(sb["left"]   * w),
                        "top":    int(sb["top"]    * h),
                        "right":  int(sb["right"]  * w),
                        "bottom": int(sb["bottom"] * h),
                    }
                else:
                    b = sb
                x0 = max(0, int(b["left"]   * scale))
                y0 = max(0, int(b["top"]    * scale))
                x1 = min(img.width,  int(b["right"]  * scale))
                y1 = min(img.height, int(b["bottom"] * scale))
                if x1 <= x0 or y1 <= y0:
                    continue
                # Sibling humans get the same GDPR pixelation as focal
                # humans, previously a person appearing as a secondary
                # detection in an animal's validation image was left
                # unblurred
                if s_label.lower() in ("human", "person", "homo_sapiens"):
                    region = img.crop((x0, y0, x1, y1))
                    region = region.resize((8, 8), Image.BOX)
                    region = region.resize((x1-x0, y1-y0), Image.NEAREST)
                    img.paste(region, (x0, y0))
                px0,py0,px1,py1 = pad_bbox(x0, y0, x1, y1,
                                           img.width, img.height)
                sdraw.rectangle([px0,py0,px1,py1],
                                outline="#FFC107",
                                width=max(2, min(4, img.width // 400)))

        # Draw bbox only if all coords are present and box has non-zero area
        if (all(v is not None for v in bbox.values()) and
                bbox["right"] > bbox["left"] and
                bbox["bottom"] > bbox["top"]):

            # Blur the person if label is human — GDPR privacy protection
            if label.lower() in ("human", "person", "homo_sapiens"):
                x0 = max(0, int(bbox["left"]   * scale))
                y0 = max(0, int(bbox["top"]    * scale))
                x1 = min(img.width,  int(bbox["right"]  * scale))
                y1 = min(img.height, int(bbox["bottom"] * scale))
                if x1 > x0 and y1 > y0:
                    region = img.crop((x0, y0, x1, y1))
                    # Pixelate to 8x8 — unrecognisable but not a solid block
                    region = region.resize((8, 8), Image.BOX)
                    region = region.resize((x1-x0, y1-y0), Image.NEAREST)
                    img.paste(region, (x0, y0))

            img = draw_bbox_on_image(img, bbox, label, conf, scale)

        img.save(dst, format="JPEG", quality=quality, optimize=True)

def sort_detections(enriched_csv, quality, max_long_edge, log,
                    classifier_id=""):
    log("\n── Step 2: Sorting, resizing and annotating images ──────────────────", "head")
    rows, fieldnames = load_csv(enriched_csv)

    required = {"detection_id","image_id","absolute_path","relative_path",
                "label","confidence","bbox_left","bbox_top","bbox_right","bbox_bottom"}
    missing = required - set(fieldnames)
    if missing:
        raise ValueError(f"Enriched CSV missing columns: {missing}")

    collection_root = enriched_csv.parent
    validation_dir  = collection_root / "validation"

    import shutil
    # Resumable / crash-safe sorting. Previously this deleted the whole
    # validation\ folder up front, so a crash partway through destroyed the
    # old output AND left an incomplete new one (no validation.csv /
    # valid_species.txt, which are written only after the loop). Now we
    # don't wipe: if a sort is already partly done for THIS dataset, we
    # resume, skipping images whose output already exists. A marker records
    # which enriched CSV the in-progress folder belongs to, so we only
    # resume on a match.
    marker = validation_dir / ".sort_source"
    if validation_dir.exists():
        try:
            prev_src = marker.read_text(encoding="utf-8").strip() if marker.exists() else ""
        except Exception:
            prev_src = ""
        if prev_src == str(enriched_csv.resolve()):
            log("  Found an incomplete validation folder for this project - "
                "resuming (already-sorted images will be skipped).", "head")
        else:
            shutil.rmtree(validation_dir)
    validation_dir.mkdir(parents=True, exist_ok=True)
    try:
        marker.write_text(str(enriched_csv.resolve()), encoding="utf-8")
    except Exception:
        pass

    log(f"  Output: {validation_dir}", "ok")

    # Use the full class list for this model, falling back to observed labels
    full_labels = get_all_labels_for_model(classifier_id)
    observed_labels = sorted({sanitise_label(r.get("label","unknown")) for r in rows})

    # folder_labels: only observed classes (no empty folders)
    # correction_labels: full ordered list for valid_species.txt (includes unidentifiable)
    if full_labels:
        folder_labels = [l for l in full_labels if l in observed_labels]
        extra = [l for l in observed_labels if l not in folder_labels]
        folder_labels = folder_labels + extra
        correction_labels = full_labels  # full list already has unidentifiable first
    elif full_labels is None:
        # SpeciesNet — dynamic ordering
        _top = {"unidentifiable", "empty"}
        species = sorted([l for l in observed_labels
                          if l not in SPECIESNET_SPECIAL
                          and not any(l.endswith(s) for s in SPECIESNET_GROUP_SUFFIXES)])
        groups  = sorted([l for l in observed_labels
                          if any(l.endswith(s) for s in SPECIESNET_GROUP_SUFFIXES)])
        special = sorted([l for l in observed_labels
                          if l in SPECIESNET_SPECIAL and l not in _top])
        folder_labels     = species + groups + special
        correction_labels = ["unidentifiable", "empty"] + species + groups + special
    else:
        folder_labels     = observed_labels
        correction_labels = ["unidentifiable", "empty"] + observed_labels

    all_labels = folder_labels  # used for folder creation and CSV writing

    validation_rows = defaultdict(list)
    success = skipped = 0
    errors  = []

    for row_num, row in enumerate(rows, start=2):
        det_id  = row.get("detection_id","").strip()
        img_id  = row.get("image_id","").strip()
        src     = build_image_path(row["absolute_path"], row["relative_path"])
        label   = sanitise_label(row.get("label","unknown"))
        conf_s  = row.get("confidence","NA")
        try:    conf = float(conf_s)
        except: conf = 0.0
        bbox_norm = row.get("bbox_normalised", "0") == "1"
        if bbox_norm:
            def _to_float(v):
                try: return float(v)
                except: return None
            bbox = {k: _to_float(row.get(f"bbox_{k}")) for k in ("left","top","right","bottom")}
        else:
            bbox = {k: _to_int(row.get(f"bbox_{k}")) for k in ("left","top","right","bottom")}

        if not src.exists():
            msg = f"Row {row_num}: not found - {src}"
            log(f"  SKIP  {msg}", "skip"); errors.append(msg); skipped += 1; continue

        dst = validation_dir / label / f"{det_id}.jpg"

        # Resume support: if this image was already sorted in a previous
        # (interrupted) run, don't redo the expensive decode/resize/draw.
        # Still record its row so validation.csv is complete.
        if dst.exists() and dst.stat().st_size > 0:
            success += 1
            validation_rows[label].append({
                "detection_id": det_id, "image_id": img_id,
                "image_name": f"{det_id}.jpg",
                "original_path": row["relative_path"].strip(),
                "datetime": row.get("DateTimeOriginal","").strip(),
                "label": label, "confidence": conf_s,
                "correct_label": "", "validated": "",
            })
            continue

        try:
            resize_draw_save(src, dst, quality, max_long_edge, bbox, label, conf,
                             bbox_normalised=bbox_norm)
            log(f"  OK    {src.name} → {label}/{det_id}.jpg  (conf {conf_s})", "ok")
            success += 1
            validation_rows[label].append({
                "detection_id": det_id, "image_id": img_id,
                "image_name": f"{det_id}.jpg",
                "original_path": row["relative_path"].strip(),
                "datetime": row.get("DateTimeOriginal","").strip(),
                "label": label, "confidence": conf_s,
                "correct_label": "", "validated": "",
            })
        except Exception as e:
            msg = f"Row {row_num}: {src.name} - {e}"
            log(f"  ERROR {msg}", "error"); errors.append(msg); skipped += 1

    val_fields = ["detection_id","image_id","image_name","original_path",
                  "datetime","label","confidence","correct_label","validated",
                  "bbox_left","bbox_top","bbox_right","bbox_bottom","bbox_normalised"]

    # Write validation.csv for folders that have images
    for label, vrows in validation_rows.items():
        vcsv = validation_dir / label / "validation.csv"
        with open(vcsv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=val_fields)
            w.writeheader(); w.writerows(vrows)
        log(f"  CSV   {label}/validation.csv  ({len(vrows):,} rows)", "ok")

    # Write valid_species.txt into every folder with images
    # Uses correction_labels (full ordered list with unidentifiable first)
    for lbl in all_labels:
        lbl_s    = sanitise_label(lbl)
        ref_path = validation_dir / lbl_s / "valid_species.txt"
        with open(ref_path, "w", encoding="utf-8") as f:
            for sp in correction_labels:
                f.write(f"{sp}\n")

    log(f"  REF   valid_species.txt written to {len(all_labels)} class folders", "ok")

    log("─"*60, "head")
    log(f"Done.  {success:,} images processed,  {skipped:,} skipped/errored.",
        "ok" if skipped == 0 else "skip")

    if errors:
        lp = validation_dir / "errors.log"
        lp.parent.mkdir(parents=True, exist_ok=True)
        lp.write_text("\n".join(errors))
        log(f"Error log: {lp}", "error")

    # Sort finished cleanly for this dataset — drop the in-progress marker
    # so a later launch knows the folder is complete and won't try to resume.
    try:
        marker = validation_dir / ".sort_source"
        if marker.exists():
            marker.unlink()
    except Exception:
        pass

    return success, skipped


def check_validation_complete(project_dir, enriched_csv=None):
    """Check whether a project's validation\\ folder is complete.

    Returns a dict:
      status : "complete" | "incomplete" | "missing" | "unknown"
      reason : short human-readable explanation
      missing_csv     : list of species folders lacking validation.csv
      missing_ref     : list of species folders lacking valid_species.txt
      in_progress     : True if a .sort_source marker is present (an
                        interrupted sort)

    "incomplete" means the sort was interrupted or some per-folder files are
    missing; the caller can offer to resume the sort to finish it.
    """
    from pathlib import Path as _P
    project_dir = _P(project_dir)
    vdir = project_dir / "validation"
    result = {"status": "unknown", "reason": "",
              "missing_csv": [], "missing_ref": [], "in_progress": False}

    if not vdir.exists():
        result["status"] = "missing"
        result["reason"] = "No validation folder found."
        return result

    marker = vdir / ".sort_source"
    if marker.exists():
        result["in_progress"] = True

    # Species folders are any subdirectory containing at least one image
    species_dirs = []
    for d in vdir.iterdir():
        if not d.is_dir():
            continue
        has_img = any(f.suffix.lower() in (".jpg", ".jpeg", ".png")
                      for f in d.iterdir() if f.is_file())
        if has_img:
            species_dirs.append(d)

    for d in species_dirs:
        if not (d / "validation.csv").exists():
            result["missing_csv"].append(d.name)
        if not (d / "valid_species.txt").exists():
            result["missing_ref"].append(d.name)

    if result["in_progress"]:
        result["status"] = "incomplete"
        result["reason"] = ("A previous sort was interrupted before it "
                            "finished. It can be resumed to complete the "
                            "validation folders.")
    elif result["missing_csv"] or result["missing_ref"]:
        result["status"] = "incomplete"
        bits = []
        if result["missing_csv"]:
            bits.append(f"{len(result['missing_csv'])} folder(s) missing validation.csv")
        if result["missing_ref"]:
            bits.append(f"{len(result['missing_ref'])} folder(s) missing valid_species.txt")
        result["reason"] = ("Validation folder looks incomplete: "
                            + "; ".join(bits) + ".")
    elif not species_dirs:
        result["status"] = "missing"
        result["reason"] = "Validation folder exists but contains no sorted images."
    else:
        result["status"] = "complete"
        result["reason"] = f"{len(species_dirs)} species folder(s), all with data files."

    return result

def run_pipeline(csv_path, quality, max_long_edge, log):
    enriched = enrich_csv(csv_path, log)
    return sort_detections(enriched, quality, max_long_edge, log)



def sort_detections_counted(enriched_csv, quality, max_long_edge, log,
                            classifier_id=""):
    """Wrapper around sort_detections that also returns per-species counts."""
    rows, fieldnames = load_csv(enriched_csv)
    required = {"detection_id","image_id","absolute_path","relative_path",
                "label","confidence","bbox_left","bbox_top","bbox_right","bbox_bottom"}
    missing = required - set(fieldnames)
    if missing:
        raise ValueError(f"Enriched CSV missing columns: {missing}")

    collection_root = enriched_csv.parent
    validation_dir  = collection_root / "validation"

    import shutil
    # Resumable / crash-safe (see sort_detections for the full rationale):
    # don't wipe the folder up front; resume if a prior sort of THIS dataset
    # was interrupted, else start fresh.
    marker = validation_dir / ".sort_source"
    if validation_dir.exists():
        try:
            prev_src = marker.read_text(encoding="utf-8").strip() if marker.exists() else ""
        except Exception:
            prev_src = ""
        if prev_src == str(enriched_csv.resolve()):
            log("  Found an incomplete validation folder for this project - "
                "resuming (already-sorted images will be skipped).", "head")
        else:
            shutil.rmtree(validation_dir)
    validation_dir.mkdir(parents=True, exist_ok=True)
    try:
        marker.write_text(str(enriched_csv.resolve()), encoding="utf-8")
    except Exception:
        pass

    log(f"  Output: {validation_dir}", "ok")

    # Use full class list for this model
    full_labels = get_all_labels_for_model(classifier_id)
    observed_labels = sorted({sanitise_label(r.get("label","unknown")) for r in rows})
    if full_labels:
        folder_labels = [l for l in full_labels if l in observed_labels]
        extra = [l for l in observed_labels if l not in folder_labels]
        folder_labels     = folder_labels + extra
        correction_labels = full_labels
    elif full_labels is None:
        _top    = {"unidentifiable", "empty"}
        species = sorted([l for l in observed_labels
                          if l not in SPECIESNET_SPECIAL
                          and not any(l.endswith(s) for s in SPECIESNET_GROUP_SUFFIXES)])
        groups  = sorted([l for l in observed_labels
                          if any(l.endswith(s) for s in SPECIESNET_GROUP_SUFFIXES)])
        special = sorted([l for l in observed_labels
                          if l in SPECIESNET_SPECIAL and l not in _top])
        folder_labels     = species + groups + special
        correction_labels = ["unidentifiable", "empty"] + species + groups + special
    else:
        folder_labels     = observed_labels
        correction_labels = ["unidentifiable", "empty"] + observed_labels

    all_labels = folder_labels

    validation_rows = defaultdict(list)
    species_counts  = defaultdict(int)
    success = skipped = 0
    errors  = []

    # Group all detections by source image so each validation image gets
    # its sibling detections drawn in permanently at sort time
    by_image = defaultdict(list)
    for r in rows:
        by_image[r.get("image_id","").strip()].append(r)

    def _sibling_boxes_for(img_id, det_id):
        sibs = []
        for s in by_image.get(img_id, []):
            if s.get("detection_id","").strip() == det_id:
                continue
            s_norm = s.get("bbox_normalised", "0") == "1"
            sb = {}
            for k in ("left","top","right","bottom"):
                v = s.get(f"bbox_{k}")
                try:
                    sb[k] = float(v) if s_norm else _to_int(v)
                except (TypeError, ValueError):
                    sb[k] = None
            sibs.append((sb, s_norm, sanitise_label(s.get("label",""))))
        return sibs

    log("\n── Step 2: Sorting, resizing and annotating images ──────────────────", "head")

    for row_num, row in enumerate(rows, start=2):
        det_id  = row.get("detection_id","").strip()
        img_id  = row.get("image_id","").strip()
        src     = build_image_path(row["absolute_path"], row["relative_path"])
        label   = sanitise_label(row.get("label","unknown"))
        conf_s  = row.get("confidence","NA")
        try:    conf = float(conf_s)
        except: conf = 0.0
        bbox_norm = row.get("bbox_normalised", "0") == "1"
        if bbox_norm:
            # Normalised 0-1 floats - keep as float for conversion at draw time
            def _to_float(v):
                try: return float(v)
                except: return None
            bbox = {k: _to_float(row.get(f"bbox_{k}")) for k in ("left","top","right","bottom")}
        else:
            bbox = {k: _to_int(row.get(f"bbox_{k}")) for k in ("left","top","right","bottom")}

        if not src.exists():
            msg = f"Row {row_num}: not found - {src}"
            log(f"  SKIP  {msg}", "skip"); errors.append(msg); skipped += 1; continue

        dst = validation_dir / label / f"{det_id}.jpg"

        # Resume: skip images already sorted by a prior interrupted run,
        # but still record the row (and count) so outputs stay complete.
        if dst.exists() and dst.stat().st_size > 0:
            success += 1
            species_counts[label] += 1
            validation_rows[label].append({
                "detection_id": det_id, "image_id": img_id,
                "image_name": f"{det_id}.jpg",
                "original_path": row["relative_path"].strip(),
                "datetime": row.get("DateTimeOriginal","").strip(),
                "label": label, "confidence": conf_s,
                "correct_label": "", "validated": "",
            })
            continue

        try:
            resize_draw_save(src, dst, quality, max_long_edge, bbox, label, conf,
                             bbox_normalised=bbox_norm,
                             sibling_boxes=_sibling_boxes_for(img_id, det_id))
            log(f"  OK    {src.name} → {label}/{det_id}.jpg  (conf {conf_s})", "ok")
            success += 1
            species_counts[label] += 1
            validation_rows[label].append({
                "detection_id": det_id, "image_id": img_id,
                "image_name": f"{det_id}.jpg",
                "original_path": row["relative_path"].strip(),
                "datetime": row.get("DateTimeOriginal","").strip(),
                "label": label, "confidence": conf_s,
                "correct_label": "", "validated": "",
            })
        except Exception as e:
            msg = f"Row {row_num}: {src.name} - {e}"
            log(f"  ERROR {msg}", "error"); errors.append(msg); skipped += 1

    val_fields = ["detection_id","image_id","image_name","original_path",
                  "datetime","label","confidence","correct_label","validated"]
    for label, vrows in validation_rows.items():
        vcsv = validation_dir / label / "validation.csv"
        with open(vcsv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=val_fields)
            w.writeheader(); w.writerows(vrows)
        log(f"  CSV   {label}/validation.csv  ({len(vrows):,} rows)", "ok")

    # Write valid_species.txt into every folder with images
    for lbl in all_labels:
        lbl_s    = sanitise_label(lbl)
        ref_path = validation_dir / lbl_s / "valid_species.txt"
        with open(ref_path, "w", encoding="utf-8") as f:
            for sp in correction_labels:
                f.write(f"{sp}\n")

    log(f"  REF   valid_species.txt written to {len(all_labels)} class folders", "ok")

    # Marker: sibling boxes are baked into these images, so the Validate
    # gallery (here or on a volunteer's machine) skips loading the master
    # CSV and drawing them live
    (validation_dir / ".siblings_baked").write_text("1", encoding="utf-8")

    log("─"*60, "head")
    log(f"Done.  {success:,} images processed,  {skipped:,} skipped/errored.",
        "ok" if skipped == 0 else "skip")

    if errors:
        lp = validation_dir / "errors.log"
        lp.parent.mkdir(parents=True, exist_ok=True)
        lp.write_text("\n".join(errors))
        log(f"Error log: {lp}", "error")

    # Completed cleanly — drop the in-progress marker.
    try:
        marker = validation_dir / ".sort_source"
        if marker.exists():
            marker.unlink()
    except Exception:
        pass

    return success, skipped, dict(species_counts)


# ══════════════════════════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════════════════════════

class WildTagApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("wildtag.ai")
        self.configure(bg=C["frost"])
        self.minsize(820, 640)
        self.geometry("960x720")

        # On Windows, a pythonw.exe-hosted app shows Python's icon in the
        # taskbar (Windows groups it under the python executable) even after
        # iconbitmap sets the window icon. Declaring our own AppUserModelID
        # makes Windows treat wildtag as its own application, so the taskbar
        # shows the wildtag icon. Harmless / no-op on non-Windows.
        try:
            import ctypes
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                "wildtag.ai.desktop.1")
        except Exception:
            pass

        ico = Path(__file__).parent / "wildtag.ico"
        if ico.exists():
            try:
                self.iconbitmap(default=str(ico))
                self.after(100, lambda: self.iconbitmap(str(ico)))
            except Exception:
                pass

        self._settings    = load_settings()
        # Apply saved theme before building UI
        theme = self._settings.get("theme", "light")
        C.update(THEMES.get(theme, THEMES["light"]))
        validate_only = not (Path(__file__).parent / "wildtag_env").exists()
        if not validate_only:
            self._repair_venv_shebangs()
        default_pane  = "validate" if validate_only else "run"
        self._active_pane = self._settings.get("last_pane", default_pane)
        # In validate-only mode, never restore to a hidden pane
        if validate_only and self._active_pane in ("run", "distribute"):
            self._active_pane = "validate"
        self._last_run    = {}
        self._last_output_dir = None   # summary data from most recent run

        # Aggregate cache — computed once from results_with_ids.csv and
        # reused by Map, Summary and the confusion matrix, so switching
        # panes never re-parses the (potentially huge) master file. Keyed
        # on (path, mtime) so it rebuilds only when the file changes.
        self._agg_cache = None          # dict of precomputed aggregates
        self._agg_cache_key = None      # (path_str, mtime)
        self._agg_loading = False

        # Run-state tracking, used by the stop button and to stop the
        # theme switcher restarting the app out from under an active run
        self._job_running = False
        self._stop_event  = threading.Event()

        # Validation gallery: a small thread pool decodes tile images off
        # the UI thread (so slow network-drive reads overlap instead of
        # freezing the window), and a generation counter lets in-flight
        # decodes be discarded when the user switches species/batch.
        from concurrent.futures import ThreadPoolExecutor
        self._val_decode_pool = ThreadPoolExecutor(max_workers=6)
        self._val_gallery_gen = 0

        self._build_styles()
        self._build_ui()
        self._restore_settings()
        self._show_pane(self._active_pane)

        # Show onboarding on first launch
        if not self._settings.get("onboarding_shown") and not validate_only:
            self.after(300, self._show_onboarding)
        elif not validate_only:
            self.after(1500, self._check_gpu_upgrade)

    # ── STYLES ────────────────────────────────────────────────────────────────

    def _build_styles(self):
        self._fonts = {
            "head":  ("Segoe UI", 11, "bold"),
            "label": ("Segoe UI", 10),
            "small": ("Segoe UI", 9),
            "mono":  ("Courier New", 9),
            "stat":  ("Segoe UI", 22, "bold"),
            "h2":    ("Segoe UI", 12, "bold"),
            "tile":  ("Segoe UI", 13, "bold"),
        }

    # ── HELPERS ───────────────────────────────────────────────────────────────

    def _show_onboarding(self):
        """First-launch onboarding dialog."""
        win = tk.Toplevel(self)
        win.title("Welcome to wildtag.ai")
        win.configure(bg=C["white"])
        win.resizable(True, True)
        win.grab_set()

        # Centre on screen — taller to fit all content
        win.update_idletasks()
        w, h = 580, 680
        x = (win.winfo_screenwidth()  - w) // 2
        y = max(0, (win.winfo_screenheight() - h) // 2)
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.minsize(520, 500)

        # Header bar
        header = tk.Frame(win, bg=C["canopy"], height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        # Logo + title in header
        hinner = tk.Frame(header, bg=C["canopy"])
        hinner.place(relx=0.5, rely=0.5, anchor="center")
        try:
            from PIL import Image as PilImg, ImageTk
            ico = Path(__file__).parent / "wildtag.ico"
            if ico.exists():
                img = PilImg.open(ico).convert("RGBA")
                img.thumbnail((40, 40))
                photo = ImageTk.PhotoImage(img)
                self._onboard_ico = photo
                tk.Label(hinner, image=photo,
                         bg=C["canopy"]).pack(side="left", padx=(0,10))
        except Exception:
            pass
        tk.Label(hinner,
                 text="wildtag.ai",
                 font=("Segoe UI", 22, "bold"),
                 bg=C["canopy"], fg="#ffffff").pack(side="left")

        # Scrollable body
        body_outer = tk.Frame(win, bg=C["white"])
        body_outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(body_outer, bg=C["white"],
                           highlightthickness=0)
        sb = tk.Scrollbar(body_outer, orient="vertical",
                          command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        body = tk.Frame(canvas, bg=C["white"])
        body_win = canvas.create_window((0,0), window=body, anchor="nw")

        def _resize(e):
            canvas.itemconfig(body_win, width=e.width)
        def _scroll_cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.bind("<Configure>", _resize)
        body.bind("<Configure>", _scroll_cfg)

        # Mouse wheel
        def _wheel(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _wheel)

        body_pad = tk.Frame(body, bg=C["white"])
        body_pad.pack(fill="both", expand=True, padx=32, pady=20)

        tk.Label(body_pad,
                 text="Before you get started",
                 font=self._fonts["h2"],
                 bg=C["white"], fg=C["canopy"],
                 anchor="w").pack(fill="x", pady=(0,12))

        # Step 1
        self._onboard_step(body_pad, "1",
            "Set up your project folder",
            "Create a folder for your survey. Inside it, create an images\\ folder "
            "containing your camera trap images, organised into site-specific "
            "subfolders - one per camera location. wildtag will find them "
            "automatically when you select the project folder.",
            example="my_project\\\n"
                    "  images\\\n"
                    "    site1\\   IMG_0001.JPG ...\n"
                    "    site2\\   IMG_0001.JPG ...\n"
                    "  deployment.csv")

        # Step 2
        self._onboard_step(body_pad, "2",
            "Add a deployment metadata file",
            "Create a deployment.csv in your project folder with one row per "
            "camera site. At minimum include: site name, latitude, longitude, "
            "deployment start date and retrieval date. This links your detections "
            "to geographic locations and enables Camtrap DP export.")

        # Template download button
        tmpl_row = tk.Frame(body_pad, bg=C["white"])
        tmpl_row.pack(fill="x", pady=(4, 16))
        tk.Label(tmpl_row, text="  ", bg=C["white"], width=4).pack(side="left")
        tk.Button(tmpl_row,
                  text="Save deployment.csv template",
                  command=lambda: self._dist_save_deployment_template(),
                  font=self._fonts["label"],
                  bg=C["frost"], fg=C["canopy"],
                  relief="flat", padx=10, pady=5,
                  cursor="hand2").pack(side="left")

        # Step 3
        self._onboard_step(body_pad, "3",
            "Run wildtag",
            "In the Run tab, click Browse and select your project folder. "
            "wildtag will find your images automatically, let you choose a "
            "species identification model, and save all results alongside "
            "your images folder.")

        # ── Pinned footer — always visible at bottom ──────────────────────
        footer = tk.Frame(win, bg=C["white"],
                          bd=0, relief="flat")
        footer.pack(fill="x", side="bottom")
        tk.Frame(footer, bg=C["border"], height=1).pack(fill="x")

        footer_pad = tk.Frame(footer, bg=C["white"])
        footer_pad.pack(fill="x", padx=32, pady=14)

        # Required checklist
        tk.Label(footer_pad,
                 text="Please confirm you have both of the above ready:",
                 font=self._fonts["small"],
                 bg=C["white"], fg=C["text_muted"],
                 anchor="w").pack(fill="x", pady=(0,6))

        check_frame = tk.Frame(footer_pad, bg=C["white"])
        check_frame.pack(fill="x", pady=(0,10))

        has_folder  = tk.BooleanVar(value=False)
        has_deploy  = tk.BooleanVar(value=False)

        # Proceed button — defined BEFORE checkboxes so _update_btn can reference it
        btn_row = tk.Frame(footer_pad, bg=C["white"])

        def _proceed():
            self._settings["onboarding_shown"] = True
            save_settings(self._settings)
            canvas.unbind_all("<MouseWheel>")
            win.destroy()
            self.after(500, self._check_gpu_upgrade)

        proceed_btn = tk.Button(btn_row,
                  text="I'm ready - let's go",
                  command=_proceed,
                  font=("Segoe UI", 11, "bold"),
                  bg=C["border"], fg=C["text_muted"],
                  relief="flat", padx=16, pady=8,
                  state="disabled")
        proceed_btn.pack(side="right")

        def _update_btn(*_):
            if has_folder.get() and has_deploy.get():
                proceed_btn.config(state="normal",
                                   bg=C["canopy"], fg="#ffffff",
                                   cursor="hand2")
            else:
                proceed_btn.config(state="disabled",
                                   bg=C["border"], fg=C["text_muted"],
                                   cursor="")

        tk.Checkbutton(check_frame,
                       text="I have a project folder with an images\\ subfolder "
                            "containing my camera trap images",
                       variable=has_folder,
                       command=_update_btn,
                       font=self._fonts["label"],
                       bg=C["white"], fg=C["text"],
                       activebackground=C["white"],
                       selectcolor=C["white"],
                       wraplength=480, justify="left",
                       cursor="hand2").pack(anchor="w", pady=(0,6))

        tk.Checkbutton(check_frame,
                       text="I have a deployment.csv in my project folder "
                            "with site names, coordinates and dates",
                       variable=has_deploy,
                       command=_update_btn,
                       font=self._fonts["label"],
                       bg=C["white"], fg=C["text"],
                       activebackground=C["white"],
                       selectcolor=C["white"],
                       wraplength=480, justify="left",
                       cursor="hand2").pack(anchor="w", pady=(0,4))

        # Pack button row after checkboxes
        btn_row.pack(fill="x", pady=(8,0))

    def _onboard_step(self, parent, num, title, body_text, example=None):
        """Render a single numbered step in the onboarding dialog."""
        row = tk.Frame(parent, bg=C["white"])
        row.pack(fill="x", pady=(0, 10))

        # Number badge
        badge = tk.Frame(row, bg=C["canopy"], width=26, height=26)
        badge.pack(side="left", anchor="n", padx=(0,10), pady=2)
        badge.pack_propagate(False)
        tk.Label(badge, text=num,
                 font=("Segoe UI", 11, "bold"),
                 bg=C["canopy"], fg="#ffffff").place(
                     relx=0.5, rely=0.5, anchor="center")

        # Text column
        col = tk.Frame(row, bg=C["white"])
        col.pack(side="left", fill="x", expand=True)

        tk.Label(col, text=title,
                 font=self._fonts["head"],
                 bg=C["white"], fg=C["canopy"],
                 anchor="w").pack(fill="x")

        tk.Label(col, text=body_text,
                 font=self._fonts["small"],
                 bg=C["white"], fg=C["text_muted"],
                 wraplength=430, justify="left",
                 anchor="w").pack(fill="x", pady=(2,0))

        if example:
            tk.Label(col, text=example,
                     font=("Courier New", 9),
                     bg=C["frost"], fg=C["canopy"],
                     justify="left", anchor="w",
                     padx=8, pady=4).pack(
                         fill="x", pady=(4,0))

    def _repair_venv_shebangs(self):
        """
        Rewrite baked absolute shebangs in wildtag_env's (and validate_env's)
        console-script wrappers, such as Scripts\\pip.exe's companion
        "-script.py" file, so they point at THIS folder's python.exe.

        pip bakes an absolute path into these wrappers at the moment a
        package is installed. Since wildtag_env is built once and then
        zipped up and extracted to a different folder on every machine it
        reaches, that baked path is almost never correct once deployed.
        Runs on every launch; harmless and near-instant once already fixed.
        """
        try:
            import fix_shebangs
            fix_shebangs.main()
        except Exception:
            pass

    def _check_gpu_upgrade(self):
        """
        Detect NVIDIA GPU without CUDA PyTorch and offer to install it.
        Re-checks on every launch — only skips if CUDA is working, the user
        declined, or a previous install attempt completed but CUDA still did
        not work on this GPU/driver combination (see gpu_install_failed).
        """
        # Check if CUDA already works — nothing to do
        try:
            import torch
            if torch.cuda.is_available():
                return
        except Exception:
            pass

        # Check for NVIDIA GPU
        try:
            import subprocess
            result = subprocess.run(
                ["nvidia-smi", "--query-gpu=name", "--format=csv,noheader"],
                capture_output=True, text=True, timeout=5,
                creationflags=_NO_WINDOW)
            has_nvidia = result.returncode == 0 and result.stdout.strip()
        except Exception:
            has_nvidia = False

        if not has_nvidia:
            return

        # User previously declined — don't ask again
        if self._settings.get("gpu_upgrade_declined"):
            return

        # A previous install completed but CUDA still wasn't available
        # afterwards — this GPU/driver combination isn't working, so don't
        # keep looping the install every launch. _gpu_install_diagnose lets
        # the user re-trigger the check manually once they've updated drivers.
        if self._settings.get("gpu_install_failed"):
            return

        gpu_name = result.stdout.strip().splitlines()[0]
        answer = messagebox.askyesno(
            "GPU detected",
            f"An NVIDIA GPU was detected ({gpu_name}) but GPU acceleration "
            f"is not currently enabled.\n\n"
            f"Would you like to enable it now?\n\n"
            f"This requires an internet connection and will take a few minutes. "
            f"wildtag will remain usable during the upgrade.")

        if not answer:
            self._settings["gpu_upgrade_declined"] = True
            save_settings(self._settings)
            return

        self._gpu_upgrade_dialog()

    def _gpu_upgrade_dialog(self):
        """Show a progress window and install GPU PyTorch silently."""
        import subprocess, threading

        win = tk.Toplevel(self)
        win.title("Enabling GPU acceleration")
        win.geometry("460x200")
        win.resizable(False, False)
        win.configure(bg=C["white"])
        win.grab_set()

        tk.Label(win, text="Installing GPU PyTorch...",
                 font=self._fonts["h2"], bg=C["white"],
                 fg=C["canopy"]).pack(pady=(24, 4))
        tk.Label(win,
                 text="This may take a few minutes. wildtag will restart automatically.",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"], wraplength=400).pack(pady=(0, 16))

        progress_var = tk.StringVar(value="Starting...")
        tk.Label(win, textvariable=progress_var,
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["forest"]).pack(pady=4)

        # Progress bar simulation
        bar_frame = tk.Frame(win, bg=C["border"], height=6)
        bar_frame.pack(fill="x", padx=32, pady=8)
        bar_fill = tk.Frame(bar_frame, bg=C["forest"], height=6, width=0)
        bar_fill.place(x=0, y=0, relheight=1)

        def _animate_bar(pct):
            try:
                w = int(bar_frame.winfo_width() * pct / 100)
                bar_fill.place(x=0, y=0, relheight=1, width=max(1, w))
            except Exception:
                pass

        def _install():
            # Resolve python.exe directly, relative to this file, and call
            # pip as a module ("-m pip") rather than through wildtag_env's
            # generated Scripts\pip.exe wrapper. That wrapper is a small
            # exe launcher with an absolute path to python.exe baked in at
            # the moment wildtag_env was built; if wildtag_env is zipped up
            # and extracted at a different path on another machine (exactly
            # what happens when we distribute it), that baked path no longer
            # resolves and the wrapper can fail. python.exe itself carries
            # no such baked path, so "-m pip" sidesteps the problem entirely.
            python_exe = Path(__file__).parent / "wildtag_env" / "python.exe"
            if not python_exe.exists():
                python_exe = Path(__file__).parent / "wildtag_env" / "Scripts" / "python.exe"
            if not python_exe.exists():
                python_exe = Path(__file__).parent / "wildtag_env" / "bin" / "python"

            cmd = [
                str(python_exe), "-m", "pip", "install", "torch", "torchvision",
                "--upgrade", "--force-reinstall",
                "--index-url", "https://download.pytorch.org/whl/cu118",
                "--quiet",
            ]

            try:
                progress_var.set("Downloading GPU PyTorch (this may take a few minutes)...")
                self.after(0, lambda: _animate_bar(10))

                proc = subprocess.Popen(
                    cmd, stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT, text=True,
                    creationflags=_NO_WINDOW)

                # Pulse the bar while waiting
                for i, line in enumerate(proc.stdout):
                    pct = min(90, 10 + i * 2)
                    self.after(0, lambda p=pct: _animate_bar(p))

                proc.wait()

                if proc.returncode != 0:
                    self.after(0, lambda: _animate_bar(100))
                    progress_var.set("Installation failed. Try running setup_gpu.bat manually.")
                    self.after(3000, win.destroy)
                    return

                # pip exiting 0 only means the wheel installed — it does not
                # mean CUDA actually works on this GPU/driver. Verify in a
                # fresh interpreter before trusting it and restarting.
                progress_var.set("Verifying CUDA...")
                verify = subprocess.run(
                    [str(python_exe), "-c",
                     "import torch, sys; sys.exit(0 if torch.cuda.is_available() else 1)"],
                    capture_output=True, timeout=60,
                    creationflags=_NO_WINDOW)
                self.after(0, lambda: _animate_bar(100))

                if verify.returncode == 0:
                    progress_var.set("Done! Restarting wildtag...")
                    self.after(1500, lambda: self._gpu_upgrade_complete(win))
                else:
                    self._settings["gpu_install_failed"] = True
                    save_settings(self._settings)
                    progress_var.set("Installed, but CUDA is still not available.")
                    self.after(0, lambda: messagebox.showwarning(
                        "GPU not available",
                        "GPU PyTorch installed successfully, but CUDA still is not "
                        "available on this NVIDIA GPU and driver combination.\n\n"
                        "This usually means the NVIDIA driver needs updating. "
                        "wildtag will keep using CPU processing for now and won't "
                        "ask again automatically.\n\n"
                        "After updating the driver from nvidia.com/drivers, delete "
                        "wildtag_settings.json next to wildtag.py to have wildtag "
                        "check again on next launch."))
                    self.after(500, win.destroy)

            except Exception as e:
                progress_var.set(f"Error: {e}")
                self.after(3000, win.destroy)

        threading.Thread(target=_install, daemon=True).start()

    def _gpu_upgrade_complete(self, win):
        """Restart wildtag after GPU PyTorch installed."""
        import subprocess, sys
        win.destroy()
        messagebox.showinfo("GPU enabled",
            "GPU acceleration has been enabled.\n\n"
            "wildtag will now restart.")
        # Restart the process
        subprocess.Popen([sys.executable] + sys.argv, creationflags=_NO_WINDOW)
        self.destroy()

    def _card(self, parent):
        outer = tk.Frame(parent, bg=C["border"], padx=1, pady=1)
        inner = tk.Frame(outer, bg=C["white"], padx=16, pady=12)
        inner.pack(fill="both", expand=True)
        return outer, inner

    def _section_label(self, parent, text):
        f = tk.Frame(parent, bg=C["frost"])
        tk.Label(f, text=text.upper(), font=self._fonts["small"],
                 bg=C["frost"], fg=C["text_muted"], anchor="w").pack(fill="x")
        tk.Frame(f, bg=C["mist"], height=1).pack(fill="x", pady=(2, 6))
        return f

    def _wrap_label(self, parent, text, **kw):
        """A label that wraps to fit its container width automatically."""
        kw.setdefault("bg",   parent["bg"])
        kw.setdefault("font", self._fonts["small"])
        kw.setdefault("fg",   C["text_muted"])
        lbl = tk.Label(parent, text=text, anchor="w", justify="left", **kw)
        lbl.bind("<Configure>",
                 lambda e: lbl.config(wraplength=max(100, e.width - 4)))
        return lbl

    def _browse_entry(self, parent, var, mode="file", title="Select"):
        """Entry + browse button row, returns the frame."""
        row = tk.Frame(parent, bg=C["white"])
        tk.Entry(row, textvariable=var, font=self._fonts["label"],
                 bg=C["frost"], fg=C["canopy"],
                 relief="flat", bd=4).pack(side="left", fill="x", expand=True)
        cmd = (lambda: self._pick_file(var, title)) if mode == "file" \
              else (lambda: self._pick_dir(var, title))
        tk.Button(row, text="Browse",
                  command=cmd, font=self._fonts["small"],
                  bg=C["border"], fg=C["canopy"],
                  relief="flat", padx=10, cursor="hand2").pack(side="left", padx=(8,0))
        return row

    def _pick_file(self, var, title):
        p = filedialog.askopenfilename(title=title,
            filetypes=[
                ("AddaxAI output files", "*.json *.csv"),
                ("JSON files", "*.json"),
                ("CSV files",  "*.csv"),
                ("All files",  "*.*"),
            ])
        if p: var.set(p)

    def _pick_dir(self, var, title):
        p = filedialog.askdirectory(title=title)
        if not p:
            return

        project = Path(p)

        # If this looks like a project folder check deployment file + alignment
        if "project" in title.lower() or "images" not in title.lower():
            dep_exists = any(
                (project / name).exists()
                for name in ["deployment.csv", "deployment.xlsx",
                             "deployments.csv", "deployments.xlsx"])

            if not dep_exists:
                answer = messagebox.askyesno(
                    "No deployment file found",
                    f"No deployment.csv was found in:\n{project}\n\n"
                    f"A deployment file is required to link your images to "
                    f"geographic locations and enable Camtrap DP export.\n\n"
                    f"Would you like to save a deployment template there now?",
                    icon="warning")
                if answer:
                    self._dist_save_deployment_template(project)
                messagebox.showwarning(
                    "Reminder",
                    "Remember to fill in deployment.csv before exporting "
                    "your data. You can do this from the Distribute tab.")
            else:
                # Deployment file exists — check site alignment
                self._check_deployment_alignment(project)

        var.set(p)

        # Auto-populate validate and distribute folder fields
        proj = Path(p)
        if hasattr(self, "_val_folder_var") and not self._val_folder_var.get():
            val = proj / "validation"
            self._val_folder_var.set(str(val))
        if hasattr(self, "_dist_folder_var") and not self._dist_folder_var.get():
            val = proj / "validation"
            self._dist_folder_var.set(str(val))
        if hasattr(self, "_dist_out_var") and not self._dist_out_var.get():
            dist = proj / "distribute"
            self._dist_out_var.set(str(dist))

    def _check_deployment_alignment(self, project: Path) -> bool:
        """
        Check that every site folder in images\\ has a row in deployment.csv
        and vice versa. Shows a warning dialog if there are mismatches.
        Returns True if all aligned, False if mismatches found.
        """
        images_dir = project / "images"
        if not images_dir.exists():
            return True  # no images folder yet — skip

        # Get site folders (subdirectories of images/)
        site_folders = {d.name for d in images_dir.iterdir() if d.is_dir()}
        if not site_folders:
            return True

        # Read deployment file
        dep_file = None
        for name in ["deployment.csv", "deployment.xlsx",
                     "deployments.csv", "deployments.xlsx"]:
            if (project / name).exists():
                dep_file = project / name
                break
        if not dep_file:
            return True  # already warned about missing file

        try:
            if dep_file.suffix.lower() == ".xlsx":
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(dep_file, data_only=True)
                    ws = wb.active
                    headers = [str(c.value or "").strip()
                               for c in next(ws.iter_rows())]
                    rows = [dict(zip(headers, [
                        str(v or "").strip() for v in row]))
                        for row in ws.iter_rows(min_row=2, values_only=True)]
                except ImportError:
                    return True
            else:
                with open(dep_file, newline="", encoding="utf-8-sig") as f:
                    rows = list(csv.DictReader(f))
        except Exception:
            return True

        # Find locationName column (case-insensitive)
        loc_col = None
        if rows:
            for col in rows[0].keys():
                if col.strip().lower() in ("locationname", "location_name",
                                           "site", "sitename", "site_name",
                                           "station", "stationname"):
                    loc_col = col
                    break

        if not loc_col:
            messagebox.showwarning(
                "Deployment file check",
                "Could not find a 'locationName' column in your deployment file.\n"
                "Please ensure one column is named 'locationName'.")
            return False

        dep_sites = {r[loc_col].strip() for r in rows if r.get(loc_col,"").strip()}

        missing_in_dep    = site_folders - dep_sites   # folders with no deployment row
        missing_in_images = dep_sites - site_folders   # deployment rows with no folder

        if not missing_in_dep and not missing_in_images:
            return True  # all good

        msg_parts = []
        if missing_in_dep:
            msg_parts.append(
                f"Site folders with no deployment row:\n"
                + "\n".join(f"  - {s}" for s in sorted(missing_in_dep)))
        if missing_in_images:
            msg_parts.append(
                f"Deployment rows with no matching image folder:\n"
                + "\n".join(f"  - {s}" for s in sorted(missing_in_images)))

        messagebox.showwarning(
            "Site mismatch",
            "The following sites do not match between your images folder "
            "and deployment.csv. Please fix before running.\n\n"
            + "\n\n".join(msg_parts))
        return False

    def _scrollable(self, parent):
        """Return a scrollable inner frame with smooth momentum scrolling."""
        canvas = tk.Canvas(parent, bg=C["frost"], highlightthickness=0)
        sb     = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=C["frost"])
        win   = canvas.create_window((0,0), window=inner, anchor="nw")

        def _cfg(e): canvas.configure(scrollregion=canvas.bbox("all"))
        def _rsz(e): canvas.itemconfig(win, width=e.width)
        inner.bind("<Configure>", _cfg)
        canvas.bind("<Configure>", _rsz)

        # Smooth momentum scroll using fraction-based movement
        _state = {"vel": 0.0, "job": None}

        def _momentum():
            if abs(_state["vel"]) < 0.1:
                _state["vel"] = 0.0
                return
            bbox = canvas.bbox("all")
            if bbox:
                content_h = bbox[3] - bbox[1]
                if content_h > 0:
                    fraction = _state["vel"] / content_h
                    cur = canvas.yview()[0]
                    canvas.yview_moveto(max(0.0, min(1.0, cur + fraction)))
            _state["vel"] *= 0.85
            _state["job"]  = canvas.after(16, _momentum)

        def _on_scroll(e):
            if _state["job"]:
                canvas.after_cancel(_state["job"])
            if e.num == 4:   _state["vel"] = -20.0
            elif e.num == 5: _state["vel"] =  20.0
            else:            _state["vel"] = -(e.delta / 120) * 20
            _momentum()

        def _on_enter(e):
            self.bind_all("<MouseWheel>", _on_scroll)
            self.bind_all("<Button-4>",   _on_scroll)
            self.bind_all("<Button-5>",   _on_scroll)

        def _on_leave(e):
            self.unbind_all("<MouseWheel>")
            self.unbind_all("<Button-4>")
            self.unbind_all("<Button-5>")

        canvas.bind("<Enter>", _on_enter)
        canvas.bind("<Leave>", _on_leave)
        inner.bind("<Enter>", _on_enter)
        inner.bind("<Leave>", _on_leave)

        return inner

    def _log(self, message, tag="plain"):
        def _a():
            self._log_box.configure(state="normal")
            self._log_box.insert("end", message+"\n", tag)
            self._log_box.see("end")
            self._log_box.configure(state="disabled")
        self.after(0, _a)

    def _set_status(self, text, colour=None):
        self._status_var.set(text)
        self._status_lbl.config(fg=colour or C["canopy"])

    def _start_progress(self):
        self._prog_running = True; self._prog_pos = 0
        self._tick_progress()

    def _tick_progress(self):
        if not self._prog_running: return
        w = self._prog_canvas.winfo_width()
        if w < 2: self.after(50, self._tick_progress); return
        seg = w // 4
        x0  = self._prog_pos % w
        x1  = min(x0 + seg, w)
        self._prog_canvas.coords(self._prog_bar, x0, 0, x1, 4)
        self._prog_pos = (self._prog_pos + 10) % w
        self.after(30, self._tick_progress)

    def _stop_progress(self):
        self._prog_running = False
        self._prog_canvas.coords(self._prog_bar, 0, 0, 0, 4)

    # ── MAIN UI ───────────────────────────────────────────────────────────────

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg=C["white"], pady=10)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=C["border"], height=1).pack(fill="x", side="bottom")

        lf = tk.Frame(hdr, bg=C["white"])
        lf.pack(side="left", padx=20)

        icon_c = tk.Canvas(lf, width=40, height=30,
                           bg=C["white"], highlightthickness=0)
        icon_c.pack(side="left", padx=(0,8))
        icon_c.create_rectangle(1,  4, 22, 26, outline=C["mist"],   width=2)
        icon_c.create_rectangle(18, 4, 39, 26, outline=C["forest"], width=2.5)

        wm = tk.Canvas(lf, bg=C["white"], highlightthickness=0, height=32, width=220)
        wm.pack(side="left")
        t1 = wm.create_text(0, 16, text="wild", anchor="w",
                            font=("Segoe UI",16,"bold"), fill=C["canopy"])
        x1 = wm.bbox(t1)[2]
        t2 = wm.create_text(x1, 16, text="tag", anchor="w",
                            font=("Segoe UI",16,"bold"), fill=C["forest"])
        x2 = wm.bbox(t2)[2]
        wm.create_text(x2, 16, text=".ai", anchor="w",
                       font=("Segoe UI",16), fill=C["text_muted"])
        wm.update_idletasks()
        wm.configure(width=(wm.bbox("all")[2] or 160) + 4)

        self._status_var = tk.StringVar(value="Ready to run")
        self._status_lbl = tk.Label(hdr, textvariable=self._status_var,
                                    font=self._fonts["small"],
                                    bg=C["frost"], fg=C["canopy"],
                                    padx=12, pady=4)
        self._status_lbl.pack(side="right", padx=20)

        # Body
        body = tk.Frame(self, bg=C["frost"])
        body.pack(fill="both", expand=True)

        # Sidebar
        self._sidebar = tk.Frame(body, bg=C["white"], width=190)
        self._sidebar.pack(side="left", fill="y")
        self._sidebar.pack_propagate(False)
        tk.Frame(self._sidebar, bg=C["border"], width=1).pack(side="right", fill="y")
        self._build_sidebar(self._sidebar)

        # Pane container
        self._pane_area = tk.Frame(body, bg=C["frost"])
        self._pane_area.pack(side="left", fill="both", expand=True)

        # Build all panes (only one shown at a time)
        self._panes = {}
        validate_only = not (Path(__file__).parent / "wildtag_env").exists()
        if not validate_only:
            self._build_pane_models()
            self._build_pane_run()
            self._build_pane_distribute()
            self._map_init_state()
        self._build_pane_validate()
        self._build_pane_summary()

    # ── SIDEBAR ───────────────────────────────────────────────────────────────

    def _build_sidebar(self, parent):
        tk.Frame(parent, bg=C["white"], height=12).pack()
        self._nav_btns = {}

        # Detect validate-only mode (no wildtag_env present)
        validate_only = not (Path(__file__).parent / "wildtag_env").exists()

        items = [
            ("models",     "Models",       not validate_only),
            ("run",        "Run wildtag",  not validate_only),
            ("validate",   "Validate",     True),
            ("distribute", "Distribute",   not validate_only),
            ("summary",    "Summary",      True),
        ]
        for key, label, visible in items:
            if not visible:
                continue
            f = tk.Frame(parent, bg=C["white"])
            f.pack(fill="x")
            bar = tk.Frame(f, bg=C["white"], width=3)
            bar.pack(side="left", fill="y")
            btn = tk.Label(f, text=label, font=self._fonts["label"],
                           bg=C["white"], fg=C["canopy"],
                           padx=16, pady=11, anchor="w", cursor="hand2")
            btn.pack(side="left", fill="x", expand=True)
            btn.bind("<Button-1>", lambda e, k=key: self._show_pane(k))
            f.bind("<Button-1>",   lambda e, k=key: self._show_pane(k))
            self._nav_btns[key] = (f, bar, btn)

        # In validate-only mode show a small notice
        if validate_only:
            tk.Label(parent, text="Validation mode",
                     font=self._fonts["small"], bg=C["white"],
                     fg=C["text_muted"]).pack(padx=16, pady=(4,0), anchor="w")

        tk.Frame(parent, bg=C["border"], height=1).pack(fill="x", padx=16, pady=8)

        # Theme toggle
        def _toggle_theme():
            if self._job_running:
                messagebox.showwarning(
                    "wildtag is running",
                    "Switching themes restarts wildtag, which would stop "
                    "the current run.\n\n"
                    "Stop the run first (or wait for it to finish), then "
                    "switch themes.")
                return
            current = self._settings.get("theme", "light")
            new_theme = "dark" if current == "light" else "light"
            self._settings["theme"] = new_theme
            save_settings(self._settings)
            import subprocess, sys
            subprocess.Popen([sys.executable] + sys.argv, creationflags=_NO_WINDOW)
            self.destroy()

        theme_icon = "🌙" if self._settings.get("theme","light") == "light" else "☀"
        theme_lbl  = f"{theme_icon}  Dark mode" if self._settings.get("theme","light") == "light" else f"{theme_icon}  Light mode"
        tk.Button(parent, text=theme_lbl,
                  command=_toggle_theme,
                  font=self._fonts["small"],
                  bg=C["white"], fg=C["text_muted"],
                  relief="flat", padx=16, pady=4,
                  activebackground=C["frost"],
                  cursor="hand2",
                  anchor="w").pack(fill="x", pady=(0,4))

        tk.Frame(parent, bg=C["border"], height=1).pack(fill="x", padx=16, pady=(0,8))
        tk.Label(parent, text="wildtag.ai  v0.1",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["mist"], padx=16).pack(anchor="w")

    def _show_pane(self, key):
        # Guard against a stale/removed pane key (e.g. "map" saved by an
        # older version) — fall back to a pane that actually exists.
        if key not in self._panes:
            key = "summary" if "summary" in self._panes else \
                  next(iter(self._panes), None)
            if key is None:
                return
        self._active_pane = key
        # Update sidebar highlight
        for k, (f, bar, btn) in self._nav_btns.items():
            if k == key:
                f.config(bg=C["frost"]); bar.config(bg=C["forest"])
                btn.config(bg=C["frost"], fg=C["forest"])
            else:
                f.config(bg=C["white"]); bar.config(bg=C["white"])
                btn.config(bg=C["white"], fg=C["canopy"])
        # Swap panes
        for k, pane in self._panes.items():
            if k == key: pane.pack(fill="both", expand=True)
            else:        pane.pack_forget()
        # Auto-populate summary stats when switching to it — both the
        # "What was processed" run stats and the validation progress,
        # otherwise a fresh launch shows all zeros until some other
        # button happens to be clicked first. Also prepare the map data
        # in the background so the Launch map button is ready.
        if key == "summary":
            self.after(100, self._restore_summary_stats)
            self.after(100, lambda: self._refresh_val_stats(auto=True))
            self.after(100, self._map_refresh)
        # Auto-populate validate folder from project when switching to it
        if key == "validate":
            proj = getattr(self, "_img_folder_var", None)
            proj = proj.get().strip() if proj else ""
            if proj:
                val = Path(proj) / "validation"
                if val.exists():
                    self._val_folder = val
                    self._val_folder_var.set(str(val))
                    self.after(100, self._val_populate_species)
                else:
                    if hasattr(self, "_val_status_lbl"):
                        self._val_status_lbl.config(
                            text="Run wildtag first to generate validation images")
            else:
                # validate-only mode — auto-detect from wildtag.py location
                local_val = Path(__file__).parent / "validation"
                if local_val.exists() and hasattr(self, "_val_status_lbl"):
                    self._val_folder = local_val
                    self._val_folder_var.set(str(local_val))
                    # Silently repair the manifest so a volunteer never sees
                    # 'Image not found' tiles; non-destructive, backs up first.
                    try:
                        self._val_repair_manifests(interactive=False)
                    except Exception:
                        pass
                    self.after(100, self._val_populate_species)

    # ── PANE: SETUP ───────────────────────────────────────────────────────────

    def _build_pane_setup(self):
        pane = tk.Frame(self._pane_area, bg=C["frost"])
        self._panes["setup"] = pane
        inner = self._scrollable(pane)
        pad   = dict(padx=24, pady=6)

        tk.Frame(inner, bg=C["frost"], height=8).pack()
        tk.Label(inner, text="Setup", font=self._fonts["h2"],
                 bg=C["frost"], fg=C["canopy"], anchor="w").pack(
                     fill="x", padx=24, pady=(0,2))
        self._wrap_label(inner,
            "Select your project folder and click Save. "
            "wildtag will remember this each time it opens.",
            bg=C["frost"]).pack(fill="x", padx=24)
        tk.Frame(inner, bg=C["frost"], height=12).pack()

        # Project folder
        self._section_label(inner, "Project folder").pack(
            fill="x", padx=24, pady=(0,4))
        o, c = self._card(inner); o.pack(fill="x", **pad)
        self._wrap_label(c,
            "Select the main folder for this camera trap deployment. "
            "It should contain one sub-folder per camera.",
            bg=C["white"]).pack(fill="x", pady=(0,8))
        self._project_var = tk.StringVar()
        self._browse_entry(c, self._project_var, mode="dir",
                           title="Select project folder").pack(fill="x")

        # Save button
        tk.Frame(inner, bg=C["frost"], height=8).pack()
        tk.Button(inner, text="Save settings",
                  command=self._save_setup,
                  font=self._fonts["label"],
                  bg=C["forest"], fg=C["white"],
                  activebackground=C["leaf"], activeforeground=C["white"],
                  relief="flat", padx=20, pady=8, cursor="hand2").pack(
                      anchor="e", padx=24, pady=(4,20))

    def _save_setup(self):
        s = self._settings
        s["project"]   = self._project_var.get()
        s["last_pane"] = self._active_pane
        save_settings(s)
        self._set_status("Settings saved", C["forest"])

    # ── PANE: MODELS ──────────────────────────────────────────────────────────

    def _build_pane_models(self):
        pane = tk.Frame(self._pane_area, bg=C["frost"])
        self._panes["models"] = pane
        self._models_inner = self._scrollable(pane)
        self._models_status_lbls = {}
        self._models_downloading = set()
        self._models_render()

    def _models_render(self):
        """(Re)draw the model list with current install status."""
        inner = self._models_inner
        for w in inner.winfo_children():
            w.destroy()
        self._models_status_lbls = {}

        tk.Frame(inner, bg=C["frost"], height=8).pack()
        tk.Label(inner, text="Models", font=self._fonts["h2"],
                 bg=C["frost"], fg=C["canopy"], anchor="w").pack(
                     fill="x", padx=24, pady=(0,2))
        self._wrap_label(inner,
            "wildtag ships with no model. Download the one you need while you "
            "have a connection; afterwards it runs completely offline.",
            bg=C["frost"]).pack(fill="x", padx=24)
        tk.Frame(inner, bg=C["frost"], height=10).pack()

        try:
            from wt_models.registry import classifiers
            models = classifiers()
        except Exception as e:
            self._wrap_label(inner, f"Could not read the model registry: {e}",
                             bg=C["frost"]).pack(fill="x", padx=24)
            return

        try:
            from wt_models.downloader import (is_ready, cache_bundle_present,
                                              model_dir)
        except Exception:
            is_ready = cache_bundle_present = model_dir = None

        for m in models:
            state = model_install_state(m, is_ready, cache_bundle_present,
                                        model_dir)
            self._models_card(inner, m, state)

        tk.Frame(inner, bg=C["frost"], height=8).pack()
        self._wrap_label(inner,
            "Models download once, then run offline. The first download needs "
            "internet; each file is checked before use.",
            bg=C["frost"]).pack(fill="x", padx=24, pady=(0,20))

    def _models_card(self, inner, m, state):
        mid = m.get("id", "")
        o, c = self._card(inner)
        o.pack(fill="x", padx=24, pady=6)

        tk.Label(c, text=m.get("name", mid), font=self._fonts["label"],
                 bg=C["white"], fg=C["canopy"], anchor="w").pack(fill="x")

        region = ", ".join(m.get("regions", [])) or ""
        arch   = m.get("architecture", "")
        meta_line = "  \u00b7  ".join(x for x in [region, arch] if x)
        if meta_line:
            self._wrap_label(c, meta_line, bg=C["white"]).pack(fill="x", pady=(2,0))

        cb   = m.get("cache_bundle") or {}
        size = fmt_model_size(cb.get("size_mb") or m.get("weights_size"))
        lic  = m.get("license", "")
        info = "  \u00b7  ".join(x for x in [
            f"Size: {size}" if size else "",
            f"Licence: {lic}" if lic else ""] if x)
        if info:
            self._wrap_label(c, info, bg=C["white"]).pack(fill="x", pady=(2,0))

        if m.get("description"):
            self._wrap_label(c, m["description"], bg=C["white"]).pack(
                fill="x", pady=(4,4))

        row = tk.Frame(c, bg=C["white"]); row.pack(fill="x", pady=(4,0))
        if state == "installed":
            tk.Label(row, text="\u2713 Installed", font=self._fonts["small"],
                     bg=C["white"], fg=C["forest"]).pack(side="left")
            tk.Button(row, text="Remove",
                      command=lambda mm=m: self._models_remove(mm),
                      font=self._fonts["small"], bg=C["border"], fg=C["canopy"],
                      relief="flat", padx=12, pady=4, cursor="hand2").pack(side="right")
        else:
            tk.Label(row, text="Not installed", font=self._fonts["small"],
                     bg=C["white"], fg=C["text_muted"]).pack(side="left")
            tk.Button(row, text="Download",
                      command=lambda mm=m: self._models_download(mm),
                      font=self._fonts["small"], bg=C["forest"], fg=C["white"],
                      activebackground=C["leaf"], activeforeground=C["white"],
                      relief="flat", padx=14, pady=4, cursor="hand2").pack(side="right")

        lbl = tk.Label(c, text="", font=self._fonts["small"], bg=C["white"],
                       fg=C["text_muted"], anchor="w")
        lbl.pack(fill="x", pady=(4,0))
        self._models_status_lbls[mid] = lbl

    def _models_download(self, m):
        mid  = m.get("id", "")
        if mid in getattr(self, "_models_downloading", set()):
            return
        lbl  = self._models_status_lbls.get(mid)
        cb   = m.get("cache_bundle") or {}
        size = fmt_model_size(cb.get("size_mb") or m.get("weights_size"))
        if not messagebox.askyesno("Download model",
            f"Download {m.get('name', mid)}"
            + (f" (about {size})" if size else "") + "?\n\n"
            "This needs an internet connection and happens once. Afterwards "
            "wildtag uses it offline."):
            return

        self._models_downloading.add(mid)

        def _log(msg, *a):
            if lbl:
                self.after(0, lambda: lbl.config(text=str(msg)[:120],
                                                 fg=C["text_muted"]))

        def _prog(done, total):
            if lbl and total:
                pct = int(done * 100 / max(1, total))
                self.after(0, lambda: lbl.config(text=f"Downloading... {pct}%",
                                                 fg=C["text_muted"]))

        def _work():
            try:
                from wt_models.downloader import ensure_model
                ensure_model(mid, _log, _prog)
                self.after(0, lambda: (
                    self._set_status(f"{m.get('name', mid)} installed", C["forest"]),
                    self._models_render()))
            except Exception as e:
                self.after(0, lambda: (lbl and lbl.config(
                    text=f"Download failed: {e}", fg=C["error"])))
            finally:
                self._models_downloading.discard(mid)

        threading.Thread(target=_work, daemon=True).start()

    def _models_remove(self, m):
        import shutil
        mid = m.get("id", "")
        if not messagebox.askyesno("Remove model",
            f"Remove the downloaded files for {m.get('name', mid)}?\n\n"
            "You can download it again later."):
            return
        try:
            from wt_models.downloader import model_dir
            d = Path(model_dir(mid))
            if d.exists():
                shutil.rmtree(d, ignore_errors=True)
            self._set_status(f"{m.get('name', mid)} removed", C["forest"])
        except Exception as e:
            self._set_status(f"Could not remove: {e}", C["error"])
        self._models_render()

    def _build_pane_run(self):
        pane  = tk.Frame(self._pane_area, bg=C["frost"])
        self._panes["run"] = pane
        inner = self._scrollable(pane)
        pad   = dict(padx=24, pady=6)

        tk.Frame(inner, bg=C["frost"], height=8).pack()
        tk.Label(inner, text="Run wildtag", font=self._fonts["h2"],
                 bg=C["frost"], fg=C["canopy"], anchor="w").pack(
                     fill="x", padx=24, pady=(0,2))
        self._wrap_label(inner,
            "Select your project folder, choose your models, and click Run. "
            "wildtag will detect and identify animals in all your images automatically. "
            "Processing time depends on dataset size — a typical survey of 100,000-500,000 "
            "images takes 1-5 days on a standard laptop. This is normal and expected. "
            "You can leave wildtag running overnight or over several days.",
            bg=C["frost"]).pack(fill="x", padx=24)
        tk.Frame(inner, bg=C["frost"], height=12).pack()

        # ── Image folder + models ─────────────────────────────────────────────
        o_main, c_main = self._card(inner)
        o_main.pack(fill="x", **pad)

        self._section_label(c_main,
            "Step 1 - Select your project folder").pack(
                fill="x", pady=(0,4))
        self._wrap_label(c_main,
            "Select the project folder containing your camera sub-folders. "
            "wildtag will scan all sub-folders for images.",
            bg=C["white"]).pack(fill="x", pady=(0,8))

        self._img_folder_var = tk.StringVar()
        self._browse_entry(c_main, self._img_folder_var,
                           mode="dir",
                           title="Select project folder").pack(fill="x")

        # When a project folder is chosen, and it's an already-processed
        # project (has results_with_ids.csv), start linking immediately so
        # Map and Summary are ready by the time the user gets there, rather
        # than waiting until first visit. Debounced so typing/partial paths
        # don't fire repeatedly.
        def _on_project_pick(*_):
            path = self._img_folder_var.get().strip()
            if not path:
                return
            if getattr(self, "_project_link_job", None):
                try: self.after_cancel(self._project_link_job)
                except Exception: pass
            self._project_link_job = self.after(400, self._prelink_project)
        self._img_folder_var.trace_add("write", _on_project_pick)

        tk.Frame(c_main, bg=C["white"], height=12).pack()

        self._section_label(c_main,
            "Step 2 - Choose models").pack(fill="x", pady=(0,4))
        self._wrap_label(c_main,
            "Models are downloaded automatically the first time they are used. "
            "This requires an internet connection and may take several minutes.",
            bg=C["white"]).pack(fill="x", pady=(0,10))

        # Classifier dropdown only — each model brings its own detector
        try:
            from wt_models.registry import classifiers
            cls_models = classifiers()
        except ImportError:
            cls_models = []

        tk.Label(c_main,
                 text="Species identification model",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w")
        self._cls_model_var = tk.StringVar(
            value=cls_models[0]["name"] if cls_models else "")
        om_cls = tk.OptionMenu(
            c_main, self._cls_model_var,
            *[m["name"] for m in cls_models])
        om_cls.config(font=self._fonts["label"], bg=C["frost"],
                      fg=C["canopy"], relief="flat",
                      activebackground=C["frost"],
                      highlightthickness=0, width=40)
        om_cls["menu"].config(font=self._fonts["small"],
                              bg=C["white"], fg=C["canopy"])
        om_cls.pack(anchor="w", pady=(2,10))

        # Confidence thresholds
        thresh_row = tk.Frame(c_main, bg=C["white"])
        thresh_row.pack(fill="x", pady=(0,4))

        for lbl, var_name, default in [
            ("Detection threshold (0.1 recommended)", "_det_conf_var", 0.1),
            ("Species confidence — below this = low_confidence", "_cls_conf_var", 0.6),
        ]:
            col = tk.Frame(thresh_row, bg=C["white"])
            col.pack(side="left", padx=(0,32))
            tk.Label(col, text=lbl, font=self._fonts["small"],
                     bg=C["white"], fg=C["text_muted"]).pack(anchor="w")
            var = tk.DoubleVar(value=default)
            setattr(self, var_name, var)
            tk.Spinbox(col, from_=0.0, to=1.0, increment=0.05,
                       textvariable=var, width=7,
                       font=self._fonts["label"], bg=C["frost"],
                       fg=C["canopy"], relief="flat", bd=2).pack(
                           anchor="w", pady=(2,0))

        # Geofence
        geo_row = tk.Frame(c_main, bg=C["white"])
        geo_row.pack(fill="x", pady=(8, 4))
        tk.Label(geo_row, text="Geographic filter (optional)",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w")
        self._geofence_var = tk.StringVar(value="GBR")
        geo_options = [
            ("GBR - United Kingdom", "GBR"),
            ("None - no filter",     ""),
        ]
        om_geo = tk.OptionMenu(
            geo_row, self._geofence_var,
            *[o[0] for o in geo_options])
        om_geo.config(font=self._fonts["label"], bg=C["frost"],
                      fg=C["canopy"], relief="flat",
                      activebackground=C["frost"],
                      highlightthickness=0, width=30)
        om_geo["menu"].config(font=self._fonts["small"],
                              bg=C["white"], fg=C["canopy"])
        # Map display names back to codes on selection
        def _geo_selected(*_):
            selected = self._geofence_var.get()
            for name, code in geo_options:
                if selected == name:
                    self._geofence_var.set(code)
                    return
        self._geofence_var.trace_add("write", _geo_selected)
        om_geo.pack(anchor="w", pady=(2, 0))
        tk.Label(geo_row,
                 text="  Filters out species unlikely to be found in the selected region.",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w", pady=(4, 0))

        # ── Performance options ───────────────────────────────────────────────
        self._section_label(inner,
            "Performance").pack(fill="x", padx=24, pady=(16,4))
        o_perf, c_perf = self._card(inner); o_perf.pack(fill="x", **pad)

        import os
        n_cpus = os.cpu_count() or 4
        try:
            import torch
            has_cuda = torch.cuda.is_available()
        except ImportError:
            has_cuda = False

        perf_row = tk.Frame(c_perf, bg=C["white"])
        perf_row.pack(fill="x", pady=(4, 8))

        # Device selector
        dev_col = tk.Frame(perf_row, bg=C["white"])
        dev_col.pack(side="left", padx=(0, 32))
        tk.Label(dev_col, text="Processing device",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w")
        self._device_var = tk.StringVar(value="cuda" if has_cuda else "cpu")
        dev_options = ["cuda (GPU)", "cpu"] if has_cuda else ["cpu"]
        om_dev = tk.OptionMenu(dev_col, self._device_var, *dev_options)
        om_dev.config(font=self._fonts["label"], bg=C["frost"],
                      fg=C["canopy"], relief="flat",
                      activebackground=C["frost"],
                      highlightthickness=0, width=14)
        om_dev["menu"].config(font=self._fonts["small"],
                              bg=C["white"], fg=C["canopy"])
        om_dev.pack(anchor="w", pady=(2, 0))
        if not has_cuda:
            tk.Label(dev_col,
                     text="  No NVIDIA GPU detected",
                     font=self._fonts["small"], bg=C["white"],
                     fg=C["text_muted"]).pack(anchor="w", pady=(2, 0))
            tk.Label(dev_col,
                     text="  GPU requires NVIDIA CUDA + GPU PyTorch",
                     font=self._fonts["small"], bg=C["white"],
                     fg=C["text_muted"]).pack(anchor="w")

        # Thread count
        thr_col = tk.Frame(perf_row, bg=C["white"])
        thr_col.pack(side="left", padx=(0, 32))
        tk.Label(thr_col, text=f"CPU threads (1–{max(1, n_cpus-1)} available)",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w")
        self._threads_var = tk.IntVar(value=max(1, n_cpus - 1))
        tk.Spinbox(thr_col, from_=1, to=max(1, n_cpus - 1),
                   textvariable=self._threads_var,
                   width=5, font=self._fonts["label"],
                   bg=C["frost"], fg=C["canopy"],
                   relief="flat", bd=2).pack(anchor="w", pady=(2, 0))

        tk.Label(c_perf,
                 text="  Tip: leave 1 CPU thread free for the rest of your computer.",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w", pady=(0, 6))

        # ── Output options ────────────────────────────────────────────────────
        self._section_label(inner,
            "Output options").pack(fill="x", padx=24, pady=(16,4))
        o3, c3 = self._card(inner); o3.pack(fill="x", **pad)

        opts = tk.Frame(c3, bg=C["white"])
        opts.pack(fill="x", pady=(0,12))

        tk.Label(opts, text="Validation image quality",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w", pady=(0,4))

        self._quality_var = tk.StringVar(value="medium")
        qual_row = tk.Frame(opts, bg=C["white"])
        qual_row.pack(anchor="w")
        for lbl, val, tip in [
            ("Low",    "low",    "Smaller files, faster"),
            ("Medium", "medium", "Recommended"),
            ("High",   "high",   "Largest files"),
        ]:
            rb = tk.Radiobutton(qual_row, text=lbl,
                                variable=self._quality_var, value=val,
                                font=self._fonts["label"],
                                bg=C["white"], fg=C["canopy"],
                                activebackground=C["white"],
                                selectcolor=C["white"],
                                relief="flat", cursor="hand2")
            rb.pack(side="left", padx=(0,16))

        # Map quality labels to JPEG values
        self._QUALITY_MAP = {"low": 40, "medium": 65, "high": 85}
        self._quality_var.trace_add(
            "write", lambda *_: self._persist_setting(
                "quality", self._quality_var))

        tk.Label(opts, text="Checkpoint frequency",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w", pady=(12,4))

        self._checkpoint_var = tk.StringVar(value="balanced")
        cp_row = tk.Frame(opts, bg=C["white"])
        cp_row.pack(anchor="w")
        for lbl, val, tip in [
            ("Frequent",   "frequent",   "Saves every 50 images"),
            ("Balanced",   "balanced",   "Recommended - every 200 images"),
            ("Infrequent", "infrequent", "Saves every 1,000 images"),
        ]:
            rb = tk.Radiobutton(cp_row, text=lbl,
                                variable=self._checkpoint_var, value=val,
                                font=self._fonts["label"],
                                bg=C["white"], fg=C["canopy"],
                                activebackground=C["white"],
                                selectcolor=C["white"],
                                relief="flat", cursor="hand2")
            rb.pack(side="left", padx=(0,16))

        tk.Label(opts,
                 text="  How often progress is saved to disk. More frequent "
                      "checkpoints lose less work if wildtag is stopped or "
                      "crashes, but add a small amount of extra disk "
                      "writing as the run goes on.",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"], anchor="w",
                 wraplength=560, justify="left").pack(anchor="w", pady=(2,0))

        # Map checkpoint frequency labels to "save every N images"
        self._CHECKPOINT_MAP = {
            "frequent": 50, "balanced": 200, "infrequent": 1000}
        self._checkpoint_var.trace_add(
            "write", lambda *_: self._persist_setting(
                "checkpoint", self._checkpoint_var))

        self._do_validation = tk.BooleanVar(value=True)
        chk = tk.Frame(c3, bg=C["frost"], padx=10, pady=8)
        chk.pack(fill="x", pady=2)
        tk.Checkbutton(chk, variable=self._do_validation, bg=C["frost"],
                       activebackground=C["frost"],
                       fg=C["canopy"], selectcolor=C["white"],
                       relief="flat").pack(side="left")
        lf2 = tk.Frame(chk, bg=C["frost"])
        lf2.pack(side="left")
        tk.Label(lf2, text="Sort images into species folders for review",
                 font=self._fonts["label"],
                 bg=C["frost"], fg=C["canopy"],
                 anchor="w").pack(anchor="w")
        tk.Label(lf2,
                 text="Resizes images, draws a box around each detected animal, and organises them by species for you to check",
                 font=self._fonts["small"],
                 bg=C["frost"], fg=C["text_muted"],
                 anchor="w").pack(anchor="w")

        # ── Run / Stop buttons ───────────────────────────────────────────────
        tk.Frame(inner, bg=C["frost"], height=4).pack()
        run_row = tk.Frame(inner, bg=C["frost"])
        run_row.pack(fill="x", padx=24, pady=(0,8))

        self._run_btn = tk.Button(
            run_row, text="Run wildtag",
            command=self._run,
            font=("Segoe UI", 12, "bold"),
            bg=C["forest"], fg=C["white"],
            activebackground=C["leaf"], activeforeground=C["white"],
            relief="flat", padx=20, pady=12, cursor="hand2")
        self._run_btn.pack(side="left", fill="x", expand=True)

        self._stop_btn = tk.Button(
            run_row, text="Stop",
            command=self._confirm_stop,
            font=("Segoe UI", 12, "bold"),
            bg=C["mist"], fg=C["canopy"],
            activebackground=C["error"], activeforeground=C["white"],
            relief="flat", padx=20, pady=12, cursor="hand2",
            state="disabled")
        self._stop_btn.pack(side="left", padx=(8,0))

        self._prog_canvas  = tk.Canvas(inner, bg=C["mist"], height=4,
                                       highlightthickness=0)
        self._prog_canvas.pack(fill="x", padx=24)
        self._prog_bar     = self._prog_canvas.create_rectangle(
            0, 0, 0, 4, fill=C["forest"], width=0)
        self._prog_running = False
        self._prog_pos     = 0

        # Log
        self._section_label(inner,
            "Progress log (what wildtag is doing)").pack(
                fill="x", padx=24, pady=(12,4))
        o4, c4 = self._card(inner)
        o4.pack(fill="x", padx=24, pady=(0,24))
        self._log_box = tk.Text(
            c4, height=14, font=self._fonts["mono"],
            bg=C["log_bg"], fg=C["canopy"],
            relief="flat", state="disabled", wrap="word")
        self._log_box.pack(fill="both", expand=True)
        for tag, col in [("ok",C["ok"]),("skip",C["skip"]),
                         ("error",C["error"]),("head",C["head"]),
                         ("plain",C["canopy"])]:
            self._log_box.tag_config(tag, foreground=col)


    # ── PANE: VALIDATE ────────────────────────────────────────────────────────

    BATCH_SIZE = 30

    def _build_pane_validate(self):
        pane = tk.Frame(self._pane_area, bg=C["frost"])
        self._panes["validate"] = pane

        # State
        self._val_folder       = None   # Path to validation/ folder
        self._val_species      = []     # list of species folder names
        self._val_species_var  = tk.StringVar()
        self._val_cols_var     = tk.IntVar(value=3)
        self._val_batch        = 0      # current batch index (0-based)
        self._val_rows         = []     # all rows from current validation.csv
        self._val_csv_path     = None   # Path to current validation.csv
        self._val_corrections  = {}     # detection_id -> correct_label
        self._val_selected     = set()  # detection_ids selected for multi-correct
        self._val_sibling_bboxes = {}   # image_id -> [(det_id, label, bbox, norm)]
        self._val_img_refs     = []     # keep PhotoImage refs alive
        self._val_all_labels   = []     # valid species list

        # ── Top controls (fixed, not scrollable) ──────────────────────────
        top = tk.Frame(pane, bg=C["white"])
        top.pack(fill="x")
        tk.Frame(top, bg=C["border"], height=1).pack(fill="x", side="bottom")

        ctrl = tk.Frame(top, bg=C["white"], pady=10)
        ctrl.pack(fill="x", padx=20)

        # Status label showing which project/folder is loaded
        self._val_folder_var = tk.StringVar()
        self._val_status_lbl = tk.Label(ctrl, text="No project loaded",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"])
        self._val_status_lbl.grid(row=0, column=0, sticky="w", padx=(0,20))

        # Species dropdown
        tk.Label(ctrl, text="Species",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).grid(row=0, column=1, sticky="w", padx=(0,8))
        self._val_species_menu = tk.OptionMenu(
            ctrl, self._val_species_var, "")
        self._val_species_menu.config(
            font=self._fonts["small"], bg=C["frost"],
            fg=C["canopy"], relief="flat",
            activebackground=C["frost"], highlightthickness=0, width=22)
        self._val_species_menu["menu"].config(
            font=self._fonts["small"], bg=C["white"], fg=C["canopy"])
        self._val_species_menu.grid(row=0, column=2, sticky="w")
        self._val_species_var.trace_add("write",
            lambda *_: self._val_load_species())

        # Column count
        tk.Label(ctrl, text="Columns",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).grid(row=0, column=5, sticky="w", padx=(20,8))
        for i, n in enumerate([1,2,3,4]):
            tk.Radiobutton(ctrl, text=str(n), variable=self._val_cols_var,
                           value=n, bg=C["white"], fg=C["canopy"],
                           selectcolor=C["frost"],
                           activebackground=C["white"],
                           font=self._fonts["small"],
                           command=self._val_refresh_gallery).grid(
                               row=0, column=6+i, padx=2)

        # Batch nav bar
        nav = tk.Frame(top, bg=C["frost"], pady=6)
        nav.pack(fill="x", padx=20)

        self._val_prev_btn = tk.Button(
            nav, text="< Previous batch",
            command=self._val_prev_batch,
            font=self._fonts["small"], bg=C["border"],
            fg=C["canopy"], relief="flat", padx=10, pady=4,
            cursor="hand2")
        self._val_prev_btn.pack(side="left")

        self._val_batch_lbl = tk.Label(
            nav, text="", font=self._fonts["small"],
            bg=C["frost"], fg=C["text_muted"])
        self._val_batch_lbl.pack(side="left", padx=16)

        self._val_next_btn = tk.Button(
            nav, text="Next batch >",
            command=self._val_next_batch,
            font=self._fonts["small"], bg=C["border"],
            fg=C["canopy"], relief="flat", padx=10, pady=4,
            cursor="hand2")
        self._val_next_btn.pack(side="left")

        self._val_repair_btn = tk.Button(
            nav, text="Repair folders",
            command=self._val_repair_manifests,
            font=self._fonts["small"], bg=C["border"],
            fg=C["canopy"], relief="flat", padx=10, pady=4,
            cursor="hand2")
        self._val_repair_btn.pack(side="left", padx=(20,0))

        self._val_complete_btn = tk.Button(
            nav, text="Mark batch complete",
            command=self._val_complete_batch,
            font=self._fonts["label"],
            bg=C["forest"], fg=C["white"],
            activebackground=C["leaf"], activeforeground=C["white"],
            relief="flat", padx=14, pady=4, cursor="hand2")
        self._val_complete_btn.pack(side="right")

        self._val_progress_lbl = tk.Label(
            nav, text="", font=self._fonts["small"],
            bg=C["frost"], fg=C["text_muted"])
        self._val_progress_lbl.pack(side="right", padx=16)

        # ── Scrollable gallery with smooth scroll ────────────────────────
        gallery_outer = tk.Frame(pane, bg=C["frost"])
        gallery_outer.pack(fill="both", expand=True)

        self._val_canvas = tk.Canvas(
            gallery_outer, bg=C["frost"], highlightthickness=0)
        val_sb = tk.Scrollbar(
            gallery_outer, orient="vertical",
            command=self._val_canvas.yview)
        self._val_canvas.configure(yscrollcommand=val_sb.set)
        val_sb.pack(side="right", fill="y")
        self._val_canvas.pack(side="left", fill="both", expand=True)

        self._val_gallery_inner = tk.Frame(
            self._val_canvas, bg=C["frost"])
        self._val_canvas_win = self._val_canvas.create_window(
            (0,0), window=self._val_gallery_inner, anchor="nw")

        def _val_cfg(e):
            self._val_canvas.configure(
                scrollregion=self._val_canvas.bbox("all"))
        def _val_rsz(e):
            self._val_canvas.itemconfig(
                self._val_canvas_win, width=e.width)
        self._val_gallery_inner.bind("<Configure>", _val_cfg)
        self._val_canvas.bind("<Configure>", _val_rsz)

        # Wheel / trackpad scrolling. Move by a fraction of the content
        # height per event so a notch travels a meaningful, consistent
        # distance (canvas "units" are only a few px and felt like the
        # gallery barely moved). Trackpads send many small deltas, so we
        # scale by delta magnitude and it stays smooth.
        def _val_scroll(e):
            try:
                bbox = self._val_canvas.bbox("all")
                view_h = self._val_canvas.winfo_height()
            except tk.TclError:
                return
            if not bbox:
                return
            content_h = bbox[3] - bbox[1]
            if content_h <= view_h:
                return  # nothing to scroll
            if getattr(e, "num", None) == 4:
                step_px = -80
            elif getattr(e, "num", None) == 5:
                step_px = 80
            else:
                # ~120 per mouse notch -> ~80px; trackpads scale down
                step_px = -(e.delta / 120.0) * 80
            frac = step_px / content_h
            cur = self._val_canvas.yview()[0]
            self._val_canvas.yview_moveto(max(0.0, min(1.0, cur + frac)))
            if hasattr(self, "_val_decode_visible"):
                self.after_idle(self._val_decode_visible)
            return "break"

        def _on_enter(e):
            self.bind_all("<MouseWheel>", _val_scroll)
            self.bind_all("<Button-4>",   _val_scroll)
            self.bind_all("<Button-5>",   _val_scroll)

        def _on_leave(e):
            self.unbind_all("<MouseWheel>")
            self.unbind_all("<Button-4>")
            self.unbind_all("<Button-5>")

        self._val_canvas.bind("<Enter>", _on_enter)
        self._val_canvas.bind("<Leave>", _on_leave)
        self._val_gallery_inner.bind("<Enter>", _on_enter)
        self._val_gallery_inner.bind("<Leave>", _on_leave)

        # Store for re-binding after refresh
        self._val_on_enter = _on_enter
        self._val_on_leave = _on_leave

    # ── VALIDATE: actions ─────────────────────────────────────────────────────

    def _val_pick_folder(self):
        p = filedialog.askdirectory(title="Select validation folder")
        if not p:
            return
        picked = Path(p)
        if (picked / "validation").exists():
            picked = picked / "validation"
        self._val_folder = picked
        self._val_folder_var.set(str(picked))
        self._val_populate_species()

    def _val_repair_manifests(self, interactive=True):
        """Trim each species validation.csv to the images actually present in
        its folder. Fixes 'Image not found' tiles left by older distribution
        packages that shipped the full species manifest to every batch.
        Non-destructive: backs up each original to validation.csv.bak and
        keeps any validating already done. Runs inside the app, so volunteers
        never need a separate script. With interactive=False it runs silently
        (used to auto-repair a folder the moment it is opened)."""
        if not self._val_folder or not self._val_folder.exists():
            if interactive:
                messagebox.showinfo("Nothing to repair",
                    "Open a validation folder first.")
            return 0

        csvs = [c for c in sorted(self._val_folder.rglob("validation.csv"))
                if "validate_env" not in c.parts]
        if not csvs:
            if interactive:
                messagebox.showinfo("Nothing to repair",
                    "No validation folders found here.")
            return 0

        if interactive and not messagebox.askyesno("Repair validation folders",
            "This removes entries for images that are not in each folder, so "
            "'Image not found' tiles disappear.\n\n"
            "Your images and any validating already done are kept, and a "
            "backup (validation.csv.bak) is saved first.\n\nContinue?"):
            return 0

        img_exts = (".jpg", ".jpeg", ".png")
        fixed = 0
        details = []
        for c in csvs:
            folder = c.parent
            present = {p.name.lower() for p in folder.iterdir()
                       if p.suffix.lower() in img_exts}
            try:
                rows, fields = load_csv(c)
            except Exception:
                continue
            if not fields or "image_name" not in fields:
                continue
            keep = [r for r in rows
                    if Path(r.get("image_name","")).name.lower() in present]
            if rows and not keep:
                continue            # no matching images here: leave untouched
            if len(keep) == len(rows):
                continue            # already correct
            bak = c.with_suffix(".csv.bak")
            if not bak.exists():
                bak.write_bytes(c.read_bytes())
            with open(c, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields)
                w.writeheader()
                w.writerows(keep)
            details.append(f"{folder.name}:  {len(rows)} -> {len(keep)} images")
            fixed += 1

        if interactive:
            # Reload the gallery so the change shows immediately.
            self._val_batch = 0
            self._val_populate_species()
            if fixed:
                msg = ("Repaired {0} folder(s):\n\n".format(fixed)
                       + "\n".join(details[:20]))
                if len(details) > 20:
                    msg += "\n... and {0} more".format(len(details) - 20)
                msg += "\n\nBackups saved as validation.csv.bak."
            else:
                msg = "Everything was already correct. Nothing needed changing."
            messagebox.showinfo("Repair complete", msg)
        return fixed

    def _val_populate_species(self):
        """Populate the species dropdown from the current validation folder."""
        if not self._val_folder or not self._val_folder.exists():
            if hasattr(self, "_val_status_lbl"):
                self._val_status_lbl.config(
                    text="No project loaded — run wildtag first")
            return
        # Fast path: if the run summary lists species, use that set to
        # decide which folders to even look at, skipping a per-folder image
        # existence scan. We still must read each validation.csv to know if
        # anything is pending (the summary is written at sort time, before
        # any validation), but this avoids the directory-listing cost per
        # species on a big project.
        summary_species = None
        try:
            summ = self._val_folder.parent / "wildtag_run_summary.txt"
            if summ.exists():
                names = set()
                in_counts = False
                for line in summ.read_text(encoding="utf-8").splitlines():
                    s = line.strip()
                    if s.startswith("Species counts"):
                        in_counts = True; continue
                    if in_counts:
                        parts = s.rsplit(None, 1)
                        if len(parts) == 2 and parts[1].replace(",","").isdigit():
                            # stored as "Red Deer" -> folder "red_deer"
                            names.add(parts[0].lower().replace(" ", "_"))
                if names:
                    summary_species = names
        except Exception:
            summary_species = None

        species = []
        for d in sorted(self._val_folder.iterdir()):
            if not d.is_dir():
                continue
            val_csv = d / "validation.csv"
            if not val_csv.exists():
                continue
            # If we have a summary species set, only bother with folders it
            # names (skips the image-existence disk check entirely)
            if summary_species is not None:
                if d.name.lower() not in summary_species:
                    continue
            else:
                has_image = next(
                    (True for f in d.iterdir()
                     if f.suffix.lower() in (".jpg", ".jpeg", ".png")), False)
                if not has_image:
                    continue
            try:
                with open(val_csv, newline="", encoding="utf-8") as f:
                    pending = any(
                        r.get("validated","").strip().lower() != "yes"
                        for r in csv.DictReader(f))
                if pending:
                    species.append(d.name)
            except Exception:
                continue

        self._val_species = species
        n_pending = sum(1 for sp in species for _ in [sp])
        if hasattr(self, "_val_status_lbl"):
            proj_name = self._val_folder.parent.name
            self._val_status_lbl.config(
                text=f"{proj_name}  —  {len(species)} species pending"
                     if species else f"{proj_name}  —  all validated")
        menu = self._val_species_menu["menu"]
        menu.delete(0, "end")
        for sp in species:
            menu.add_command(
                label=sp.replace("_", " ").title(),
                command=lambda s=sp: self._val_species_var.set(s))

        # Deliberately do NOT auto-select a species. Landing on the
        # Validate pane should render nothing (fast, no image decoding)
        # until the user explicitly chooses a folder from the dropdown.
        self._val_species_var.set("")
        for w in self._val_gallery_inner.winfo_children():
            w.destroy()
        if species:
            tk.Label(self._val_gallery_inner,
                     text="Choose a species from the dropdown above to begin.",
                     font=self._fonts["h2"], bg=C["frost"],
                     fg=C["text_muted"], anchor="w").pack(
                         fill="x", padx=16, pady=32)
        else:
            tk.Label(self._val_gallery_inner,
                     text="All species folders have been fully validated.",
                     font=self._fonts["h2"], bg=C["frost"],
                     fg=C["forest"], anchor="w").pack(
                         fill="x", padx=16, pady=32)

    def _val_load_species(self):
        sp = self._val_species_var.get()
        if not sp or not self._val_folder:
            return

        csv_path = self._val_folder / sp / "validation.csv"
        if not csv_path.exists():
            return

        self._val_load_favourites()
        self._val_csv_path   = csv_path
        self._val_corrections = {}
        self._val_selected    = set()  # multi-select: set of det_ids

        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            self._val_rows = list(reader)

        # Load class list from valid_species.txt
        txt = self._val_folder / sp / "valid_species.txt"
        if txt.exists():
            lines = txt.read_text(encoding="utf-8").splitlines()
            self._val_all_labels = [l.strip() for l in lines if l.strip()]
        else:
            self._val_all_labels = []

        # Merge in user-defined custom labels (subspecies, coarser groupings
        # like "unidentifiable_ungulate", local names, anything the model's
        # own classes don't cover). Stored once per project in
        # validation/custom_species.txt, so they survive a re-sort and apply
        # across every species folder.
        self._val_custom_labels = self._val_load_custom_labels()
        for lbl in self._val_custom_labels:
            if lbl not in self._val_all_labels:
                self._val_all_labels.append(lbl)

        # Build image_id -> list of rows lookup for sibling bbox drawing.
        # The master results_with_ids.csv covers the WHOLE project, so
        # after a big run it can be very large. Parse it once in a
        # background thread and cache the result (keyed on file mtime),
        # instead of re-parsing the entire file on the UI thread every
        # time a species is opened, which froze the window for a long
        # time on large projects. The gallery renders immediately, and
        # sibling boxes appear once the parse completes.
        self._val_sibling_bboxes = {}
        # If this project was sorted with baked-in sibling boxes, they're
        # already drawn into the image pixels, skip the master CSV load
        # and live drawing entirely. Projects sorted before that change
        # fall back to the live path below.
        baked  = (self._val_folder / ".siblings_baked").exists()
        master = self._val_folder.parent / "results_with_ids.csv"
        if master.exists() and not baked:
            try:
                master_mtime = master.stat().st_mtime
            except OSError:
                master_mtime = None
            cache = getattr(self, "_val_sibling_cache", None)
            if cache and cache[0] == master and cache[1] == master_mtime:
                self._val_sibling_bboxes = cache[2]
            elif not getattr(self, "_val_sibling_loading", False):
                self._val_sibling_loading = True

                def _parse_siblings():
                    lookup = {}
                    try:
                        with open(master, newline="", encoding="utf-8") as f:
                            for r in csv.DictReader(f):
                                img_id = r.get("image_id","").strip()
                                if not img_id:
                                    continue
                                try:
                                    bbox = {
                                        "left":   float(r.get("bbox_left",0) or 0),
                                        "top":    float(r.get("bbox_top",0) or 0),
                                        "right":  float(r.get("bbox_right",0) or 0),
                                        "bottom": float(r.get("bbox_bottom",0) or 0),
                                    }
                                    norm = r.get("bbox_normalised","") == "1"
                                    det_id = r.get("detection_id","")
                                    lbl    = r.get("label","")
                                    if img_id not in lookup:
                                        lookup[img_id] = []
                                    lookup[img_id].append(
                                        (det_id, lbl, bbox, norm))
                                except Exception:
                                    pass
                    except Exception:
                        pass
                    # Drop single-detection images entirely, they have no
                    # siblings to draw (the gallery only draws when an
                    # image has more than one detection), and on typical
                    # camera trap data they're the large majority of rows
                    lookup = {k: v for k, v in lookup.items() if len(v) > 1}

                    def _siblings_done():
                        self._val_sibling_loading = False
                        self._val_sibling_cache = (master, master_mtime, lookup)
                        self._val_sibling_bboxes = lookup
                        # Redraw the current batch so sibling boxes appear
                        if lookup:
                            self._val_refresh_gallery()

                    self.after(0, _siblings_done)

                threading.Thread(target=_parse_siblings, daemon=True).start()

        # Find first unvalidated batch
        self._val_batch = 0
        for i, row in enumerate(self._val_rows):
            if row.get("validated","").strip().lower() != "yes":
                self._val_batch = i // self.BATCH_SIZE
                break

        self._val_refresh_gallery()

    def _val_refresh_gallery(self):
        if not self._val_rows:
            return

        # New gallery generation: any decode workers still in flight from
        # the previous species/batch will see the changed gen and drop
        # their results instead of drawing into the rebuilt grid.
        self._val_gallery_gen += 1

        # Only show unvalidated rows
        pending = [r for r in self._val_rows
                   if r.get("validated","").strip().lower() != "yes"]

        # Clear gallery and reset scroll
        for w in self._val_gallery_inner.winfo_children():
            w.destroy()
        self._val_img_refs.clear()
        self._val_tiles = {}
        self._val_canvas.yview_moveto(0)

        cols      = self._val_cols_var.get()
        start     = self._val_batch * self.BATCH_SIZE
        end       = min(start + self.BATCH_SIZE, len(pending))
        batch     = pending[start:end]
        n_batches = max(1, (len(pending) + self.BATCH_SIZE - 1) // self.BATCH_SIZE)
        sp        = self._val_species_var.get()

        # Progress
        n_done = len(self._val_rows) - len(pending)
        self._val_progress_lbl.config(
            text=f"{n_done} of {len(self._val_rows)} validated")
        self._val_batch_lbl.config(
            text=f"Batch {self._val_batch+1} of {n_batches}  "
                 f"({len(pending)} remaining)")
        self._val_prev_btn.config(
            state="normal" if self._val_batch > 0 else "disabled")
        self._val_next_btn.config(
            state="normal" if self._val_batch < n_batches-1 else "disabled")

        # All validated
        if not pending:
            tk.Label(self._val_gallery_inner,
                     text=f"All images in {sp.replace('_',' ').title()} have been validated.",
                     font=self._fonts["h2"], bg=C["frost"],
                     fg=C["forest"], anchor="w").pack(
                         fill="x", padx=16, pady=32)
            return

        # Gallery header
        sp_display = sp.replace("_"," ").title()
        tk.Label(self._val_gallery_inner,
                 text=f"{sp_display}  |  {len(pending)} images remaining",
                 font=self._fonts["h2"], bg=C["frost"],
                 fg=C["canopy"], anchor="w").pack(
                     fill="x", padx=16, pady=(12,4))
        tk.Label(self._val_gallery_inner,
                 text="Click an image to flag it as incorrectly labelled.",
                 font=self._fonts["small"], bg=C["frost"],
                 fg=C["text_muted"], anchor="w").pack(
                     fill="x", padx=16, pady=(0,2))
        tk.Label(self._val_gallery_inner,
                 text="To select multiple images, hold Ctrl and click each "
                      "one (selected tiles turn blue), then click any selected "
                      "tile to correct them all together.",
                 font=self._fonts["small"], bg=C["frost"],
                 fg=C["forest"], anchor="w", justify="left").pack(
                     fill="x", padx=16, pady=(0,10))

        # Grid frame
        grid = tk.Frame(self._val_gallery_inner, bg=C["frost"])
        grid.pack(fill="both", padx=12, pady=4)

        # Calculate tile width. Prefer the canvas width (the actual
        # viewport), which is reliable once the pane is visible; fall back
        # to the inner frame. Retry only if genuinely not yet drawn.
        self._val_gallery_inner.update_idletasks()
        avail_w = self._val_canvas.winfo_width()
        if avail_w < 50:
            avail_w = self._val_gallery_inner.winfo_width()
        if avail_w < 50:
            # Window not yet rendered — schedule a retry
            self.after(150, self._val_refresh_gallery)
            return
        avail_w = max(400, avail_w - 32)

        # Each card consumes tile_w PLUS grid padx (4 each side) and the
        # 2px border frame each side, ~16px of per-column overhead. Compute
        # the width so `cols` columns genuinely fit inside the viewport. If
        # they can't (very narrow window), drop a column rather than let the
        # rightmost tile and its expand button spill off the right edge.
        per_col_overhead = 16
        MIN_TILE = 120
        while cols > 1 and (avail_w // cols) - per_col_overhead < MIN_TILE:
            cols -= 1
        tile_w = (avail_w // cols) - per_col_overhead
        tile_w = max(MIN_TILE, tile_w)

        img_folder = self._val_folder / sp

        # Camera trap images all share one aspect ratio, so size the tile
        # HEIGHT to match the image rather than forcing a square, otherwise
        # a 16:9 photo sits in a square tile with big grey bands above and
        # below. Sample the aspect from the first image in the batch (cheap:
        # PIL reads only the header for .size, no full decode). Fall back to
        # 16:9 if that fails. Every tile uses the same height, so the grid
        # stays perfectly uniform.
        aspect = 9.0 / 16.0  # height/width fallback
        try:
            from PIL import Image as _PilProbe
            if batch:
                probe = img_folder / batch[0].get("image_name","")
                if probe.exists():
                    with _PilProbe.open(probe) as _pi:
                        pw, ph = _pi.size
                        if pw > 0:
                            aspect = ph / pw
        except Exception:
            pass
        tile_h = max(90, int(tile_w * aspect))

        def _build_card(idx, row):
            """Build the tile frame and caption instantly, no image decode.
            Returns the info needed to decode the image later."""
            row_idx = idx // cols
            col_idx = idx  % cols

            det_id    = row.get("detection_id","")
            img_id    = row.get("image_id","")
            img_name  = row.get("image_name","")
            conf      = row.get("confidence","")
            validated = row.get("validated","").strip().lower() == "yes"
            corrected = det_id in self._val_corrections
            selected  = det_id in self._val_selected

            # Card frame
            card = tk.Frame(grid, bg=C["white"], padx=3, pady=3)
            card.grid(row=row_idx, column=col_idx, padx=4, pady=4, sticky="n")

            # Border colour — selected overrides corrected overrides default.
            # Selected tiles get a thick, clearly visible border (a 2px one
            # was nearly invisible behind the image); others get a thin one.
            if selected:
                border_col = "#2471A3"   # blue for multi-select
                bpad = 4
            elif corrected:
                border_col = "#C0392B"   # red for correction
                bpad = 3
            elif validated:
                border_col = C["mist"]
                bpad = 2
            else:
                border_col = C["border"]
                bpad = 2

            outer = tk.Frame(card, bg=border_col, padx=bpad, pady=bpad)
            outer.pack()

            # Placeholder box, sized to the tile (image aspect ratio, not
            # square), filled in when the image decodes. Fixed size keeps
            # layout stable so scrolling works immediately.
            holder = tk.Frame(outer, bg="#ECECEC",
                              width=tile_w, height=tile_h)
            holder.pack()
            holder.pack_propagate(False)
            ph = tk.Label(holder, text="…", font=self._fonts["small"],
                          bg="#ECECEC", fg="#BBBBBB")
            ph.place(relx=0.5, rely=0.5, anchor="center")

            # Confidence / correction caption
            try:
                conf_pct = f"{float(conf)*100:.0f}% confident"
            except (ValueError, TypeError):
                conf_pct = conf
            if selected:
                info_text, info_fg = "Selected", "#2471A3"
            elif corrected:
                info_text = f"→ {self._val_corrections[det_id].replace('_',' ')}"
                info_fg   = "#C0392B"
            else:
                info_text, info_fg = conf_pct, C["text_muted"]

            info_lbl = tk.Label(card, text=info_text,
                     font=self._fonts["tile"],
                     bg=C["white"], fg=info_fg,
                     anchor="center")
            info_lbl.pack(fill="x", pady=(4, 2))

            self._val_tiles[det_id] = {"outer": outer, "info": info_lbl,
                                       "row": row, "holder": holder}
            return {"holder": holder, "ph": ph, "outer": outer, "card": card,
                    "row": row, "det_id": det_id, "img_id": img_id,
                    "img_name": img_name, "decoded": False}

        def _decode_pil(t):
            """Heavy part: read the file and produce a finished PIL image.
            Safe to run on a worker thread (no Tk calls). Returns the PIL
            image, or None on failure/cache-hit-handled-elsewhere."""
            from PIL import Image as PilImg, ImageDraw as PilDraw
            det_id   = t["det_id"]
            img_id   = t["img_id"]
            img_name = t["img_name"]
            img_path = img_folder / img_name
            siblings = self._val_sibling_bboxes.get(img_id, [])

            img = PilImg.open(img_path)
            orig_w, orig_h = img.size
            img.draft("RGB", (tile_w, tile_h))
            img.thumbnail((tile_w, tile_h))
            img = img.convert("RGB")
            tw_, th_ = img.size
            if len(siblings) > 1:
                draw = PilDraw.Draw(img)
                sx = tw_ / orig_w
                sy = th_ / orig_h
                for sib_det, sib_lbl, sib_bbox, sib_norm in siblings:
                    if sib_det == det_id:
                        continue
                    if sib_norm:
                        x0 = int(sib_bbox["left"]*tw_); y0 = int(sib_bbox["top"]*th_)
                        x1 = int(sib_bbox["right"]*tw_); y1 = int(sib_bbox["bottom"]*th_)
                    else:
                        x0 = int(sib_bbox["left"]*sx); y0 = int(sib_bbox["top"]*sy)
                        x1 = int(sib_bbox["right"]*sx); y1 = int(sib_bbox["bottom"]*sy)
                    x0,y0,x1,y1 = pad_bbox(x0, y0, x1, y1, tw_, th_)
                    draw.rectangle([x0,y0,x1,y1], outline="#FFC107", width=2)
            return img

        def _place_decoded(t, pil_img):
            """Light part: runs on the Tk main thread. Turns the finished
            PIL image into a PhotoImage and places it in the tile."""
            from PIL import ImageTk
            from collections import OrderedDict
            det_id   = t["det_id"]
            img_id   = t["img_id"]
            img_name = t["img_name"]
            row      = t["row"]
            outer    = t["outer"]
            card     = t["card"]
            img_path = img_folder / img_name
            try:
                if not t["holder"].winfo_exists():
                    return
            except tk.TclError:
                return

            if not hasattr(self, "_val_thumb_cache"):
                self._val_thumb_cache = OrderedDict()
            cache    = self._val_thumb_cache
            siblings = self._val_sibling_bboxes.get(img_id, [])
            ckey     = (img_name, det_id, tile_w, len(siblings))

            photo = cache.get(ckey)
            if photo is None:
                if pil_img is None:
                    # decode failed
                    try:
                        for ch in t["holder"].winfo_children():
                            ch.destroy()
                        tk.Label(t["holder"], text="Image\nnot found",
                                 font=self._fonts["small"], bg="#ECECEC",
                                 fg=C["text_muted"]).place(
                                     relx=0.5, rely=0.5, anchor="center")
                    except tk.TclError:
                        pass
                    return
                photo = ImageTk.PhotoImage(pil_img)
                cache[ckey] = photo
                while len(cache) > 400:
                    cache.popitem(last=False)
            else:
                cache.move_to_end(ckey)

            self._val_img_refs.append(photo)

            try:
                t["ph"].destroy()
            except Exception:
                pass
            holder = t["holder"]
            for ch in holder.winfo_children():
                try: ch.destroy()
                except Exception: pass
            lbl_img = tk.Label(holder, image=photo, bg=C["white"],
                               cursor="hand2", bd=0)
            lbl_img.place(relx=0.5, rely=0.5, anchor="center")

            exp = tk.Label(holder, text="⛶", font=("Segoe UI", 10),
                           bg="#1A1A1A", fg="#FFFFFF",
                           cursor="hand2", padx=4, pady=1)
            exp.place(relx=1.0, x=-3, y=3, anchor="ne")
            exp.bind("<Button-1>",
                lambda e, p=img_path, i=img_id, d=det_id:
                    self._val_expand_image(p, i, d))

            is_fav = det_id in self._val_favourites
            heart = tk.Label(holder, text=("♥" if is_fav else "♡"),
                             font=("Segoe UI", 11), bg="#1A1A1A",
                             fg=("#FF4D6D" if is_fav else "#FFFFFF"),
                             cursor="hand2", padx=4, pady=0)
            heart.place(relx=1.0, x=-30, y=3, anchor="ne")
            heart.bind("<Button-1>",
                lambda e, d=det_id, hb=heart:
                    self._val_toggle_favourite(d, hb))
            if det_id in self._val_tiles:
                self._val_tiles[det_id]["heart"] = heart

            lbl_img.bind("<Button-1>",
                lambda e, r=row, c=card, o=outer, d=det_id:
                    self._val_click_image(r, c, o, d))
            lbl_img.bind("<Control-Button-1>",
                lambda e, d=det_id: self._val_toggle_select(d))
            if hasattr(self, "_val_on_enter"):
                lbl_img.bind("<Enter>", self._val_on_enter, add="+")
                lbl_img.bind("<Leave>", self._val_on_leave, add="+")

        def _decode_tile(t):
            """Queue one tile for threaded decode. The slow file read + PIL
            work runs on a worker thread; the finished image is handed back
            to the Tk main thread for placement. Multiple slow (network)
            reads thus overlap instead of serialising on the UI thread."""
            if t.get("decoded") or t.get("decoding"):
                return
            try:
                if not t["holder"].winfo_exists():
                    return
            except tk.TclError:
                return

            # Cache hit: place immediately on the main thread, no worker
            from collections import OrderedDict
            if not hasattr(self, "_val_thumb_cache"):
                self._val_thumb_cache = OrderedDict()
            siblings = self._val_sibling_bboxes.get(t["img_id"], [])
            ckey = (t["img_name"], t["det_id"], tile_w, len(siblings))
            if ckey in self._val_thumb_cache:
                t["decoded"] = True
                _place_decoded(t, None)
                return

            t["decoding"] = True
            gen = self._val_gallery_gen

            def _worker():
                try:
                    pil_img = _decode_pil(t)
                except Exception:
                    pil_img = None
                def _finish():
                    # Ignore if the gallery moved on (species/batch changed)
                    if gen != self._val_gallery_gen:
                        return
                    t["decoding"] = False
                    t["decoded"]  = True
                    _place_decoded(t, pil_img)
                self.after(0, _finish)

            self._val_decode_pool.submit(_worker)

        # Build every card frame up front, instant since no image is
        # touched, so the grid, scrollbar and captions all appear right
        # away. Then decode images lazily: whatever's currently scrolled
        # into view first, the rest in the background and on scroll.
        self._val_pending_decode = []
        for i in range(len(batch)):
            self._val_pending_decode.append(_build_card(i, batch[i]))

        if hasattr(self, "_val_on_enter"):
            self._val_gallery_inner.update_idletasks()
            def _rebind(w):
                w.bind("<Enter>", self._val_on_enter, add="+")
                w.bind("<Leave>", self._val_on_leave, add="+")
                for ch in w.winfo_children():
                    _rebind(ch)
            _rebind(self._val_gallery_inner)

        def _decode_visible():
            """Submit tiles currently within (or near) the viewport for
            threaded decode, prioritising them ahead of off-screen tiles."""
            try:
                top = self._val_canvas.canvasy(0)
                bot = top + self._val_canvas.winfo_height()
            except tk.TclError:
                return
            margin = self._val_canvas.winfo_height()
            for t in self._val_pending_decode:
                if t.get("decoded") or t.get("decoding"):
                    continue
                try:
                    if not t["holder"].winfo_exists():
                        continue
                    y = t["holder"].winfo_rooty() - \
                        self._val_gallery_inner.winfo_rooty()
                    h = t["holder"].winfo_height()
                except tk.TclError:
                    continue
                if y + h >= top - margin and y <= bot + margin:
                    _decode_tile(t)

        def _decode_rest():
            """Submit every remaining undecoded tile to the pool. The pool's
            worker limit bounds how many run at once, so this doesn't
            overwhelm a network drive; visible tiles were submitted first
            and so are already ahead in the queue."""
            for t in self._val_pending_decode:
                if not t.get("decoded") and not t.get("decoding"):
                    _decode_tile(t)

        # Show the grid immediately (placeholders). Submit the visible
        # screenful first, then everything else. All decoding happens on
        # worker threads, so slow (network) reads overlap and the UI never
        # blocks, tiles fill in as each read completes.
        self._val_gallery_inner.update_idletasks()
        self._val_decode_visible = _decode_visible
        def _kick():
            _decode_visible()
            self.after(50, _decode_rest)
        self.after(1, _kick)

    def _val_expand_image(self, img_path, img_id, det_id):
        """Open a single validation image large, fitted to the screen,
        with a toggle for an enhanced (contrast/sharpness) view."""
        try:
            from PIL import Image as PilImg, ImageTk, ImageDraw as PilDraw
            from PIL import ImageEnhance, ImageOps, ImageFilter
            img = PilImg.open(img_path)
            orig_w, orig_h = img.size
        except Exception as e:
            messagebox.showerror("Error", f"Could not open image:\n{e}")
            return

        win = tk.Toplevel(self)
        win.title(img_path.name)
        win.configure(bg="#111111")

        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        img.thumbnail((int(sw * 0.92), int(sh * 0.85)))
        base = img.convert("RGB")
        tw_, th_ = base.size

        def _draw_sibs(im):
            """Draw padded sibling boxes onto a copy at display size."""
            siblings = self._val_sibling_bboxes.get(img_id, [])
            if len(siblings) <= 1:
                return im
            draw = PilDraw.Draw(im)
            sx = tw_ / orig_w
            sy = th_ / orig_h
            for sib_det, sib_lbl, sib_bbox, sib_norm in siblings:
                if sib_det == det_id:
                    continue
                if sib_norm:
                    x0 = int(sib_bbox["left"]   * tw_)
                    y0 = int(sib_bbox["top"]    * th_)
                    x1 = int(sib_bbox["right"]  * tw_)
                    y1 = int(sib_bbox["bottom"] * th_)
                else:
                    x0 = int(sib_bbox["left"]   * sx)
                    y0 = int(sib_bbox["top"]    * sy)
                    x1 = int(sib_bbox["right"]  * sx)
                    y1 = int(sib_bbox["bottom"] * sy)
                x0,y0,x1,y1 = pad_bbox(x0, y0, x1, y1, tw_, th_)
                draw.rectangle([x0,y0,x1,y1],
                               outline="#FFC107", width=3)
            return im

        photo_norm = ImageTk.PhotoImage(_draw_sibs(base.copy()))
        win._photos = [photo_norm]  # keep references or Tk drops them

        lbl = tk.Label(win, image=photo_norm, bg="#111111")
        lbl.pack(padx=8, pady=(8,4))

        bottom = tk.Frame(win, bg="#111111")
        bottom.pack(pady=(0,8))

        state = {"mode": 0, "photos": {0: photo_norm}}
        MODES = ("Original", "Brightened", "High contrast", "Detail")

        def _apply_mode(mode):
            # Three distinct ways to make a hard-to-see subject clearer,
            # each targets a different failure case:
            if mode == 1:
                # Brightened: gamma-lift the shadows. Best when the animal
                # itself is dark, night IR shots, shaded undergrowth. Lifts
                # dark tones a lot while barely touching highlights, so it
                # reveals the subject without blowing out the bright bits.
                lut = [int(255 * (v / 255) ** 0.55) for v in range(256)]
                im = base.point(lut * 3)
                im = ImageEnhance.Contrast(im).enhance(1.08)
            elif mode == 2:
                # High contrast: per-channel histogram stretch plus a
                # contrast push. Best for flat, washed-out or hazy images
                # where subject and background tones are close together.
                im = ImageOps.autocontrast(base, cutoff=2)
                im = ImageEnhance.Contrast(im).enhance(1.25)
                im = ImageEnhance.Brightness(im).enhance(0.95)
            else:
                # Detail: half-blend with a fully equalised copy (a rough
                # local-contrast reveal) plus unsharp masking. Best for
                # picking out texture, fur, antlers, markings, when the
                # shape is visible but features are muddy.
                im = PilImg.blend(base, ImageOps.equalize(base), 0.55)
                im = im.filter(ImageFilter.UnsharpMask(radius=2, percent=130))
            return im

        def _cycle_contrast(evt=None):
            mode = (state["mode"] + 1) % len(MODES)
            if mode not in state["photos"]:
                im = base if mode == 0 else _apply_mode(mode)
                state["photos"][mode] = ImageTk.PhotoImage(_draw_sibs(im.copy()))
                win._photos.append(state["photos"][mode])
            lbl.config(image=state["photos"][mode])
            mode_lbl.config(text=MODES[mode])
            state["mode"] = mode

        contrast_btn = tk.Button(bottom, text="Change contrast (E)",
                                 command=_cycle_contrast,
                                 font=self._fonts["label"],
                                 bg="#E8E8E8", fg="#222222",
                                 activebackground="#FFFFFF",
                                 activeforeground="#222222",
                                 relief="raised", bd=2, padx=12, pady=4,
                                 cursor="hand2")
        contrast_btn.pack(side="left", padx=(0,8))

        mode_lbl = tk.Label(bottom, text=MODES[0], width=13, anchor="w",
                            font=self._fonts["small"],
                            bg="#111111", fg="#FFC107")
        mode_lbl.pack(side="left", padx=(0,16))

        if not hasattr(self, "_val_favourites"):
            self._val_favourites = set()
        fav_on = det_id in self._val_favourites
        fav_btn = tk.Button(bottom,
                            text=("♥ Favourited" if fav_on else "♡ Favourite"),
                            font=self._fonts["label"],
                            bg="#E8E8E8",
                            fg=("#C0392B" if fav_on else "#222222"),
                            activebackground="#FFFFFF",
                            relief="raised", bd=2, padx=12, pady=4,
                            cursor="hand2")
        def _fav_click():
            self._val_toggle_favourite(det_id)
            now_on = det_id in self._val_favourites
            fav_btn.config(text=("♥ Favourited" if now_on else "♡ Favourite"),
                           fg=("#C0392B" if now_on else "#222222"))
            # Reflect the change on the gallery tile's heart too
            tile = getattr(self, "_val_tiles", {}).get(det_id)
            if tile and tile.get("heart"):
                try:
                    tile["heart"].config(
                        text=("♥" if now_on else "♡"),
                        fg=("#FF4D6D" if now_on else "#FFFFFF"))
                except tk.TclError:
                    pass
        fav_btn.config(command=_fav_click)
        fav_btn.pack(side="left", padx=(0,16))

        tk.Label(bottom, text="Click the image or press Esc to close",
                 font=self._fonts["small"], bg="#111111",
                 fg="#999999").pack(side="left")

        # Center on screen
        win.update_idletasks()
        x = max(0, (sw - win.winfo_width())  // 2)
        y = max(0, (sh - win.winfo_height()) // 2 - 20)
        win.geometry(f"+{x}+{y}")

        lbl.bind("<Button-1>", lambda e: win.destroy())
        win.bind("<Escape>",   lambda e: win.destroy())
        win.bind("<KeyPress-e>", _cycle_contrast)
        win.bind("<KeyPress-E>", _cycle_contrast)
        win.focus_set()

    def _val_custom_labels_path(self):
        """Project-level custom species list, alongside the per-species
        folders in validation/. One list for the whole project."""
        if not self._val_folder:
            return None
        return self._val_folder / "custom_species.txt"

    def _val_load_custom_labels(self):
        p = self._val_custom_labels_path()
        if not p:
            return []
        if not p.exists():
            # Create it (with a short header) the first time, so it always
            # exists for a project, including older projects made before
            # this feature, and can be edited directly at source.
            try:
                p.write_text(
                    "# wildtag.ai custom species labels\n"
                    "# One label per line. These appear in the correction\n"
                    "# dropdown alongside the model's own classes. Lines\n"
                    "# starting with # are ignored. Edit this file directly,\n"
                    "# or add labels from the Validate correction dialog.\n",
                    encoding="utf-8")
            except Exception:
                pass
            return []
        try:
            lines = p.read_text(encoding="utf-8").splitlines()
            return [l.strip() for l in lines
                    if l.strip() and not l.strip().startswith("#")]
        except Exception:
            return []

    def _val_add_custom_label(self, raw_text):
        """Sanitise a user-typed label to the same convention as model
        classes (lowercase, underscores), add it to custom_species.txt and
        the in-memory list, and return the stored form (or None if invalid
        or already present)."""
        if not raw_text or not raw_text.strip():
            return None
        # Match the model-class convention: lowercase, spaces -> underscores,
        # strip anything that isn't a word char, underscore or hyphen
        import re as _re
        lbl = raw_text.strip().lower().replace(" ", "_")
        lbl = _re.sub(r"[^a-z0-9_\-]", "", lbl)
        if not lbl:
            return None
        if lbl in self._val_all_labels:
            return lbl  # already available, nothing to add
        # Persist to the project custom list
        customs = self._val_load_custom_labels()
        if lbl not in customs:
            customs.append(lbl)
            p = self._val_custom_labels_path()
            if p:
                try:
                    # Preserve any existing comment header; append the new
                    # label rather than rewriting (and losing) the file.
                    if p.exists():
                        existing = p.read_text(encoding="utf-8")
                        if existing and not existing.endswith("\n"):
                            existing += "\n"
                        p.write_text(existing + lbl + "\n", encoding="utf-8")
                    else:
                        p.write_text(lbl + "\n", encoding="utf-8")
                except Exception:
                    pass
        # Add to the live list so it appears immediately
        if lbl not in self._val_all_labels:
            self._val_all_labels.append(lbl)
            self._val_custom_labels = customs
        return lbl

    def _val_click_image(self, row, card, outer, det_id):
        """Open a correction dialog. If images are multi-selected, applies to all."""
        win = tk.Toplevel(self)
        win.title("Correct label")
        win.configure(bg=C["white"])
        win.resizable(False, False)
        win.grab_set()

        # Show multi-select context if applicable
        n_selected = len(self._val_selected)
        if n_selected > 0 and det_id not in self._val_selected:
            # Clicked image not in selection — just correct this one
            targets = [det_id]
            header  = f"Current label: {row.get('label','').replace('_',' ')}"
        elif n_selected > 1 and det_id in self._val_selected:
            targets = list(self._val_selected)
            header  = f"Apply correction to {n_selected} selected images"
        else:
            targets = [det_id]
            header  = f"Current label: {row.get('label','').replace('_',' ')}"

        tk.Label(win, text=header,
                 font=self._fonts["label"], bg=C["white"],
                 fg=C["canopy"]).pack(padx=20, pady=(16,4))
        tk.Label(win, text="Select the correct species:",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(padx=20, pady=(0,6))

        lb_frame = tk.Frame(win, bg=C["white"])
        lb_frame.pack(padx=20, pady=(0,8))
        scrollbar = tk.Scrollbar(lb_frame)
        scrollbar.pack(side="right", fill="y")

        # Display labels in sentence case, store raw internally
        def _to_display(lbl):
            return lbl.replace("_", " ").capitalize()

        display_labels = [_to_display(l) for l in self._val_all_labels]

        lb = tk.Listbox(lb_frame, listvariable=tk.StringVar(
                            value=display_labels),
                        font=self._fonts["small"],
                        bg=C["frost"], fg=C["canopy"],
                        selectbackground=C["forest"],
                        selectforeground=C["white"],
                        relief="flat", width=36, height=12,
                        yscrollcommand=scrollbar.set)
        lb.pack(side="left")
        scrollbar.config(command=lb.yview)

        # Mouse-wheel / trackpad scrolling for the listbox. A tk Listbox
        # doesn't scroll on the wheel by default on Windows, so bind it,
        # accumulating small trackpad deltas like the main gallery does.
        _lb_accum = {"v": 0.0}
        def _lb_wheel(e):
            if getattr(e, "num", None) == 4:
                lb.yview_scroll(-1, "units")
            elif getattr(e, "num", None) == 5:
                lb.yview_scroll(1, "units")
            else:
                _lb_accum["v"] += -e.delta * (1.0 / 120.0)
                u = int(_lb_accum["v"])
                _lb_accum["v"] -= u
                if u:
                    lb.yview_scroll(u, "units")
            return "break"
        lb.bind("<MouseWheel>", _lb_wheel)   # Windows / macOS
        lb.bind("<Button-4>", _lb_wheel)     # Linux up
        lb.bind("<Button-5>", _lb_wheel)     # Linux down
        # Also scroll when the pointer is anywhere over the dialog
        win.bind("<MouseWheel>", _lb_wheel)

        # Pre-select current label
        current = row.get("label","")
        if current in self._val_all_labels:
            idx = self._val_all_labels.index(current)
            lb.selection_set(idx)
            lb.see(idx)

        # Add-new-label row: lets a validator create a label the model
        # doesn't have (subspecies, coarse groupings, local names). Added
        # to the project custom list and selected immediately.
        add_frame = tk.Frame(win, bg=C["white"])
        add_frame.pack(padx=20, pady=(0,8), fill="x")
        tk.Label(add_frame, text="Not listed? Add a label:",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"]).pack(anchor="w")
        entry_row = tk.Frame(add_frame, bg=C["white"])
        entry_row.pack(fill="x", pady=(2,0))
        new_lbl_var = tk.StringVar()
        new_entry = tk.Entry(entry_row, textvariable=new_lbl_var,
                             font=self._fonts["small"], width=26,
                             relief="solid", bd=1)
        new_entry.pack(side="left")

        def _add_label():
            stored = self._val_add_custom_label(new_lbl_var.get())
            if not stored:
                return
            # Rebuild the listbox contents and select the new/matched label
            display_labels = [_to_display(l) for l in self._val_all_labels]
            lb.delete(0, "end")
            for d in display_labels:
                lb.insert("end", d)
            idx = self._val_all_labels.index(stored)
            lb.selection_clear(0, "end")
            lb.selection_set(idx)
            lb.see(idx)
            new_lbl_var.set("")

        tk.Button(entry_row, text="Add",
                  command=_add_label,
                  font=self._fonts["small"],
                  bg=C["mist"], fg=C["canopy"],
                  activebackground=C["leaf"], activeforeground=C["white"],
                  relief="flat", padx=10, pady=2,
                  cursor="hand2").pack(side="left", padx=(6,0))
        new_entry.bind("<Return>", lambda e: _add_label())

        btn_row = tk.Frame(win, bg=C["white"])
        btn_row.pack(padx=20, pady=(0,16))

        def _confirm():
            sel = lb.curselection()
            if not sel:
                return
            new_label = self._val_all_labels[sel[0]]  # raw label, not display
            for tid in targets:
                if new_label != self._get_row_label(tid):
                    self._val_corrections[tid] = new_label
            # Update affected tiles in place, both the correction targets
            # and anything that was multi-selected (deselection changes
            # its border back), instead of rebuilding the whole gallery
            affected = set(targets) | set(self._val_selected)
            self._val_selected.clear()
            win.destroy()
            for tid in affected:
                self._val_update_tile(tid)

        def _remove():
            for tid in targets:
                self._val_corrections.pop(tid, None)
            affected = set(targets) | set(self._val_selected)
            self._val_selected.clear()
            win.destroy()
            for tid in affected:
                self._val_update_tile(tid)

        tk.Button(btn_row, text="Confirm correction",
                  command=_confirm,
                  font=self._fonts["label"],
                  bg=C["forest"], fg=C["white"],
                  activebackground=C["leaf"],
                  relief="flat", padx=12, pady=6,
                  cursor="hand2").pack(side="left", padx=(0,8))
        tk.Button(btn_row, text="Remove flag",
                  command=_remove,
                  font=self._fonts["small"],
                  bg=C["border"], fg=C["canopy"],
                  relief="flat", padx=10, pady=6,
                  cursor="hand2").pack(side="left")

    def _get_row_label(self, det_id):
        """Return the original label for a detection id."""
        for r in self._val_rows:
            if r.get("detection_id","") == det_id:
                return r.get("label","")
        return ""

    def _val_update_tile(self, det_id):
        """Refresh one tile's border colour and caption in place, without
        rebuilding the gallery (which re-decodes every image)."""
        tile = getattr(self, "_val_tiles", {}).get(det_id)
        if not tile:
            return
        outer, info_lbl, row = tile["outer"], tile["info"], tile["row"]
        try:
            if not outer.winfo_exists():
                return
        except tk.TclError:
            return

        validated = row.get("validated","").strip().lower() == "yes"
        corrected = det_id in self._val_corrections
        selected  = det_id in self._val_selected

        if selected:
            border_col = "#2471A3"; bpad = 4
        elif corrected:
            border_col = "#C0392B"; bpad = 3
        elif validated:
            border_col = C["mist"];  bpad = 2
        else:
            border_col = C["border"]; bpad = 2
        outer.config(bg=border_col, padx=bpad, pady=bpad)

        # A clear "selected" badge in the tile's top-left corner, so
        # multi-select is obvious at a glance rather than relying on a
        # thin border. Created on demand, removed when deselected.
        holder = tile.get("holder")
        badge  = tile.get("sel_badge")
        if selected and holder is not None:
            if badge is None or not badge.winfo_exists():
                try:
                    badge = tk.Label(holder, text="✓", font=("Segoe UI", 11, "bold"),
                                     bg="#2471A3", fg="#FFFFFF", padx=5, pady=0)
                    badge.place(relx=0.0, x=3, y=3, anchor="nw")
                    tile["sel_badge"] = badge
                except tk.TclError:
                    pass
        elif badge is not None:
            try: badge.destroy()
            except tk.TclError: pass
            tile["sel_badge"] = None

        conf = row.get("confidence","")
        try:
            conf_pct = f"{float(conf)*100:.0f}% confident"
        except (ValueError, TypeError):
            conf_pct = conf

        if selected:
            info_lbl.config(text="Selected", fg="#2471A3")
        elif corrected:
            correct = self._val_corrections[det_id]
            info_lbl.config(text=f"→ {correct.replace('_',' ')}",
                            fg="#C0392B")
        else:
            info_lbl.config(text=conf_pct, fg=C["text_muted"])

    def _val_favourites_path(self):
        """favourites.json lives in the project's validation folder so it
        travels with the project and is scoped to it, not global."""
        if not self._val_folder:
            return None
        return self._val_folder / "favourites.json"

    def _val_load_favourites(self):
        if not hasattr(self, "_val_favourites"):
            self._val_favourites = set()
        p = self._val_favourites_path()
        if p and p.exists():
            try:
                data = json.loads(p.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    self._val_favourites = set(data)
            except Exception:
                pass

    def _val_save_favourites(self):
        p = self._val_favourites_path()
        if not p:
            return
        try:
            p.write_text(json.dumps(sorted(self._val_favourites)),
                         encoding="utf-8")
        except Exception:
            pass

    def _val_toggle_favourite(self, det_id, heart_widget=None):
        if not hasattr(self, "_val_favourites"):
            self._val_favourites = set()
        if det_id in self._val_favourites:
            self._val_favourites.discard(det_id)
            on = False
        else:
            self._val_favourites.add(det_id)
            on = True
        self._val_save_favourites()
        if heart_widget is not None:
            try:
                heart_widget.config(text=("♥" if on else "♡"),
                                    fg=("#FF4D6D" if on else "#FFFFFF"))
            except tk.TclError:
                pass

    def _val_export_favourites(self):
        """Copy every favourited image into a favourites\\ folder (alongside
        validation\\, distribute\\, etc.), organised into per-species
        subfolders. Triggered from the Distribute pane."""
        # Resolve the validation folder without relying on the Validate pane
        # having been opened this session.
        val_folder = getattr(self, "_val_folder", None)
        if not val_folder or not val_folder.exists():
            proj = self._find_project_root()
            if proj:
                cand = proj / "validation"
                if cand.exists():
                    val_folder = cand
        if not val_folder or not val_folder.exists():
            messagebox.showinfo("No project",
                "Load a processed project first (Run tab).")
            return

        # Load favourites from this project's favourites.json
        favs = set()
        fav_json = val_folder / "favourites.json"
        if fav_json.exists():
            try:
                data = json.loads(fav_json.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    favs = set(data)
            except Exception:
                pass
        # Merge any in-memory favourites from the current session too
        favs |= getattr(self, "_val_favourites", set())

        if not favs:
            messagebox.showinfo(
                "No favourites",
                "No favourite images have been marked yet. In the Validate "
                "pane, click the heart in the corner of an image (or the "
                "Favourite button in full-screen view) to mark it.")
            return

        import shutil
        dest_root = val_folder.parent / "favourites"
        dest_root.mkdir(exist_ok=True)

        # Favourites are keyed by detection_id; the sorted image lives at
        # validation/{species}/{detection_id}.jpg. Copy each into a matching
        # per-species subfolder under favourites/.
        remaining = set(favs)
        copied = 0
        species_hit = set()
        for sp_dir in val_folder.iterdir():
            if not sp_dir.is_dir():
                continue
            for det_id in list(remaining):
                src = sp_dir / f"{det_id}.jpg"
                if src.exists():
                    sp_dest = dest_root / sp_dir.name
                    sp_dest.mkdir(exist_ok=True)
                    try:
                        shutil.copy2(src, sp_dest / f"{det_id}.jpg")
                        copied += 1
                        species_hit.add(sp_dir.name)
                        remaining.discard(det_id)
                    except Exception:
                        pass

        messagebox.showinfo(
            "Favourites exported",
            f"Copied {copied} favourite image{'s' if copied != 1 else ''} "
            f"across {len(species_hit)} species folder"
            f"{'s' if len(species_hit) != 1 else ''} to:\n{dest_root}")

    def _find_project_root(self):
        """Best-effort project root (the folder containing validation\\)."""
        if getattr(self, "_last_output_dir", None):
            p = Path(self._last_output_dir).parent
            if p.exists():
                return p
        v = getattr(self, "_img_folder_var", None)
        if v and v.get().strip():
            p = Path(v.get().strip())
            if p.exists():
                return p
        s = self._settings.get("project", "").strip()
        if s and Path(s).exists():
            return Path(s)
        return None

    def _val_toggle_select(self, det_id):
        """Toggle multi-select for a tile. Ctrl+click."""
        if det_id in self._val_selected:
            self._val_selected.discard(det_id)
        else:
            self._val_selected.add(det_id)
        self._val_update_tile(det_id)

    def _val_prev_batch(self):
        if self._val_batch > 0:
            self._val_batch -= 1
            self._val_refresh_gallery()

    def _val_next_batch(self):
        n_batches = (len(self._val_rows) + self.BATCH_SIZE - 1) // self.BATCH_SIZE
        if self._val_batch < n_batches - 1:
            self._val_batch += 1
            self._val_refresh_gallery()

    def _val_merge_to_master(self):
        """
        Silently merge all validation CSVs back into results_with_ids.csv.
        Uses detection_id as the join key. Runs in a background thread.
        """
        def _merge():
            try:
                if not self._val_folder:
                    return
                master_path = self._val_folder.parent / "results_with_ids.csv"
                if not master_path.exists():
                    return

                # Read master
                with open(master_path, newline="", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    master_fields = list(reader.fieldnames)
                    master_rows   = list(reader)

                # Add columns if not present
                for col in ("correct_label", "validated"):
                    if col not in master_fields:
                        master_fields.append(col)
                        for r in master_rows:
                            r.setdefault(col, "")

                # Build lookup from all validation CSVs
                corrections = {}  # detection_id -> {correct_label, validated}
                for sp_dir in self._val_folder.iterdir():
                    val_csv = sp_dir / "validation.csv"
                    if not sp_dir.is_dir() or not val_csv.exists():
                        continue
                    with open(val_csv, newline="", encoding="utf-8-sig") as f:
                        for row in csv.DictReader(f):
                            det_id = row.get("detection_id","").strip()
                            if det_id:
                                corrections[det_id] = {
                                    "correct_label": row.get("correct_label",""),
                                    "validated":     row.get("validated",""),
                                }

                # Apply to master
                for row in master_rows:
                    det_id = row.get("detection_id","").strip()
                    if det_id in corrections:
                        row["correct_label"] = corrections[det_id]["correct_label"]
                        row["validated"]     = corrections[det_id]["validated"]

                # Write back
                with open(master_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=master_fields)
                    writer.writeheader()
                    writer.writerows(master_rows)

            except Exception as e:
                print(f"Merge error: {e}")

        threading.Thread(target=_merge, daemon=True).start()

    def _val_complete_batch(self):
        if not self._val_csv_path or not self._val_rows:
            return

        pending = [r for r in self._val_rows
                   if r.get("validated","").strip().lower() != "yes"]
        start   = self._val_batch * self.BATCH_SIZE
        end     = min(start + self.BATCH_SIZE, len(pending))
        batch   = pending[start:end]

        if not batch:
            return

        confirmed = messagebox.askyesno(
            "Mark batch complete",
            f"By clicking Yes you are confirming that you have checked all "
            f"{len(batch)} images in this batch and corrected any incorrect "
            f"labels.\n\nThis will mark them as validated in the spreadsheet. "
            f"Continue?")
        if not confirmed:
            return

        # Apply corrections and mark validated
        for row in batch:
            det_id = row.get("detection_id","")
            if det_id in self._val_corrections:
                row["correct_label"] = self._val_corrections[det_id]
            row["validated"] = "yes"

        # Write updated validation.csv
        fields = list(self._val_rows[0].keys())
        with open(self._val_csv_path, "w", newline="",
                  encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(self._val_rows)

        # Clear corrections and selection for completed rows
        for row in batch:
            self._val_corrections.pop(row.get("detection_id",""), None)
        self._val_selected.clear()

        # Merge to master CSV
        self._val_merge_to_master()

        # Check if this species is now fully validated
        still_pending = [r for r in self._val_rows
                         if r.get("validated","").strip().lower() != "yes"]

        if not still_pending:
            # Species fully done — auto-advance to next unvalidated species
            messagebox.showinfo(
                "Batch complete",
                f"{len(batch)} images marked as validated.\n\n"
                f"All images in this species have been validated!")

            # Find next species with pending images
            current_sp = self._val_species_var.get()
            advanced   = False
            for sp in self._val_species:
                if sp == current_sp:
                    continue
                sp_csv = self._val_folder / sp / "validation.csv"
                if not sp_csv.exists():
                    continue
                with open(sp_csv, newline="", encoding="utf-8") as f:
                    rows = list(csv.DictReader(f))
                if any(r.get("validated","").strip().lower() != "yes"
                       for r in rows):
                    self._val_species_var.set(sp)
                    advanced = True
                    break

            if not advanced:
                messagebox.showinfo(
                    "All validated",
                    "All species folders have been fully validated!")
                self._val_refresh_gallery()
        else:
            # More images remain — reset to first batch and scroll to top
            self._val_batch = 0
            self._val_refresh_gallery()
            # Reset scroll to top
            self._val_canvas.yview_moveto(0)

            messagebox.showinfo(
                "Batch complete",
                f"{len(batch)} images marked as validated and saved.\n"
                f"{len(still_pending)} images remaining.")

    # ── PANE: DISTRIBUTE ──────────────────────────────────────────────────────

    def _build_pane_distribute(self):
        pane  = tk.Frame(self._pane_area, bg=C["frost"])
        self._panes["distribute"] = pane
        inner = self._scrollable(pane)
        pad   = dict(padx=24, pady=6)

        tk.Frame(inner, bg=C["frost"], height=8).pack()
        tk.Label(inner, text="Distribute", font=self._fonts["h2"],
                 bg=C["frost"], fg=C["canopy"], anchor="w").pack(
                     fill="x", padx=24, pady=(0,2))
        self._wrap_label(inner,
            "Prepare species folders as zip files to send to validators, "
            "and import validated zips when they are returned.",
            bg=C["frost"]).pack(fill="x", padx=24)
        tk.Frame(inner, bg=C["frost"], height=12).pack()

        # ── Part A: Prepare packages ─────────────────────────────────────
        self._section_label(inner, "Part A - Prepare packages for validators").pack(
            fill="x", padx=24, pady=(0,4))
        oA, cA = self._card(inner); oA.pack(fill="x", **pad)

        self._wrap_label(cA,
            "wildtag creates one zip per species folder and saves them to "
            "project\\distribute\\. Each zip contains the images, validation "
            "spreadsheet, species reference list, the wildtag app, and a README. "
            "Send the zip to your validator. When done they zip the folder back "
            "up and return it to you.",
            bg=C["white"]).pack(fill="x", pady=(0,12))

        self._dist_folder_var = tk.StringVar()
        self._dist_out_var    = tk.StringVar()

        batch_row = tk.Frame(cA, bg=C["white"])
        batch_row.pack(anchor="w", pady=(0,4))
        tk.Label(batch_row, text="Max images per zip:",
                 font=self._fonts["label"], bg=C["white"],
                 fg=C["canopy"]).pack(side="left", padx=(0,8))
        self._dist_batch_var = tk.IntVar(
            value=int(self._settings.get("dist_batch", 250)))
        tk.Spinbox(batch_row, from_=25, to=100000, increment=25, width=8,
                   textvariable=self._dist_batch_var,
                   font=self._fonts["label"],
                   relief="solid", bd=1).pack(side="left")
        self._wrap_label(cA,
            "Smaller zips (the default 250) are easy to email or host "
            "online for remote volunteers. Larger zips suit handing over "
            "a hard drive to someone working through images in bulk.",
            bg=C["white"]).pack(fill="x", pady=(0,10))

        tk.Button(cA, text="Prepare zip packages",
                  command=self._dist_prepare,
                  font=("Segoe UI", 11, "bold"),
                  bg=C["forest"], fg=C["white"],
                  activebackground=C["leaf"], activeforeground=C["white"],
                  relief="flat", padx=16, pady=8,
                  cursor="hand2").pack(anchor="w", pady=(4,0))

        self._dist_prep_log = tk.Label(cA, text="",
                                       font=self._fonts["small"],
                                       bg=C["white"], fg=C["text_muted"],
                                       anchor="w", justify="left")
        self._dist_prep_log.pack(fill="x", pady=(6,0))

        # ── Part B: Collect validated zips ───────────────────────────────
        self._section_label(inner, "Part B - Collect returned validated zips").pack(
            fill="x", padx=24, pady=(20,4))
        oB, cB = self._card(inner); oB.pack(fill="x", **pad)

        self._wrap_label(cB,
            "When validators return their zips, drop them into the collect\\ "
            "folder inside your project folder. Then click the button below "
            "to merge all returned validations into your master files at once. "
            "Processed zips are moved to collect\\processed\\ automatically.",
            bg=C["white"]).pack(fill="x", pady=(0,12))

        tk.Button(cB, text="Merge all returned validations",
                  command=self._dist_collect,
                  font=("Segoe UI", 11, "bold"),
                  bg=C["forest"], fg=C["white"],
                  activebackground=C["leaf"], activeforeground=C["white"],
                  relief="flat", padx=16, pady=8,
                  cursor="hand2").pack(anchor="w")

        self._dist_import_log = tk.Label(cB, text="",
                                         font=self._fonts["small"],
                                         bg=C["white"], fg=C["text_muted"],
                                         anchor="w", justify="left")
        self._dist_import_log.pack(fill="x", pady=(6,0))

        # ── Part C: Export Camtrap DP ─────────────────────────────────────
        self._section_label(inner, "Part C - Export as Camtrap DP").pack(
            fill="x", padx=24, pady=(20,4))
        oC, cC = self._card(inner); oC.pack(fill="x", **pad)

        self._wrap_label(cC,
            "Export your project as a Camtrap DP data package - the TDWG open "
            "standard for camera trap data exchange. Requires a deployment.csv "
            "file in your project folder with site names, coordinates and dates. "
            "Produces deployments.csv, media.csv, observations.csv and "
            "datapackage.json ready for submission to GBIF, Zenodo or camtrapR.",
            bg=C["white"]).pack(fill="x", pady=(0,8))

        btn_row_c = tk.Frame(cC, bg=C["white"])
        btn_row_c.pack(anchor="w", pady=(0,8))

        tk.Button(btn_row_c, text="Export Camtrap DP",
                  command=self._dist_camtrapdp,
                  font=("Segoe UI", 11, "bold"),
                  bg=C["forest"], fg=C["white"],
                  activebackground=C["leaf"], activeforeground=C["white"],
                  relief="flat", padx=16, pady=8,
                  cursor="hand2").pack(side="left", padx=(0,8))

        tk.Button(btn_row_c, text="Save deployment template",
                  command=self._dist_save_deployment_template,
                  font=self._fonts["label"],
                  bg=C["border"], fg=C["canopy"],
                  relief="flat", padx=10, pady=8,
                  cursor="hand2").pack(side="left")

        self._dist_camtrapdp_log = tk.Label(cC, text="",
                                            font=self._fonts["small"],
                                            bg=C["white"], fg=C["text_muted"],
                                            anchor="w", justify="left")
        self._dist_camtrapdp_log.pack(fill="x", pady=(6,0))

        # ── Part D: Export favourites ────────────────────────────────────
        self._section_label(inner, "Part D - Export favourite images").pack(
            fill="x", padx=24, pady=(20,4))
        oD, cD = self._card(inner); oD.pack(fill="x", **pad)

        self._wrap_label(cD,
            "Collect the images you hearted during validation into a "
            "favourites\\ folder (alongside validation\\ and distribute\\), "
            "organised into per-species subfolders. Handy for pulling out "
            "your best shots for reports, presentations or sharing.",
            bg=C["white"]).pack(fill="x", pady=(0,8))

        tk.Button(cD, text="♥ Export favourites",
                  command=self._val_export_favourites,
                  font=("Segoe UI", 11, "bold"),
                  bg=C["forest"], fg=C["white"],
                  activebackground=C["leaf"], activeforeground=C["white"],
                  relief="flat", padx=16, pady=8,
                  cursor="hand2").pack(anchor="w")

    # ── DISTRIBUTE: actions ───────────────────────────────────────────────────

    def _dist_camtrapdp(self):
        """Export project as Camtrap DP package."""
        # Ask for project folder
        project_str = filedialog.askdirectory(
            title="Select project folder")
        if not project_str:
            return
        project = Path(project_str)

        # Check deployment file exists
        dep_file = None
        for name in ["deployment.csv", "deployment.xlsx",
                     "deployments.csv", "deployments.xlsx"]:
            if (project / name).exists():
                dep_file = project / name
                break

        if not dep_file:
            if messagebox.askyesno(
                "No deployment file",
                f"No deployment.csv found in:\n{project}\n\n"
                f"Would you like to save a template there now?\n"
                f"Fill it in then click Export Camtrap DP again."):
                self._dist_save_deployment_template(project)
            return

        self._dist_camtrapdp_log.config(
            text="Exporting...", fg=C["forest"])
        self.update_idletasks()

        def _export():
            try:
                from wt_models.camtrapdp import export
                log_lines = []
                out = export(
                    project_dir = project,
                    log         = lambda msg: log_lines.append(msg))
                summary = f"Export complete.\nOutput: {out}"
                self.after(0, lambda: self._dist_camtrapdp_log.config(
                    text=summary, fg=C["forest"]))
                self.after(0, lambda: messagebox.showinfo(
                    "Camtrap DP exported",
                    f"Package written to:\n{out}\n\n"
                    f"Contains: deployments.csv, media.csv, "
                    f"observations.csv, datapackage.json"))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: self._dist_camtrapdp_log.config(
                    text=f"Error: {err}", fg=C["warn"]))
                self.after(0, lambda: messagebox.showerror(
                    "Export failed", err))

        threading.Thread(target=_export, daemon=True).start()

    def _dist_save_deployment_template(self, dest_dir=None):
        """Save a deployment.csv template to the project folder."""
        if dest_dir is None:
            dest_dir = filedialog.askdirectory(
                title="Select project folder to save template")
            if not dest_dir:
                return
            dest_dir = Path(dest_dir)

        template = dest_dir / "deployment.csv"
        if template.exists():
            if not messagebox.askyesno(
                "File exists",
                f"deployment.csv already exists in:\n{dest_dir}\n\nOverwrite?"):
                return

        fields = [
            "locationName", "latitude", "longitude",
            "deploymentStart", "deploymentEnd",
            "cameraModel", "habitat", "setupBy", "deploymentComments"
        ]
        example = {
            "locationName":    "Site A",
            "latitude":        "51.5074",
            "longitude":       "-0.1278",
            "deploymentStart": "2024-11-01",
            "deploymentEnd":   "2024-11-30",
            "cameraModel":     "Browning-StrikeForce",
            "habitat":         "Mixed deciduous woodland",
            "setupBy":         "",
            "deploymentComments": "",
        }
        with open(template, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            w.writerow(example)

        messagebox.showinfo(
            "Template saved",
            f"deployment.csv saved to:\n{template}\n\n"
            f"Fill in your site details - one row per camera trap deployment.\n"
            f"Then click Export Camtrap DP.")

    def _dist_pick_folder(self):
        p = filedialog.askdirectory(title="Select project folder")
        if p:
            picked = Path(p)
            # Auto-detect validation subfolder
            if (picked / "validation").exists():
                p = str(picked / "validation")
            self._dist_folder_var.set(p)
            if not self._dist_out_var.get():
                self._dist_out_var.set(p)

    def _make_readme(self, species_name: str) -> str:
        sp_display = species_name.replace("_", " ").title()
        return f"""wildtag.ai - Validation package
================================

Species folder: {sp_display}

WHAT TO DO
----------

1. Unzip this file into a folder on your computer.

2. Inside the folder you will find:
   - Run wildtag.bat   (double-click this to start)
   - wildtag.py        (the wildtag app, do not open this directly)
   - validate_env/     (a private copy of Python, already set up for you)
   - {species_name}/   (the species folder to validate)

3. Double-click Run wildtag.bat to open wildtag.
   A black window will appear briefly, that is normal, wildtag itself
   will open in its own window a few seconds later.
   You do not need Python installed, everything needed is already
   included in this folder.

4. Click the Validate tab in the left panel.

5. Click Browse next to Validation folder and select the
   {species_name} folder inside this unzipped folder.

6. The gallery will show images that need reviewing.
   - If a label is CORRECT, do nothing.
   - If a label is WRONG, click the image and select the correct species.

7. When you have checked all images in a batch, click
   Mark batch complete and confirm.

8. When you have finished all batches, close wildtag.

SENDING BACK
------------

1. Find the {species_name} folder (not the whole unzipped folder).
2. Right-click it and choose Send to > Compressed (zipped) folder.
3. Name the zip:  {species_name}_validated.zip
4. Send the zip back to the project lead.

NOTES
-----

- You only need to flag images that are WRONG.
- If you genuinely cannot identify an animal, select unidentifiable.
- Do not rename, move, or delete any image files.
- Do not edit the validation.csv directly in Excel.

Thank you for your help.
"""

    def _dist_prepare(self):
        import zipfile, math, io

        proj_str = self._img_folder_var.get().strip()
        if not proj_str:
            messagebox.showerror("No project",
                "Please select a project folder in the Run tab first.")
            return

        project  = Path(proj_str)
        val_path = project / "validation"
        out_path = project / "distribute"

        if not val_path.exists():
            messagebox.showerror("No validation folder",
                f"No validation folder found in:\n{project}\n\n"
                "Please run wildtag first.")
            return

        out_path.mkdir(parents=True, exist_ok=True)
        (project / "collect").mkdir(parents=True, exist_ok=True)

        species_dirs = [
            d for d in sorted(val_path.iterdir())
            if d.is_dir() and (d / "validation.csv").exists()
        ]
        if not species_dirs:
            messagebox.showerror("Nothing to package",
                "No species folders with validation.csv found.")
            return

        script_dir   = Path(__file__).parent
        wildtag_py   = script_dir / "wildtag.py"
        wildtag_ico  = script_dir / "wildtag.ico"
        validate_env = script_dir / "validate_env"

        has_env = validate_env.exists()
        if not has_env:
            if not messagebox.askyesno(
                "No validate_env found",
                "validate_env\\ not found — volunteer zips will not include "
                "a Python environment.\n\nRun setup_validate_env.bat to build it.\n\n"
                "Continue without it?"):
                return

        try:
            BATCH = int(self._dist_batch_var.get())
        except Exception:
            BATCH = 250
        BATCH = max(1, min(BATCH, 100000))
        if self._settings.get("dist_batch") != BATCH:
            self._settings["dist_batch"] = BATCH
            save_settings(self._settings)
        created = []
        errors  = []

        # Never send the same image twice. A ledger of image_names packaged in
        # previous runs is kept in distribute/, and images already validated
        # locally are skipped too. Delete distributed_images.txt to allow a
        # full re-send.
        sent_log = out_path / "distributed_images.txt"
        already_sent = set()
        if sent_log.exists():
            already_sent = {ln.strip() for ln in
                            sent_log.read_text(encoding="utf-8").splitlines()
                            if ln.strip()}
        newly_sent   = []
        skipped_done = 0

        # Sibling-bbox drawing (other detections in the same source image,
        # shown as thin muted boxes behind the main detection). If this
        # project was sorted with baked-in sibling boxes, the boxes are
        # already in the image pixels, nothing extra to ship. Projects
        # sorted before that change drew siblings live from
        # results_with_ids.csv, so for those, load it once here and write
        # a filtered subset into each zip, just the rows for images that
        # are actually included in that particular batch, so cross-species
        # siblings still resolve without leaking every other image's data
        # to a volunteer who only received one species' folder.
        siblings_baked = (project / "validation" / ".siblings_baked").exists()
        master_csv  = project / "results_with_ids.csv"
        master_rows = []
        if master_csv.exists() and not siblings_baked:
            try:
                master_rows, _ = load_csv(master_csv)
            except Exception:
                master_rows = []

        for sp_dir in species_dirs:
            sp   = sp_dir.name
            imgs = sorted([f for f in sp_dir.iterdir()
                           if f.suffix.lower() in (".jpg",".jpeg",".png")])
            meta = [f for f in sp_dir.iterdir()
                    if f.suffix.lower() not in (".jpg",".jpeg",".png")]
            n_batches = max(1, math.ceil(len(imgs) / BATCH))

            # detection_id -> image_id for this species, from its own
            # validation.csv, so batch filenames (named by detection_id)
            # can be resolved to the image_id sibling lookups need
            det_to_img = {}
            sp_rows, sp_fields = [], []
            sp_csv = sp_dir / "validation.csv"
            if sp_csv.exists():
                try:
                    sp_rows, sp_fields = load_csv(sp_csv)
                    det_to_img = {r.get("detection_id",""): r.get("image_id","")
                                  for r in sp_rows}
                except Exception:
                    det_to_img = {}
            # image_name -> row, so each batch can ship only its own manifest
            # rows. Copying the whole validation.csv into every batch is what
            # left volunteers with rows for images not in their zip ("Image
            # not found"). Fall back to shipping the raw file if we couldn't
            # parse it.
            rows_by_name = {r.get("image_name",""): r for r in sp_rows}
            if sp_fields:
                other_meta = [f for f in meta if f.name != "validation.csv"]
            else:
                other_meta = meta

            # Only package images that still need validation and have not been
            # sent before, so no volunteer ever receives the same image twice.
            validated_names = {r.get("image_name","") for r in sp_rows
                               if r.get("validated","").strip().lower() == "yes"}
            before_n = len(imgs)
            imgs = [f for f in imgs
                    if f.name not in already_sent
                    and f.name not in validated_names]
            skipped_done += before_n - len(imgs)
            if not imgs:
                continue
            n_batches = max(1, math.ceil(len(imgs) / BATCH))

            for idx in range(n_batches):
                batch = imgs[idx*BATCH:(idx+1)*BATCH]
                sfx   = f"_{idx+1:03d}" if n_batches > 1 else ""
                zname = f"{sp}{sfx}_validation.zip"
                zpath = out_path / zname
                try:
                    # Open with STORED (no compression) as the default.
                    # Camera-trap images are JPEGs, already compressed, so
                    # running DEFLATE over them costs a lot of CPU for almost
                    # no size saving, this is the main reason zipping 1000+
                    # images was slow. We store images as-is (near-instant)
                    # and explicitly DEFLATE only the genuinely compressible
                    # files (text, CSV, Python, the validate_env) via the
                    # compress_type argument per write.
                    DEF = zipfile.ZIP_DEFLATED
                    with zipfile.ZipFile(zpath, "w", zipfile.ZIP_STORED) as zf:
                        if wildtag_py.exists():
                            zf.write(wildtag_py,  "wildtag.py", compress_type=DEF)
                        if wildtag_ico.exists():
                            zf.write(wildtag_ico, "wildtag.ico")

                        # Include wt_models (Python files only, no weights)
                        wt_models_dir = script_dir / "wt_models"
                        if wt_models_dir.exists():
                            for f in wt_models_dir.rglob("*.py"):
                                rel = f.relative_to(script_dir)
                                zf.write(f, str(rel).replace("\\", "/"),
                                         compress_type=DEF)
                        bat = ("@echo off\r\n"
                               "cd /d \"%~dp0\"\r\n"
                               "set TCL_LIBRARY=%~dp0validate_env\\tcl\\tcl8.6\r\n"
                               "set TK_LIBRARY=%~dp0validate_env\\tcl\\tk8.6\r\n"
                               "validate_env\\python.exe wildtag.py\r\n"
                               "pause\r\n")
                        zf.writestr("Run wildtag.bat", bat)
                        zf.writestr(f"README_{sp}.txt", self._make_readme(sp))
                        if siblings_baked:
                            zf.writestr("validation/.siblings_baked", "1")
                        # Ship the project-level custom species list so
                        # volunteers see the same user-defined labels. Ensure
                        # it exists first (it may not if Validate was never
                        # opened this session), so it's always in the zip.
                        custom_txt = project / "validation" / "custom_species.txt"
                        if not custom_txt.exists():
                            try:
                                custom_txt.write_text(
                                    "# wildtag.ai custom species labels\n"
                                    "# One label per line. Lines starting with "
                                    "# are ignored.\n",
                                    encoding="utf-8")
                            except Exception:
                                pass
                        if custom_txt.exists():
                            zf.write(custom_txt, "validation/custom_species.txt",
                                     compress_type=DEF)
                        # Images: STORED (default) — already compressed, so
                        # no recompression, this is the big speedup.
                        for f in batch:
                            zf.write(f, f"validation/{sp}/{f.name}")
                        # Per-batch validation.csv: only the rows for images in
                        # THIS batch, so the manifest matches the images shipped.
                        if sp_fields:
                            buf = io.StringIO()
                            w = csv.DictWriter(buf, fieldnames=sp_fields)
                            w.writeheader()
                            for f in batch:
                                r = rows_by_name.get(f.name)
                                if r:
                                    w.writerow(r)
                            zf.writestr(f"validation/{sp}/validation.csv",
                                        buf.getvalue(), compress_type=DEF)
                        # Other metadata (valid_species.txt etc.) copied as-is
                        for f in other_meta:
                            zf.write(f, f"validation/{sp}/{f.name}")
                        if has_env:
                            for f in validate_env.rglob("*"):
                                if f.is_file():
                                    rel = f.relative_to(validate_env)
                                    zf.write(f, f"validate_env/{rel}",
                                             compress_type=DEF)

                        # Filtered sibling-bbox data for this batch only
                        if master_rows:
                            batch_img_ids = {det_to_img.get(f.stem, "")
                                              for f in batch}
                            batch_img_ids.discard("")
                            sub_rows = [r for r in master_rows
                                        if r.get("image_id","") in batch_img_ids]
                            if sub_rows:
                                buf = io.StringIO()
                                w = csv.DictWriter(buf, fieldnames=list(sub_rows[0].keys()))
                                w.writeheader()
                                w.writerows(sub_rows)
                                zf.writestr("results_with_ids.csv", buf.getvalue())
                    created.append(zname)
                    newly_sent.extend(f.name for f in batch)
                except Exception as e:
                    errors.append(f"{zname}: {e}")

        # Record everything packaged this run so it is never sent again.
        if newly_sent:
            with open(sent_log, "a", encoding="utf-8") as f:
                for name in newly_sent:
                    f.write(name + "\n")

        if not created:
            self._dist_prep_log.config(
                text="Nothing new to send", fg=C["forest"])
            messagebox.showinfo("Nothing to package",
                "Every image has already been sent or validated.\n\n"
                "To re-send from scratch, delete distributed_images.txt in the "
                "distribute\\ folder and try again.")
            return

        msg = (f"{len(created)} zip(s) created in:\n{out_path}\n"
               f"{len(newly_sent)} new image(s) packaged; "
               f"{skipped_done} skipped (already sent or validated).\n"
               f"Max {BATCH} images per zip.")
        if errors:
            msg += "\n\nErrors:\n" + "\n".join(errors)
        self._dist_prep_log.config(
            text=f"{len(created)} package(s) created",
            fg=C["forest"] if not errors else C["error"])
        messagebox.showinfo("Packages ready", msg)


    def _dist_collect(self):
        import zipfile, tempfile, shutil

        # Derive project from current run folder
        proj_str = self._img_folder_var.get().strip()
        if not proj_str:
            messagebox.showerror("No project",
                "Please select a project folder in the Run tab first.")
            return

        project       = Path(proj_str)
        local_val     = project / "validation"
        collect_dir   = project / "collect"
        processed_dir = collect_dir / "processed"
        distribute_dir = project / "distribute"

        if not local_val.exists():
            messagebox.showerror("No validation folder",
                "Run wildtag first to generate validation images.")
            return

        if not collect_dir.exists():
            collect_dir.mkdir(parents=True)
            messagebox.showinfo("Collect folder created",
                f"Created collect\\ folder at:\n{collect_dir}\n\n"
                f"Drop validated zips there and click this button again.")
            return

        zips = list(collect_dir.glob("*.zip"))
        if not zips:
            messagebox.showinfo("Nothing to collect",
                f"No zip files found in:\n{collect_dir}\n\n"
                f"Drop validated zips there and try again.")
            return

        processed_dir.mkdir(exist_ok=True)

        all_preview  = []
        all_merge    = {}
        failed_zips  = []
        read_ok      = []

        for zip_path in zips:
            try:
                tmp = Path(tempfile.mkdtemp())
                with zipfile.ZipFile(zip_path, "r") as zf:
                    zf.extractall(tmp)

                for csv_path in tmp.rglob("validation.csv"):
                    sp = csv_path.parent.name
                    if sp == "validation":
                        continue
                    with open(csv_path, newline="", encoding="utf-8-sig") as f:
                        rows = list(csv.DictReader(f))
                    validated = [r for r in rows
                                 if r.get("validated","").strip().lower() == "yes"]
                    if not validated:
                        continue
                    corrected = [r for r in validated
                                 if r.get("correct_label","").strip()]
                    all_preview.append(
                        f"{zip_path.name} — {sp.replace('_',' ').title()}: "
                        f"{len(validated)} validated, {len(corrected)} corrections")
                    for r in corrected:
                        all_preview.append(
                            f"    {r.get('image_name','')}: "
                            f"{r.get('label','')} -> {r.get('correct_label','')}")
                    if sp not in all_merge:
                        all_merge[sp] = {}
                    for r in validated:
                        did = r.get("detection_id","").strip()
                        if did:
                            all_merge[sp][did] = r

                shutil.rmtree(tmp, ignore_errors=True)
                read_ok.append(zip_path)

            except Exception as e:
                failed_zips.append(f"{zip_path.name}: {e}")

        if not all_merge:
            messagebox.showinfo("Nothing to merge",
                "No validated rows found in any of the collected zips.")
            return

        preview_text = "\n".join(all_preview[:40])
        if len(all_preview) > 40:
            preview_text += f"\n... and {len(all_preview)-40} more"
        if failed_zips:
            preview_text += "\n\nFailed:\n" + "\n".join(failed_zips)

        confirmed = messagebox.askyesno("Merge preview",
            f"Ready to merge from {len(zips)} zip(s):\n\n"
            f"{preview_text}\n\n"
            f"Merge into your local validation files?")
        if not confirmed:
            return

        # Apply merges to local validation CSVs
        merged_sp = []
        for sp, lookup in all_merge.items():
            local_csv = local_val / sp / "validation.csv"
            if not local_csv.exists():
                continue
            with open(local_csv, newline="", encoding="utf-8-sig") as f:
                local_rows = list(csv.DictReader(f))
                f.seek(0)
                fields = list(csv.DictReader(f).fieldnames)
            for row in local_rows:
                det_id = row.get("detection_id","")
                if det_id in lookup:
                    imp = lookup[det_id]
                    row["validated"] = "yes"
                    if imp.get("correct_label","").strip():
                        row["correct_label"] = imp["correct_label"]
            with open(local_csv, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fields)
                writer.writeheader()
                writer.writerows(local_rows)
            merged_sp.append(sp)

        # Move successfully-read zips from collect/ to collect/processed/ and
        # remove the matching zip from distribute/. Zips that failed to process
        # are left in collect/ so they can be retried and stay visible as
        # outstanding in distribute/.
        for zip_path in read_ok:
            try:
                zip_path.rename(processed_dir / zip_path.name)
            except Exception:
                pass
            # Remove from distribute/ too - it's been returned and merged
            dist_zip = distribute_dir / zip_path.name
            if dist_zip.exists():
                try:
                    dist_zip.unlink()
                except Exception:
                    pass

        # Merge to master CSV
        self._val_folder = local_val
        self._val_merge_to_master()

        result = (f"Merged {len(merged_sp)} species folder(s): "
                  f"{', '.join(merged_sp)}\n"
                  f"Zips moved to collect\\processed\\ and removed from distribute\\\n"
                  f"Master results_with_ids.csv updated.")
        if failed_zips:
            result += (f"\n\n{len(failed_zips)} zip(s) could not be read and were "
                       f"left in collect\\ to retry:\n" + "\n".join(failed_zips))
        self._dist_import_log.config(
            text=result, fg=C["forest"] if not failed_zips else C["error"])
        messagebox.showinfo("Collect complete", result)

    def _dist_import(self):
        import zipfile, tempfile, shutil

        zip_path = filedialog.askopenfilename(
            title="Select validated zip to import",
            filetypes=[("Zip files","*.zip"),("All files","*.*")])
        if not zip_path:
            return

        zip_path = Path(zip_path)

        # Ask for local validation folder to merge into
        local_val = filedialog.askdirectory(
            title="Select your local validation folder to merge into")
        if not local_val:
            return
        local_val = Path(local_val)

        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                names = zf.namelist()

            # Find species folder inside zip
            sp_names = sorted({
                n.split("/")[0] for n in names
                if "/" in n and not n.startswith("wildtag")
                and not n.startswith("README")
            })

            if not sp_names:
                messagebox.showerror("Invalid zip",
                    "Could not find a species folder inside this zip.")
                return

            # Extract to temp folder for inspection
            tmp = Path(tempfile.mkdtemp())
            with zipfile.ZipFile(zip_path, "r") as zf:
                zf.extractall(tmp)

            # Build preview of changes
            preview_lines = []
            merge_data = {}  # sp -> list of row dicts

            for sp in sp_names:
                tmp_csv = tmp / sp / "validation.csv"
                local_csv = local_val / sp / "validation.csv"

                if not tmp_csv.exists():
                    preview_lines.append(
                        f"{sp}: no validation.csv found in zip")
                    continue

                with open(tmp_csv, newline="", encoding="utf-8-sig") as f:
                    imported_rows = list(csv.DictReader(f))

                validated = [r for r in imported_rows
                             if r.get("validated","").strip().lower() == "yes"]
                corrected = [r for r in validated
                             if r.get("correct_label","").strip()]

                preview_lines.append(
                    f"{sp.replace('_',' ').title()}: "
                    f"{len(validated)} validated, "
                    f"{len(corrected)} corrections")
                for r in corrected:
                    preview_lines.append(
                        f"    {r.get('image_name','')}:  "
                        f"{r.get('label','')} -> "
                        f"{r.get('correct_label','')}")

                merge_data[sp] = imported_rows

            if not merge_data:
                messagebox.showinfo("Nothing to import",
                    "No validated data found in this zip.")
                shutil.rmtree(tmp, ignore_errors=True)
                return

            # Show preview and ask to confirm
            preview_text = "\n".join(preview_lines)
            confirmed = messagebox.askyesno(
                "Import preview",
                f"Found the following validated data:\n\n"
                f"{preview_text}\n\n"
                f"Merge this into your local validation files?")

            if not confirmed:
                shutil.rmtree(tmp, ignore_errors=True)
                return

            # Merge
            merged_sp = []
            for sp, imported_rows in merge_data.items():
                local_csv = local_val / sp / "validation.csv"
                if not local_csv.exists():
                    continue

                with open(local_csv, newline="", encoding="utf-8-sig") as f:
                    local_rows = list(csv.DictReader(f))
                    f.seek(0)
                    fields     = list(csv.DictReader(f).fieldnames)

                # Build lookup from imported data
                imported_lookup = {
                    r["detection_id"]: r
                    for r in imported_rows
                    if r.get("detection_id")
                }

                # Apply to local rows
                for row in local_rows:
                    det_id = row.get("detection_id","")
                    if det_id in imported_lookup:
                        imp = imported_lookup[det_id]
                        if imp.get("validated","").strip().lower() == "yes":
                            row["validated"]     = "yes"
                            if imp.get("correct_label","").strip():
                                row["correct_label"] = imp["correct_label"]

                with open(local_csv, "w", newline="",
                          encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=fields)
                    writer.writeheader()
                    writer.writerows(local_rows)

                merged_sp.append(sp)

            shutil.rmtree(tmp, ignore_errors=True)

            # Set val_folder so merge to master works
            self._val_folder = local_val
            self._val_merge_to_master()

            result = (f"Merged {len(merged_sp)} species folder(s):\n"
                      f"{', '.join(merged_sp)}\n\n"
                      f"Master results_with_ids.csv updated.")
            self._dist_import_log.config(text=result, fg=C["forest"])
            messagebox.showinfo("Import complete", result)

        except Exception as e:
            messagebox.showerror("Import error", str(e))

    def _map_init_state(self):
        """Initialise map data state. The map itself is launched from a
        button in the Summary pane and opens in the browser, so there's no
        dedicated Map pane any more, just the underlying data + HTML
        generation."""
        self._map_species_var = tk.StringVar(value="all")
        self._map_deps      = []
        self._map_site_obs  = {}
        self._map_html_path = None
        self._map_ready     = False

    def _map_placeholder_html(self):
        return ""  # not used in browser mode

    def _map_open_browser(self):
        if self._map_html_path and Path(self._map_html_path).exists():
            import webbrowser, time
            # Add timestamp to bust cache and force browser refresh
            url = f"file:///{self._map_html_path}?t={int(time.time())}"
            webbrowser.open(url)

    def _map_render(self):
        """Write the map HTML to a temp file. The browser opens it on
        demand from the Launch map button in Summary."""
        if not getattr(self, "_map_deps", []):
            return
        import tempfile, os as _os
        filt = self._map_species_var.get()
        html = self._map_generate_html(self._map_deps, self._map_site_obs, filt)
        if not self._map_html_path:
            fd, path = tempfile.mkstemp(suffix=".html", prefix="wildtag_map_")
            _os.close(fd)
            self._map_html_path = path
        with open(self._map_html_path, "w", encoding="utf-8") as f:
            f.write(html)
        self._map_ready = True
        # Enable the Summary launch button if it exists yet
        btn = getattr(self, "_summary_map_btn", None)
        if btn is not None:
            try:
                btn.config(state="normal", bg=C["forest"], cursor="hand2")
            except tk.TclError:
                pass

    def _map_status_set(self, text):
        """Route map status text to the Summary pane's map status label,
        if it exists. Safe no-op otherwise (there's no dedicated map pane)."""
        lbl = getattr(self, "_summary_map_status", None)
        if lbl is not None:
            try:
                lbl.config(text=text)
            except tk.TclError:
                pass

    def _map_refresh(self):
        """Load deployment + observation data then render the map."""

        project_str = self._img_folder_var.get().strip()
        if not project_str:
            self._map_status_set(
                "No project selected - use the Run tab first")
            return

        project = Path(project_str)

        # Read deployment file - prefer Camtrap DP export if available
        dep_rows   = None
        dep_source = None

        cdp_path = project / "camtrapdp" / "deployments.csv"
        if cdp_path.exists():
            try:
                with open(cdp_path, newline="", encoding="utf-8-sig") as f:
                    dep_rows = list(csv.DictReader(f))
                dep_source = "camtrapdp/deployments.csv"
            except Exception:
                pass

        if not dep_rows:
            for name in ["deployment.csv","deployment.xlsx",
                         "deployments.csv","deployments.xlsx"]:
                p = project / name
                if not p.exists():
                    continue
                try:
                    if p.suffix.lower() == ".xlsx":
                        import openpyxl
                        wb = openpyxl.load_workbook(p, data_only=True)
                        ws = wb.active
                        headers = [str(c.value or "").strip()
                                   for c in next(ws.iter_rows())]
                        dep_rows = [dict(zip(headers,[
                            str(v or "").strip() for v in row]))
                            for row in ws.iter_rows(min_row=2, values_only=True)]
                    else:
                        with open(p, newline="", encoding="utf-8-sig") as f:
                            dep_rows = list(csv.DictReader(f))
                    dep_source = name
                    break
                except Exception as e:
                    self._map_status_set(f"Error reading {name}: {e}")
                    return

        if not dep_rows:
            self._map_status_set("No deployment file found in project folder")
            return

        def _col(row, *candidates):
            for c in candidates:
                for k in row.keys():
                    if k.strip().lower() == c.lower():
                        return k
            return None

        sample   = dep_rows[0]
        lat_col  = _col(sample, "latitude", "lat", "y")
        lon_col  = _col(sample, "longitude", "lon", "long", "x")
        site_col = _col(sample, "locationname", "location_name",
                        "site", "sitename", "site_name", "station")

        if not all([lat_col, lon_col, site_col]):
            missing = [n for n,c in [("latitude",lat_col),
                                      ("longitude",lon_col),
                                      ("locationName",site_col)] if not c]
            self._map_status_set(
                f"Deployment file missing: {', '.join(missing)}")
            return

        deps = []
        for row in dep_rows:
            try:
                lat  = float(row.get(lat_col,0) or 0)
                lon  = float(row.get(lon_col,0) or 0)
                site = row.get(site_col,"").strip()
                if not site or (lat==0 and lon==0):
                    continue
                deps.append({
                    "site":  site, "lat": lat, "lon": lon,
                    "start": row.get("deploymentStart",
                             row.get("deployment_start","")).strip(),
                    "end":   row.get("deploymentEnd",
                             row.get("deployment_end","")).strip(),
                })
            except (ValueError, TypeError):
                continue

        if not deps:
            self._map_status_set(
                "No valid coordinates found in deployment file")
            return

        # Pre-index deployment site names so the aggregate builder's
        # site keys (and any name-style variants) line up with them
        site_by_key = {}
        for d in deps:
            s = d["site"]
            site_by_key[s.lower()] = d["site"]
            site_by_key[s.lower().replace(" ", "_")] = d["site"]
            site_by_key[s.lower().replace("_", " ")] = d["site"]
        self._map_site_keys = site_by_key

        # Pull per-site species counts from the shared aggregate cache
        # instead of re-parsing the whole results file on every map open.
        agg = self._get_aggregates(on_ready=self._map_refresh)
        if agg is None:
            self._map_status_set("Loading detection data…")
            return

        site_obs = {d["site"]: {} for d in deps}
        for site_name, spp in agg["site_species"].items():
            # Map the aggregate's site key onto a deployment site name
            matched = site_by_key.get(site_name.lower())
            if not matched:
                continue
            bucket = site_obs[matched]
            for sp, n in spp.items():
                bucket[sp] = bucket.get(sp, 0) + n

        self._map_deps     = deps
        self._map_site_obs = site_obs

        # Species filtering now happens inside the browser map, so there's
        # no app-side dropdown to populate here.

        total_dets = sum(n for obs in site_obs.values()
                         for sp,n in obs.items()
                         if sp not in ("empty","low_confidence"))
        self._map_status_set(
            f"{len(deps)} sites, {total_dets} detections from {dep_source}")

        self._map_render()

    def _map_generate_html(self, deps, site_obs, filt):
        import json as _json
        centre_lat = sum(d["lat"] for d in deps) / len(deps)
        centre_lon = sum(d["lon"] for d in deps) / len(deps)
        SKIP = {"empty","low_confidence","unclassified","human","vehicle"}
        deps_esc = _json.dumps(deps).replace(chr(39), chr(92)+chr(39))
        obs_esc  = _json.dumps(site_obs).replace(chr(39), chr(92)+chr(39))
        filt_json = _json.dumps(filt)
        skip_json = _json.dumps(list(SKIP))
        return f"""<!DOCTYPE html>
<html><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>wildtag.ai</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  html,body{{margin:0;padding:0;height:100%;font-family:system-ui,sans-serif}}
  #map{{height:100vh}}
  .pt{{font-weight:600;font-size:14px;margin-bottom:6px;color:#1a2e1e}}
  .pd{{font-size:11px;color:#888;margin-bottom:8px}}
  table.sp{{width:100%;border-collapse:collapse;font-size:13px}}
  table.sp td{{padding:3px 6px 3px 0}}
  table.sp td:last-child{{text-align:right;font-weight:500;padding-left:12px}}
  .tr{{border-top:1px solid #eee;font-weight:600}}
  .er{{color:#bbb}}
  #ctrl{{position:absolute;top:10px;right:10px;z-index:1000;background:#fff;
    padding:10px 12px;border-radius:8px;box-shadow:0 1px 6px rgba(0,0,0,.3);
    font-size:13px;color:#1a2e1e}}
  #ctrl label{{font-weight:600;margin-right:6px}}
  #ctrl select{{font-size:13px;padding:3px 6px;border:1px solid #ccc;
    border-radius:4px;max-width:200px}}
  #ctrl .cnt{{margin-top:6px;color:#888;font-size:11px}}
</style>
</head><body>
<div id="map"></div>
<div id="ctrl">
  <label for="sp">Species</label>
  <select id="sp"></select>
  <div class="cnt" id="cnt"></div>
</div>
<script>
const deps=JSON.parse('{deps_esc}');
const obs=JSON.parse('{obs_esc}');
const SKIP=new Set({skip_json});
let filter={filt_json};

const map=L.map('map').setView([{centre_lat},{centre_lon}],12);
const positron=L.tileLayer('https://{{s}}.basemaps.cartocdn.com/light_all/{{z}}/{{x}}/{{y}}{{r}}.png',{{attribution:'© OpenStreetMap contributors © CARTO',maxZoom:19}});
const satellite=L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}',{{attribution:'Tiles © Esri',maxZoom:19}});
positron.addTo(map);
L.control.layers({{'Light':positron,'Satellite':satellite}}).addTo(map);
L.control.scale({{imperial:false}}).addTo(map);

// Build the species list from ALL sites' data, so filtering happens
// entirely in-page, no need to regenerate from the app.
const spp=new Set();
Object.values(obs).forEach(o=>Object.keys(o).forEach(s=>{{if(!SKIP.has(s))spp.add(s);}}));
const sel=document.getElementById('sp');
const cnt=document.getElementById('cnt');
const optAll=document.createElement('option');
optAll.value='all';optAll.textContent='All species';sel.appendChild(optAll);
[...spp].sort().forEach(s=>{{
  const o=document.createElement('option');
  o.value=s;o.textContent=s.replace(/_/g,' ');sel.appendChild(o);
}});
sel.value=(filter&&filter!=='all')?filter:'all';

let markers=[];
function draw(){{
  markers.forEach(m=>map.removeLayer(m));
  markers=[];
  let activeSites=0,totalDets=0;
  deps.forEach(dep=>{{
    const o=obs[dep.site]||{{}};
    const dets=Object.entries(o).filter(([s])=>!SKIP.has(s));
    const total=dets.reduce((a,[,v])=>a+v,0);
    const empty=o['empty']||0;
    const active=filter==='all'?total>0:(o[filter]||0)>0;
    if(active)activeSites++;
    totalDets+=(filter==='all'?total:(o[filter]||0));
    const sizeBasis=filter==='all'?total:(o[filter]||0);
    const r=Math.max(7,Math.min(14,7+Math.sqrt(Math.max(0,sizeBasis))*1.0));
    const col=active?'#2D7A45':'#c0392b';
    const d2=(r+5)*2;
    const svg=`<svg xmlns="http://www.w3.org/2000/svg" width="${{d2}}" height="${{d2}}"><circle cx="${{d2/2}}" cy="${{d2/2}}" r="${{r}}" fill="${{col}}" stroke="${{active?'#1a4a28':'#7a2418'}}" stroke-width="1.5" opacity="${{active?1:0.7}}"/></svg>`;
    const icon=L.divIcon({{html:svg,className:'',iconSize:[d2,d2],iconAnchor:[d2/2,d2/2],popupAnchor:[0,-d2/2]}});
    const rows=dets.sort(([,a],[,b])=>b-a).slice(0,12).map(([s,n])=>`<tr${{s===filter?' style="background:#eafaf0"':''}}><td>${{s.replace(/_/g,' ')}}</td><td>${{n}}</td></tr>`).join('');
    const popup=`<div class="pt">${{dep.site}}</div>${{dep.start?`<div class="pd">${{dep.start}} – ${{dep.end||''}}</div>`:''}}<table class="sp"><tr class="tr"><td>Detections</td><td>${{total}}</td></tr>${{rows}}${{empty?`<tr class="er"><td>empty</td><td>${{empty}}</td></tr>`:''}}</table>`;
    const mk=L.marker([dep.lat,dep.lon],{{icon}}).bindPopup(popup,{{maxWidth:220}}).addTo(map);
    markers.push(mk);
  }});
  const lbl=filter==='all'?'all species':filter.replace(/_/g,' ');
  cnt.textContent=`${{activeSites}} of ${{deps.length}} sites · ${{totalDets}} detections (${{lbl}})`;
}}
sel.addEventListener('change',()=>{{filter=sel.value;draw();}});
draw();
</script></body></html>"""


    def _build_pane_summary(self):
        pane = tk.Frame(self._pane_area, bg=C["frost"])
        self._panes["summary"] = pane
        inner = self._scrollable(pane)
        pad   = dict(padx=24, pady=6)

        tk.Frame(inner, bg=C["frost"], height=8).pack()
        tk.Label(inner, text="Summary", font=self._fonts["h2"],
                 bg=C["frost"], fg=C["canopy"], anchor="w").pack(
                     fill="x", padx=24, pady=(0,2))
        self._wrap_label(inner,
            "This page shows what was processed and the current state of validation.",
            bg=C["frost"]).pack(fill="x", padx=24)
        tk.Frame(inner, bg=C["frost"], height=12).pack()

        # ── Pipeline stats ───────────────────────────────────────────────
        self._section_label(inner, "What was processed").pack(
            fill="x", padx=24, pady=(0,4))
        stats_row = tk.Frame(inner, bg=C["frost"])
        stats_row.pack(fill="x", padx=24, pady=6)

        self._stat_images  = tk.StringVar(value="0")
        self._stat_species = tk.StringVar(value="0")
        self._stat_errors  = tk.StringVar(value="0")
        self._stat_time    = tk.StringVar(value="0")  # kept for run summary, not shown

        for var, label in [
            (self._stat_images,  "Images sorted"),
            (self._stat_species, "Species found"),
            (self._stat_errors,  "Images skipped"),
        ]:
            sc = tk.Frame(stats_row, bg=C["white"], padx=14, pady=12)
            sc.pack(side="left", fill="x", expand=True, padx=(0,8))
            tk.Label(sc, textvariable=var, font=self._fonts["stat"],
                     bg=C["white"], fg=C["forest"]).pack()
            tk.Label(sc, text=label, font=self._fonts["small"],
                     bg=C["white"], fg=C["text_muted"]).pack()

        # Launch map, occupies the slot the low-value "Last run" card used
        # to. Opens the interactive deployment map in the browser (with its
        # own species filter). Disabled until the map data is ready.
        map_card = tk.Frame(stats_row, bg=C["white"], padx=14, pady=12)
        map_card.pack(side="left", fill="x", expand=True, padx=(0,8))
        self._summary_map_btn = tk.Button(map_card, text="🗺  Launch map",
                  command=self._map_open_browser,
                  font=("Segoe UI", 11, "bold"),
                  bg=C["mist"], fg=C["canopy"],
                  activebackground=C["leaf"], activeforeground=C["white"],
                  relief="flat", padx=12, pady=6, cursor="hand2",
                  state="disabled")
        self._summary_map_btn.pack()
        self._summary_map_status = tk.Label(map_card, text="deployment map",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"])
        self._summary_map_status.pack()

        # ── Validation stats ─────────────────────────────────────────────
        self._section_label(inner, "Validation progress").pack(
            fill="x", padx=24, pady=(16,4))
        val_row = tk.Frame(inner, bg=C["frost"])
        val_row.pack(fill="x", padx=24, pady=6)

        self._stat_validated   = tk.StringVar(value="0")
        self._stat_pct         = tk.StringVar(value="0%")
        self._stat_corrections = tk.StringVar(value="0")

        for var, label in [
            (self._stat_validated,   "Detections validated"),
            (self._stat_pct,         "Percent validated"),
            (self._stat_corrections, "Labels corrected"),
        ]:
            sc = tk.Frame(val_row, bg=C["white"], padx=14, pady=12)
            sc.pack(side="left", fill="x", expand=True, padx=(0,8))
            tk.Label(sc, textvariable=var, font=self._fonts["stat"],
                     bg=C["white"], fg=C["forest"]).pack()
            tk.Label(sc, text=label, font=self._fonts["small"],
                     bg=C["white"], fg=C["text_muted"]).pack()

        # Refresh validation stats button
        tk.Button(inner, text="Refresh validation stats",
                  command=self._refresh_val_stats,
                  font=self._fonts["small"],
                  bg=C["border"], fg=C["canopy"],
                  relief="flat", padx=10, pady=4,
                  cursor="hand2").pack(anchor="e", padx=24)

        # ── Confusion matrix ─────────────────────────────────────────────
        self._section_label(inner, "Confusion matrix").pack(
            fill="x", padx=24, pady=(16,4))

        cm_ctrl = tk.Frame(inner, bg=C["frost"])
        cm_ctrl.pack(fill="x", padx=24, pady=(0,6))
        self._wrap_label(cm_ctrl,
            "Shows how AI labels compare to validated labels. "
            "Rows are what the AI predicted, columns are the validated labels.",
            bg=C["frost"]).pack(side="left", fill="x", expand=True)

        self._cm_mode = tk.StringVar(value="errors")
        tk.Radiobutton(cm_ctrl, text="Errors only",
                       variable=self._cm_mode, value="errors",
                       bg=C["frost"], fg=C["canopy"],
                       selectcolor=C["white"],
                       activebackground=C["frost"],
                       font=self._fonts["small"],
                       command=self._refresh_val_stats).pack(
                           side="right", padx=(8,0))
        tk.Radiobutton(cm_ctrl, text="All validated",
                       variable=self._cm_mode, value="all",
                       bg=C["frost"], fg=C["canopy"],
                       selectcolor=C["white"],
                       activebackground=C["frost"],
                       font=self._fonts["small"],
                       command=self._refresh_val_stats).pack(
                           side="right", padx=(8,0))

        o_cm, c_cm = self._card(inner); o_cm.pack(fill="x", **pad)
        self._cm_frame = c_cm
        tk.Label(c_cm,
                 text="Confusion matrix will appear here once images have been validated.",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"], anchor="w").pack(fill="x")

        # ── Species breakdown ────────────────────────────────────────────
        self._section_label(inner, "How many detections per species").pack(
            fill="x", padx=24, pady=(16,4))
        o2, c2 = self._card(inner); o2.pack(fill="x", **pad)
        self._species_frame = c2
        self._wrap_label(c2,
            "Once you have run the pipeline, you will see a count of detections per species here.",
            bg=C["white"]).pack(fill="x")

        # ── Output location ──────────────────────────────────────────────
        self._section_label(inner, "Where your files have been saved").pack(
            fill="x", padx=24, pady=(16,4))
        o3, c3 = self._card(inner); o3.pack(fill="x", **pad)
        self._output_lbl = tk.Label(
            c3,
            text="The path to your validation folder will appear here once the pipeline has finished.",
            font=self._fonts["small"], bg=C["white"],
            fg=C["text_muted"], anchor="w")
        self._output_lbl.pack(fill="x", pady=(0,4))

    def _prelink_project(self):
        """Called shortly after a project folder is selected. If the project
        is already processed (results_with_ids.csv exists), start building
        the aggregate cache now, with the linking popup, so Map and Summary
        are ready by the time the user opens them. Does nothing for an
        unprocessed folder (nothing to link yet)."""
        path = self._img_folder_var.get().strip()
        if not path:
            return
        master = Path(path) / "results_with_ids.csv"
        if not master.exists():
            return  # not processed yet, nothing to link
        # Kick off the build (shows the linking popup). No on_ready callback
        # needed: this is purely to warm the cache ahead of time.
        self._get_aggregates()

        # Also check whether the validation folder is complete. A crash
        # during the sort step (e.g. an interrupted GPU run) can leave it
        # half-built. If so, offer to finish it rather than let the user
        # discover missing images/folders later during validation.
        try:
            info = check_validation_complete(path)
            if info["status"] == "incomplete":
                self.after(300, lambda: self._offer_resume_sort(path, info))
        except Exception:
            pass

    def _offer_resume_sort(self, project_path, info):
        """Prompt to finish an interrupted/incomplete sort, then run it. The
        sort is resumable, so this only fills in what's missing."""
        do_it = messagebox.askyesno(
            "Finish building validation folders?",
            "wildtag detected that the validation folders for this project "
            "are incomplete, most likely a previous run was interrupted "
            "before it finished.\n\n"
            f"{info['reason']}\n\n"
            "Finish building them now? (Already-sorted images are kept and "
            "skipped, so this only completes the missing work.)")
        if not do_it:
            return
        enriched = Path(project_path) / "results_with_ids.csv"
        if not enriched.exists():
            enriched = Path(project_path) / "results.csv"
        if not enriched.exists():
            messagebox.showerror(
                "Cannot resume",
                "The results file needed to rebuild the validation folders "
                "was not found in this project.")
            return

        def _work():
            try:
                quality = self._QUALITY_MAP.get(
                    self._quality_var.get(), 65) if hasattr(self, "_quality_var") else 65
                cls = getattr(self, "_last_classifier_id", "") or ""
                sort_detections_counted(
                    enriched, quality, None,
                    self._log,
                    classifier_id=cls)
                self.after(0, lambda: messagebox.showinfo(
                    "Validation folders complete",
                    "The validation folders have been finished. You can now "
                    "validate as normal."))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror(
                    "Resume failed",
                    f"Could not finish the validation folders:\n{e}"))
        threading.Thread(target=_work, daemon=True).start()

    def _show_linking_popup(self):
        """Small modal shown while the aggregate cache builds, so the user
        knows wildtag is syncing rather than frozen."""
        if getattr(self, "_linking_popup", None):
            return
        try:
            win = tk.Toplevel(self)
            win.transient(self)
            win.title("wildtag.ai")
            win.configure(bg=C["white"])
            win.resizable(False, False)
            frame = tk.Frame(win, bg=C["white"],
                             highlightbackground=C["forest"],
                             highlightthickness=2)
            frame.pack(fill="both", expand=True)
            tk.Label(frame, text="Linking wildtag.ai to your project",
                     font=self._fonts["h2"], bg=C["white"],
                     fg=C["canopy"]).pack(padx=40, pady=(28,6))
            tk.Label(frame,
                     text="Reading detection data. This happens once, then "
                          "switching between panes is instant.",
                     font=self._fonts["small"], bg=C["white"],
                     fg=C["text_muted"], wraplength=340,
                     justify="center").pack(padx=40, pady=(0,10))
            bar = tk.Frame(frame, bg=C["mist"], height=4, width=320)
            bar.pack(padx=40, pady=(0,24))
            fill = tk.Frame(bar, bg=C["forest"], height=4)
            fill.place(x=0, y=0, relheight=1, width=40)

            # Indeterminate sliding bar
            state = {"x": 0, "dir": 1}
            def _pulse():
                if not getattr(self, "_linking_popup", None):
                    return
                state["x"] += state["dir"] * 12
                if state["x"] > 280: state["x"] = 280; state["dir"] = -1
                elif state["x"] < 0: state["x"] = 0; state["dir"] = 1
                try:
                    fill.place(x=state["x"], y=0, relheight=1, width=40)
                    win.after(40, _pulse)
                except tk.TclError:
                    pass

            self._linking_popup = win
            # Position before showing
            self.update_idletasks()
            px = self.winfo_rootx() + (self.winfo_width()  - 420) // 2
            py = self.winfo_rooty() + (self.winfo_height() - 180) // 2
            win.geometry(f"420x180+{max(0,px)}+{max(0,py)}")
            # Force the window to actually map and paint NOW. update_idletasks
            # alone processes geometry but does not reliably render the
            # Toplevel before the caller starts the heavy work, so without
            # this the popup can exist invisibly and get torn down before it
            # ever shows. lift + a full update() guarantees it's on screen.
            win.lift()
            win.attributes("-topmost", True)
            win.update()
            import time as _t
            self._linking_shown_at = _t.time()
            _pulse()
        except Exception:
            self._linking_popup = None

    def _hide_linking_popup(self):
        win = getattr(self, "_linking_popup", None)
        if not win:
            return
        # Keep it visible for at least ~600ms so it's always perceptible,
        # even if the build finishes quickly
        import time as _t
        shown_at = getattr(self, "_linking_shown_at", 0)
        elapsed_ms = (_t.time() - shown_at) * 1000
        remaining = int(max(0, 600 - elapsed_ms))
        def _destroy():
            self._linking_popup = None
            try: win.destroy()
            except Exception: pass
        if remaining > 0:
            self.after(remaining, _destroy)
        else:
            _destroy()

    def _find_master_csv(self):
        """Locate results_with_ids.csv from the current project context,
        without prompting. Returns a Path or None."""
        if getattr(self, "_last_output_dir", None):
            c = Path(self._last_output_dir).parent / "results_with_ids.csv"
            if c.exists():
                return c
        v = getattr(self, "_img_folder_var", None)
        if v and v.get().strip():
            c = Path(v.get().strip()) / "results_with_ids.csv"
            if c.exists():
                return c
        s = self._settings.get("project", "").strip()
        if s:
            c = Path(s) / "results_with_ids.csv"
            if c.exists():
                return c
        return None

    def _get_aggregates(self, on_ready=None, force=False):
        """Return precomputed aggregates from results_with_ids.csv, or None
        if not yet available. Reads the big file at most once per file
        version: subsequent pane switches reuse the cached result.

        The heavy parse runs in a background thread. If a fresh parse is
        needed, this returns None immediately and calls on_ready() once the
        aggregates are built. If the cache is already valid, it returns the
        aggregates synchronously.

        Aggregates (all small) computed in a single pass:
          site_species  : {site: {species: count}}
          species_total : {species: count}
          n_rows         : total detections
          n_validated    : validated rows
          n_corrected    : validated rows whose label was changed
          cm_errors      : {(predicted, actual): count}  (corrections only)
          cm_all         : {(predicted, actual): count}  (all validated)
        """
        master = self._find_master_csv()
        if not master:
            return None
        try:
            mtime = master.stat().st_mtime
        except OSError:
            return None
        key = (str(master), mtime)

        if not force and self._agg_cache_key == key and self._agg_cache:
            return self._agg_cache

        if self._agg_loading:
            return None
        self._agg_loading = True
        self._show_linking_popup()

        # Pre-index deployment site names (for map matching) once, off the
        # aggregate build so it stays purely about the results file
        site_keys = getattr(self, "_map_site_keys", None)

        def _build():
            from collections import defaultdict
            site_species  = defaultdict(lambda: defaultdict(int))
            species_total = defaultdict(int)
            n_rows = n_validated = n_corrected = 0
            cm_errors = defaultdict(int)
            cm_all    = defaultdict(int)
            try:
                with open(master, newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    for r in reader:
                        n_rows += 1
                        label   = r.get("label","").strip()
                        correct = r.get("correct_label","").strip()
                        final   = correct or label
                        if label:
                            species_total[final] += 1

                        # site aggregation for the map
                        loc = r.get("locationName","").strip()
                        if not loc:
                            rel = r.get("relative_path","").strip().replace("\\","/")
                            parts = [p for p in rel.split("/") if p]
                            for p in parts[:-1]:
                                loc = p
                                break
                        if loc and final:
                            site_species[loc][final] += 1

                        # validation aggregation for summary + matrix
                        if r.get("validated","").strip().lower() == "yes":
                            n_validated += 1
                            if label:
                                cm_all[(label, correct or label)] += 1
                            if correct:
                                n_corrected += 1
                                cm_errors[(label, correct)] += 1
            except Exception:
                pass

            agg = {
                "site_species":  {k: dict(v) for k,v in site_species.items()},
                "species_total": dict(species_total),
                "n_rows":        n_rows,
                "n_validated":   n_validated,
                "n_corrected":   n_corrected,
                "cm_errors":     dict(cm_errors),
                "cm_all":        dict(cm_all),
            }

            def _done():
                self._agg_cache     = agg
                self._agg_cache_key = key
                self._agg_loading   = False
                self._hide_linking_popup()
                if on_ready:
                    try: on_ready()
                    except Exception: pass
            self.after(0, _done)

        threading.Thread(target=_build, daemon=True).start()
        return None

    def _refresh_val_stats(self, auto=False):
        """Update validation stats + confusion matrix from precomputed
        aggregates. The heavy file parse happens at most once per file
        version (see _get_aggregates); this method itself is cheap.
        auto=True stays silent if no project is found yet."""
        master = self._find_master_csv()
        if not master:
            if auto:
                return
            p = filedialog.askopenfilename(
                title="Select results_with_ids.csv",
                filetypes=[("CSV files","*.csv"),("All files","*.*")])
            if not p:
                return
            # Point the app at the chosen file's folder so the cache can
            # find it, then fall through to the cache path
            self._last_output_dir = str(Path(p).parent / "validation")

        agg = self._get_aggregates(
            on_ready=lambda: self._refresh_val_stats(auto=True))
        if agg is None:
            # Parsing in the background; show a loading state in both the
            # cards and the matrix area so nothing reads as a misleading
            # zero. on_ready re-runs this once the cache is built.
            self._stat_validated.set("…")
            self._stat_pct.set("…")
            self._stat_corrections.set("…")
            for w in self._cm_frame.winfo_children():
                w.destroy()
            tk.Label(self._cm_frame,
                     text="Loading validation data…",
                     font=self._fonts["small"], bg=C["white"],
                     fg=C["text_muted"], anchor="w").pack(fill="x")
            return

        total  = agg["n_rows"]
        n_val  = agg["n_validated"]
        n_corr = agg["n_corrected"]
        pct    = f"{n_val/total*100:.1f}%" if total else "0%"

        self._stat_validated.set(f"{n_val:,}")
        self._stat_pct.set(pct)
        self._stat_corrections.set(f"{n_corr:,}")

        # Confusion matrix, from precomputed pair counts
        mode   = self._cm_mode.get()
        counts = dict(agg["cm_errors"] if mode == "errors" else agg["cm_all"])
        n_cm   = sum(counts.values())

        # Clear old matrix
        for w in self._cm_frame.winfo_children():
            w.destroy()

        if not counts:
            tk.Label(self._cm_frame,
                     text="No validated data to display yet.",
                     font=self._fonts["small"], bg=C["white"],
                     fg=C["text_muted"], anchor="w").pack(fill="x")
            return

        predicted_labels = sorted({p[0] for p in counts})
        actual_labels    = sorted({p[1] for p in counts})
        all_labels       = sorted(set(predicted_labels) | set(actual_labels))

        # Render as a table
        n = len(all_labels)
        cell_w = 14

        tk.Label(self._cm_frame,
                 text=f"{'Errors only' if mode=='errors' else 'All validated'}  "
                      f"|  {n_cm:,} detections  "
                      f"|  rows = AI prediction, columns = validated label",
                 font=self._fonts["small"], bg=C["white"],
                 fg=C["text_muted"], anchor="w").pack(fill="x", pady=(0,6))

        tbl = tk.Frame(self._cm_frame, bg=C["white"])
        tbl.pack(anchor="w")

        tk.Label(tbl, text="", width=22, bg=C["white"],
                 font=self._fonts["small"]).grid(row=0, column=0, padx=1)
        for j, lbl in enumerate(all_labels):
            short = lbl.replace("_"," ")[:12]
            tk.Label(tbl, text=short, width=cell_w,
                     bg=C["mist"], fg=C["canopy"],
                     font=("Segoe UI", 8, "bold"),
                     relief="flat", anchor="center").grid(
                         row=0, column=j+1, padx=1, pady=1)

        for i, pred in enumerate(all_labels):
            short = pred.replace("_"," ")[:20]
            tk.Label(tbl, text=short, width=22,
                     bg=C["mist"], fg=C["canopy"],
                     font=("Segoe UI", 8, "bold"),
                     anchor="w").grid(row=i+1, column=0, padx=1, pady=1)
            for j, actual in enumerate(all_labels):
                count = counts.get((pred, actual), 0)
                on_diag = pred == actual
                if count == 0:
                    bg = C["white"]; fg = C["mist"]
                elif on_diag:
                    bg = C["frost"]; fg = C["forest"]
                else:
                    bg = "#FFD6D6"; fg = "#C0392B"
                tk.Label(tbl, text=str(count) if count else "",
                         width=cell_w, bg=bg, fg=fg,
                         font=self._fonts["small"],
                         relief="flat", anchor="center").grid(
                             row=i+1, column=j+1, padx=1, pady=1)

    def _rebuild_species_list(self, species_counts):
        for w in self._species_frame.winfo_children():
            w.destroy()
        if species_counts:
            for sp, count in sorted(species_counts.items(),
                                    key=lambda x: -x[1]):
                row = tk.Frame(self._species_frame, bg=C["white"])
                row.pack(fill="x", pady=2)
                tk.Label(row, text=sp.replace("_"," ").title(),
                         font=self._fonts["label"],
                         bg=C["white"], fg=C["canopy"],
                         anchor="w").pack(side="left")
                tk.Label(row, text=f"{count:,}",
                         font=self._fonts["label"],
                         bg=C["white"], fg=C["forest"],
                         anchor="e").pack(side="right")
        else:
            tk.Label(self._species_frame,
                     text="No species data available.",
                     font=self._fonts["small"], bg=C["white"],
                     fg=C["text_muted"], anchor="w").pack(fill="x")

    def _update_summary(self, success, skipped, species_counts, output_dir):
        import datetime
        self._last_output_dir = output_dir
        self._stat_images.set(f"{success:,}")
        self._stat_species.set(str(len(species_counts)))
        self._stat_errors.set(str(skipped))
        self._stat_time.set(datetime.datetime.now().strftime("%H:%M"))
        self._output_lbl.config(text=str(output_dir), fg=C["canopy"])
        self._rebuild_species_list(species_counts)

        # Persist so the Summary pane can restore these after a restart,
        # previously they lived only in memory and showed 0 next session
        try:
            stats = {
                "images":  success,
                "species": len(species_counts),
                "skipped": skipped,
                "when":    datetime.datetime.now().strftime("%d %b %H:%M"),
                "species_counts": dict(species_counts),
            }
            (Path(output_dir).parent / "wildtag_run_stats.json").write_text(
                json.dumps(stats), encoding="utf-8")
        except Exception:
            pass

    def _restore_summary_stats(self):
        """Repopulate the 'What was processed' cards after a restart. The
        numbers are set in memory when a run completes; on a fresh launch
        they're read back from wildtag_run_stats.json, or reconstructed
        from the validation CSVs for projects processed before that file
        existed."""
        # Already populated this session (a run finished, or restored)
        if self._stat_images.get() not in ("", "0"):
            return

        proj = None
        if getattr(self, "_last_output_dir", None):
            proj = Path(self._last_output_dir).parent
        if not proj:
            v = getattr(self, "_img_folder_var", None)
            if v and v.get().strip():
                proj = Path(v.get().strip())
        if not proj:
            s = self._settings.get("project", "").strip()
            if s:
                proj = Path(s)
        if not proj or not proj.exists():
            return

        # Preferred source: the human-readable run summary written at the
        # end of every run. It already contains the exact figures, so
        # there's no need to reconstruct anything when it exists. It's also
        # written fresh on the most recent run, so it reflects the current
        # state of the project.
        summary_txt = proj / "wildtag_run_summary.txt"
        if summary_txt.exists():
            try:
                vals = {}
                counts = {}
                in_counts = False
                for line in summary_txt.read_text(encoding="utf-8").splitlines():
                    s = line.strip()
                    if s.startswith("Species counts"):
                        in_counts = True
                        continue
                    if in_counts:
                        # "  Red Deer                     1,234"
                        parts = s.rsplit(None, 1)
                        if len(parts) == 2 and parts[1].replace(",","").isdigit():
                            counts[parts[0]] = int(parts[1].replace(",",""))
                        continue
                    if ":" in line:
                        k, _, v = line.partition(":")
                        vals[k.strip().lower()] = v.strip()
                if vals:
                    def _num(key, default="0"):
                        return vals.get(key, default).replace(",","") or "0"
                    self._stat_images.set(f"{int(_num('processed')):,}")
                    self._stat_species.set(_num("species"))
                    self._stat_errors.set(_num("skipped"))
                    self._stat_time.set(vals.get("date","-"))
                    self._rebuild_species_list(counts)
                    self._output_lbl.config(text=str(proj / "validation"),
                                            fg=C["canopy"])
                    return
            except Exception:
                pass

        # Exact stats saved when the run finished, if available
        stats_file = proj / "wildtag_run_stats.json"
        if stats_file.exists():
            try:
                st = json.loads(stats_file.read_text(encoding="utf-8"))
                self._stat_images.set(f"{int(st.get('images', 0)):,}")
                self._stat_species.set(str(st.get("species", 0)))
                self._stat_errors.set(str(st.get("skipped", 0)))
                self._stat_time.set(st.get("when", "-"))
                self._rebuild_species_list(st.get("species_counts") or {})
                self._output_lbl.config(text=str(proj / "validation"),
                                        fg=C["canopy"])
                return
            except Exception:
                pass

        # Project processed before the stats file existed, reconstruct
        # from the validation CSVs. Row counting scales with project size,
        # so do it off the UI thread and populate when done.
        val_dir = proj / "validation"
        if not val_dir.exists():
            return

        def _reconstruct():
            species_counts = {}
            skipped = 0
            try:
                for d in sorted(val_dir.iterdir()):
                    vcsv = d / "validation.csv"
                    if not d.is_dir() or not vcsv.exists():
                        continue
                    with open(vcsv, newline="", encoding="utf-8") as f:
                        n = sum(1 for _ in csv.DictReader(f))
                    if n:
                        species_counts[d.name] = n
                elog = val_dir / "errors.log"
                if elog.exists():
                    skipped = sum(1 for ln in
                                  elog.read_text(encoding="utf-8").splitlines()
                                  if ln.strip())
            except Exception:
                return
            if not species_counts:
                # Nothing to show, clear the loading placeholder back to 0
                def _clear():
                    if self._stat_images.get() == "…":
                        self._stat_images.set("0")
                        self._stat_species.set("0")
                self.after(0, _clear)
                return

            import datetime
            try:
                mt = (proj / "results_with_ids.csv").stat().st_mtime
                when = datetime.datetime.fromtimestamp(mt).strftime("%d %b")
            except OSError:
                when = "-"

            def _apply():
                total = sum(species_counts.values())
                self._stat_images.set(f"{total:,}")
                self._stat_species.set(str(len(species_counts)))
                self._stat_errors.set(str(skipped))
                self._stat_time.set(when)
                self._rebuild_species_list(species_counts)
                self._output_lbl.config(text=str(val_dir), fg=C["canopy"])

            self.after(0, _apply)

        # Placeholder while reconstruction reads the validation CSVs, so
        # the cards don't read a misleading zero in the meantime
        if self._stat_images.get() in ("", "0"):
            self._stat_images.set("…")
            self._stat_species.set("…")
        threading.Thread(target=_reconstruct, daemon=True).start()

    # ── RUN ───────────────────────────────────────────────────────────────────

    def _end_run(self):
        """Reset Run/Stop button state once a run finishes, stops, or errors."""
        self._job_running = False
        self._stop_event.clear()
        self._run_btn.configure(state="normal", bg=C["forest"])
        self._stop_btn.configure(state="disabled", text="Stop", bg=C["mist"])

    def _confirm_stop(self):
        if not self._job_running or self._stop_event.is_set():
            return
        if messagebox.askyesno(
            "Stop run?",
            "Stop the current run?\n\n"
            "wildtag will finish the batch of images it's part-way "
            "through, save everything classified so far, and stop. "
            "You can resume from where it left off later."):
            self._stop_event.set()
            self._stop_btn.configure(state="disabled", text="Stopping...")
            self._set_status("Stopping...", C["skip"])
            self._log("\nStop requested by user - finishing the current "
                       "batch and saving progress...", "skip")

    def _run(self):
        self._run_from_images()

    def _run_from_images(self):
        """Mode A: run detection + classification directly on images."""
        import sys, datetime

        folder_str = self._img_folder_var.get().strip()
        if not folder_str:
            messagebox.showerror(
                "No folder",
                "Please select a project folder first.")
            return
        project = Path(folder_str)
        if not project.exists():
            messagebox.showerror(
                "Not found", f"Cannot find:\n{project}")
            return

        # Look for images\ subfolder
        images_dir = project / "images"
        if not images_dir.exists():
            if messagebox.askyesno(
                "No images folder",
                f"No 'images' folder found in:\n{project}\n\n"
                f"Create it now?"):
                images_dir.mkdir()
                messagebox.showinfo(
                    "Created",
                    f"Created:\n{images_dir}\n\n"
                    f"Add your images there and run again.")
            return

        # Check deployment file exists and sites are aligned
        dep_exists = any(
            (project / name).exists()
            for name in ["deployment.csv", "deployment.xlsx",
                         "deployments.csv", "deployments.xlsx"])
        if not dep_exists:
            if not messagebox.askyesno(
                "No deployment file",
                "No deployment.csv was found in your project folder.\n\n"
                "You can still run wildtag, but you will need a deployment "
                "file before exporting to Camtrap DP.\n\n"
                "Continue anyway?"):
                return
        else:
            # Deployment exists — check site alignment, block if mismatch
            if not self._check_deployment_alignment(project):
                return

        # Check if already run
        json_out = project / "image_recognition_file.json"
        if json_out.exists():
            import datetime as dt
            mtime = json_out.stat().st_mtime
            ts    = dt.datetime.fromtimestamp(mtime).strftime(
                "%A %d %B %Y at %H:%M")
            if not messagebox.askyesno(
                "Already processed",
                f"This project has already been processed.\n\n"
                f"Results were last written on {ts}.\n\n"
                f"Running again will overwrite the existing results. Continue?"):
                return

        folder = images_dir  # images scanned from here

        # If a previous run was interrupted (crash, stop, power cut), its
        # checkpoint will still be sitting in the project folder - offer
        # to pick up where it left off instead of starting from scratch.
        resume = False
        from wt_models.engine import checkpoint_status
        cp_status = checkpoint_status(str(project))
        if cp_status["exists"] and cp_status["count"] > 0:
            ts = ""
            if cp_status["mtime"]:
                import datetime as _cdt
                ts = _cdt.datetime.fromtimestamp(
                    cp_status["mtime"]).strftime(" (%A %d %B %Y at %H:%M)")
            resume = messagebox.askyesno(
                "Resume previous run?",
                f"wildtag found an unfinished run for this project with "
                f"{cp_status['count']:,} images already processed{ts}.\n\n"
                f"Resume and continue from there? Choosing 'No' starts a "
                f"fresh run and discards that saved progress.")

        # Get classifier selection from registry
        from wt_models.registry import REGISTRY
        cls_name = self._cls_model_var.get()
        cls_id   = next((m["id"] for m in REGISTRY
                         if m["name"] == cls_name), None)

        if not cls_id:
            messagebox.showerror("Model error",
                "Could not find selected model. "
                "Please restart wildtag and try again.")
            return

        # SpeciesNet ships its model files separately (they're large and
        # not bundled in the installer). If the user picked SpeciesNet but
        # those files aren't on this machine yet, offer to download them now
        # or fall back to DeepFaune, rather than failing mid-run.
        try:
            from wt_models.registry import get_model as _get_model
            from wt_models.downloader import cache_bundle_present
            _meta = _get_model(cls_id)

            # SpeciesNet is a GPU-scale model. On a CPU-only machine it runs
            # very slowly (roughly 10+ seconds per image: it runs a full
            # MegaDetector plus a large classifier plus geolocation on every
            # image, with no empty-frame skipping). DeepFaune is far faster
            # on CPU and excellent for UK/European wildlife. If the user
            # picked SpeciesNet with the device set to CPU, warn them up
            # front and offer to switch, before any download happens.
            _dev = getattr(self, "_device_var", None)
            _dev = _dev.get().split()[0] if _dev else "cpu"
            if _meta.get("cache_bundle") and _dev != "cuda":
                use_df = messagebox.askyesno(
                    "SpeciesNet is slow without a GPU",
                    "SpeciesNet runs very slowly on a computer without a "
                    "graphics card (GPU), often around 10 seconds or more per "
                    "image. No GPU was detected on this machine, so a project "
                    "of a few thousand images could take many hours.\n\n"
                    "For UK and European wildlife, DeepFaune is far faster "
                    "and highly accurate, and it's already installed.\n\n"
                    "Switch to DeepFaune (recommended)?\n\n"
                    "Yes  -  use DeepFaune (fast)\n"
                    "No   -  continue with SpeciesNet anyway (slow)")
                if use_df:
                    cls_id = "deepfaune-v1.4"
                    df = next((m["name"] for m in REGISTRY
                               if m["id"] == "deepfaune-v1.4"), None)
                    if df:
                        self._cls_model_var.set(df)
                    _meta = _get_model(cls_id)  # re-fetch so the download
                                                # check below sees DeepFaune

            _cb = _meta.get("cache_bundle")
            if _cb and not cache_bundle_present(_cb):
                size_mb = _cb.get("size_mb", 0)
                choice = messagebox.askyesno(
                    "Download SpeciesNet?",
                    f"To use SpeciesNet, wildtag needs to download its model "
                    f"files (about {size_mb} MB). This happens once and is "
                    f"reused afterwards.\n\n"
                    f"Yes  -  download SpeciesNet now\n"
                    f"No   -  use DeepFaune (Europe) instead, already installed\n\n"
                    f"Download SpeciesNet now?")
                if not choice:
                    # Fall back to DeepFaune
                    cls_id = "deepfaune-v1.4"
                    df = next((m["name"] for m in REGISTRY
                               if m["id"] == "deepfaune-v1.4"), None)
                    if df:
                        self._cls_model_var.set(df)
                # If Yes, the actual download happens inside ensure_model()
                # during the run's model-prep step, with progress shown in
                # the run log, so nothing more to do here.
        except Exception:
            # If the check itself fails, don't block the run; ensure_model
            # will handle setup (or surface a clear error) during the run.
            pass

        det_conf  = self._det_conf_var.get()
        cls_conf  = self._cls_conf_var.get()
        quality   = self._QUALITY_MAP.get(self._quality_var.get(), 65)
        self._last_classifier_id = cls_id  # remembered for resume-sort
        do_val    = self._do_validation.get()
        chk_every = self._CHECKPOINT_MAP.get(self._checkpoint_var.get(), 200)
        geofence  = getattr(self, "_geofence_var", None)
        geofence  = geofence.get() if geofence else ""
        device    = getattr(self, "_device_var",   None)
        device    = device.get().split()[0] if device else "cpu"
        threads   = getattr(self, "_threads_var",  None)
        threads   = threads.get() if threads else max(1, (os.cpu_count() or 4) - 1)

        # Collect all images recursively
        exts = {".jpg", ".jpeg", ".png", ".JPG", ".JPEG", ".PNG"}
        image_paths = sorted([
            p for p in folder.rglob("*")
            if p.suffix in exts
        ])

        if not image_paths:
            messagebox.showerror(
                "No images found",
                f"No JPG or PNG images found in:\n{folder}")
            return

        self._run_btn.configure(state="disabled", bg=C["mist"])
        self._stop_event.clear()
        self._job_running = True
        self._stop_btn.configure(state="normal", text="Stop", bg=C["mist"])
        self._set_status("Running...", C["skip"])
        self._start_progress()
        import datetime as _dt
        run_start = _dt.datetime.now()

        self._log(
            f"Starting wildtag pipeline...\n"
            f"Found {len(image_paths):,} images in {project.name}/images\n",
            "head")

        # Estimate processing time — batched detection ~0.3s/image on CPU
        secs_per_img = 0.3
        est_hours    = len(image_paths) * secs_per_img / 3600
        if est_hours < 1:
            est_str = f"~{int(est_hours*60)} minutes"
        elif est_hours < 24:
            est_str = f"~{est_hours:.1f} hours"
        else:
            est_str = f"~{est_hours/24:.1f} days"
        self._log(
            f"Estimated processing time: {est_str} on CPU "
            f"(varies by machine speed and how many animals are in images). "
            f"wildtag can run overnight — leave it running and check back.\n",
            "plain")

        def _progress(done, total):
            if total > 0:
                frac = done / total
                w    = self._prog_canvas.winfo_width()
                self._prog_canvas.coords(
                    self._prog_bar, 0, 0, int(w * frac), 4)

        def worker():
            try:
                import sys
                # Add wt_models to path
                wt_dir = Path(__file__).parent
                if str(wt_dir) not in sys.path:
                    sys.path.insert(0, str(wt_dir))

                from wt_models.engine import run_pipeline

                # Run detection + classification
                results, stopped = run_pipeline(
                    image_paths   = image_paths,
                    classifier_id = cls_id,
                    det_confidence= det_conf,
                    cls_confidence= cls_conf,
                    geofence      = geofence,
                    device        = device,
                    threads       = threads,
                    project_dir   = str(project),
                    log           = self._log,
                    progress      = lambda d,t: self.after(
                        0, lambda: _progress(d, t)),
                    stop_flag     = self._stop_event.is_set,
                    resume        = resume,
                    checkpoint_every = chk_every,
                )

                if not results:
                    self._log("No detections found.", "skip")
                    self.after(0, self._stop_progress)
                    self.after(0, self._end_run)
                    return

                # Write results as JSON in the same format wildtag already
                # knows how to parse, then hand off to existing pipeline
                self._log(
                    f"\nWriting results for {len(results):,} detections...",
                    "head")

                # Build a lightweight results CSV directly from engine output
                import csv as csv_mod
                results_csv = project / "results.csv"

                # Delete stale files from previous runs
                for stale in [results_csv,
                               project / "results_with_ids.csv"]:
                    try: stale.unlink()
                    except FileNotFoundError: pass

                fields = list(results[0].keys())
                with open(results_csv, "w", newline="",
                          encoding="utf-8") as f:
                    w = csv_mod.DictWriter(f, fieldnames=fields)
                    w.writeheader()
                    w.writerows(results)

                self._log(f"  Saved: {results_csv.name}", "ok")

                # Always generate unique IDs, then sort if requested
                success = skipped = 0
                species_counts = {}
                output_dir = project / "validation"

                enriched = enrich_csv(results_csv, self._log)
                if do_val:
                    success, skipped, species_counts = \
                        sort_detections_counted(
                            enriched, quality, None, self._log,
                            classifier_id=cls_id)

                self.after(0, lambda: self._update_summary(
                    success, skipped, species_counts, output_dir))
                self.after(0, lambda: self._set_status(
                    "Stopped" if stopped else "Complete",
                    C["skip"] if stopped else C["forest"]))

                # Write run summary file
                run_end      = _dt.datetime.now()
                run_duration = run_end - run_start
                hours, rem   = divmod(int(run_duration.total_seconds()), 3600)
                mins, secs   = divmod(rem, 60)
                duration_str = (f"{hours}h {mins}m {secs}s" if hours
                                else f"{mins}m {secs}s")

                # Count sites from results
                n_sites = len({
                    r.get("locationName","").strip()
                    for r in results if r.get("locationName","").strip()
                })

                summary_lines = [
                    "wildtag.ai - Run Summary" + (" (STOPPED EARLY)" if stopped else ""),
                    "=" * 40,
                    f"Project:       {project.name}",
                    f"Date:          {run_start.strftime('%A %d %B %Y')}",
                    f"Start time:    {run_start.strftime('%H:%M:%S')}",
                    f"End time:      {run_end.strftime('%H:%M:%S')}",
                    f"Duration:      {duration_str}",
                    "",
                    "Model",
                    "-" * 40,
                    f"Classifier:    {cls_name}",
                    f"Device:        {device}",
                    f"Threads:       {threads}",
                    f"Det threshold: {det_conf}",
                    f"Cls threshold: {cls_conf}",
                    f"Geofence:      {geofence or 'none'}",
                    "",
                    "Results",
                    "-" * 40,
                    f"Images:        {len(image_paths):,}",
                    f"Detections:    {len(results):,}",
                    f"Sites:         {n_sites}",
                    f"Species:       {len(species_counts)}",
                    f"Processed:     {success:,}",
                    f"Skipped:       {skipped:,}",
                    "",
                    "Species counts",
                    "-" * 40,
                ]
                for sp, count in sorted(species_counts.items(),
                                        key=lambda x: -x[1]):
                    summary_lines.append(
                        f"  {sp.replace('_',' ').title():<30} {count:,}")

                summary_path = project / "wildtag_run_summary.txt"
                with open(summary_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(summary_lines) + "\n")
                self._log(f"  Saved run summary: {summary_path.name}", "ok")

                self.after(0, lambda: self._show_pane("summary"))
                if stopped:
                    self.after(0, lambda: messagebox.showinfo(
                        "wildtag.ai",
                        f"Run stopped.\n\n"
                        f"{success:,} images processed before stopping\n"
                        f"{len(species_counts)} species found so far\n"
                        f"Duration: {duration_str}\n\n"
                        f"Progress was saved. Start a run on this project "
                        f"again to resume from here.\n\n"
                        f"Run summary saved to:\n{summary_path.name}"))
                else:
                    self.after(0, lambda: messagebox.showinfo(
                        "wildtag.ai",
                        f"Pipeline complete.\n\n"
                        f"{success:,} images processed\n"
                        f"{len(species_counts)} species found\n"
                        f"Duration: {duration_str}\n\n"
                        f"Run summary saved to:\n{summary_path.name}"))

            except Exception as e:
                err = str(e)
                self._log(f"\nERROR: {err}", "error")
                self.after(0, lambda: self._set_status(
                    "Error", C["error"]))
                self.after(0, lambda m=err: messagebox.showerror(
                    "Error", m))
            finally:
                self.after(0, self._stop_progress)
                self.after(0, self._end_run)

        threading.Thread(target=worker, daemon=True).start()


    def _persist_setting(self, key, var):
        self._settings[key] = var.get()
        save_settings(self._settings)

    def _restore_settings(self):
        s = self._settings
        if s.get("csv"):     self._csv_var.set(s["csv"])
        if s.get("quality") in ("low","medium","high"):
            self._quality_var.set(s["quality"])
        if s.get("checkpoint") in ("frequent","balanced","infrequent"):
            self._checkpoint_var.set(s["checkpoint"])



# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="wildtag.ai camera trap pipeline")
    parser.add_argument("--csv",           default="")
    parser.add_argument("--quality",       type=int, default=60)
    parser.add_argument("--max-long-edge", type=int, default=1920)
    parser.add_argument("--no-gui",        action="store_true")
    args = parser.parse_args()

    if args.no_gui:
        if not args.csv:
            sys.exit("--no-gui requires --csv")
        p = Path(args.csv)
        if not p.exists():
            sys.exit(f"CSV not found: {p}")
        mle = args.max_long_edge if args.max_long_edge > 0 else None
        run_pipeline(p, args.quality, mle, lambda m, t="plain": print(m))
    else:
        try:
            WildTagApp().mainloop()
        except Exception as e:
            import traceback, datetime
            tb = traceback.format_exc()
            # Write crash log alongside wildtag.py
            log_path = Path(__file__).parent / "wildtag_crash.log"
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n{'='*60}\n")
                f.write(f"Crash at {datetime.datetime.now()}\n")
                f.write(tb)
            # Try to show a dialog
            try:
                import tkinter.messagebox as _mb
                _mb.showerror("wildtag.ai crashed",
                    f"An error occurred:\n\n{e}\n\n"
                    f"Details saved to:\n{log_path}")
            except Exception:
                pass
            raise

if __name__ == "__main__":
    main()
