"""
wt_models/camtrapdp.py
======================
Exports wildtag project data as a Camtrap DP package.

Camtrap DP is the TDWG open standard for camera trap data exchange.
See: https://camtrap-dp.tdwg.org

Generates three CSV files and a datapackage.json:
  deployments.csv   - camera trap placements
  media.csv         - image files with timestamps
  observations.csv  - species detections

Usage:
  from wt_models.camtrapdp import export
  export(project_dir, output_dir=None)
"""

import csv
import json
import hashlib
import re
from datetime import datetime, timezone
from pathlib import Path


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_id(*parts):
    """Make a short stable ID from parts."""
    raw = "_".join(str(p).strip().lower().replace(" ", "_") for p in parts)
    raw = re.sub(r"[^a-z0-9_]", "", raw)
    return raw[:64] or "unknown"

def _iso_datetime(date_str: str, end_of_day: bool = False) -> str:
    """
    Convert YYYY-MM-DD (or already ISO) to ISO 8601 with UTC timezone.
    end_of_day=True sets time to 23:59:59 for deployment end dates.
    """
    if not date_str or not date_str.strip():
        return ""
    s = date_str.strip()
    # Already has time component
    if "T" in s:
        if not s.endswith("Z") and "+" not in s[-6:]:
            s += "Z"
        return s
    # Date only — add time
    try:
        datetime.strptime(s, "%Y-%m-%d")
        t = "23:59:59" if end_of_day else "00:00:00"
        return f"{s}T{t}Z"
    except ValueError:
        # Try DD/MM/YYYY
        try:
            dt = datetime.strptime(s, "%d/%m/%Y")
            t = "23:59:59" if end_of_day else "00:00:00"
            return dt.strftime(f"%Y-%m-%dT{t}Z")
        except ValueError:
            return s

def _read_deployment_file(project_dir: Path):
    """
    Read deployment.csv or deployment.xlsx from project folder.
    Returns list of row dicts.
    """
    for name in ["deployment.csv", "deployment.xlsx",
                 "deployments.csv", "deployments.xlsx"]:
        p = project_dir / name
        if p.exists():
            if p.suffix.lower() == ".xlsx":
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(p, data_only=True)
                    ws = wb.active
                    headers = [str(c.value).strip() for c in next(ws.iter_rows())]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(v is not None for v in row):
                            rows.append(dict(zip(headers, [
                                str(v).strip() if v is not None else ""
                                for v in row])))
                    return rows
                except ImportError:
                    raise ImportError(
                        "openpyxl is required to read .xlsx files. "
                        "Please save your deployment file as .csv instead.")
            else:
                with open(p, newline="", encoding="utf-8-sig") as f:
                    return list(csv.DictReader(f))
    return None


# ── Main export ───────────────────────────────────────────────────────────────

def export(project_dir: Path, output_dir: Path = None, log=None) -> Path:
    """
    Export a wildtag project as a Camtrap DP package.

    Args:
        project_dir: Path to the project folder (contains images/, results_with_ids.csv)
        output_dir:  Where to write camtrapdp/ folder. Defaults to project_dir.
        log:         Optional logging callable

    Returns:
        Path to the output folder.
    """
    if log is None:
        log = print

    project_dir = Path(project_dir)
    if output_dir is None:
        output_dir = project_dir / "camtrapdp"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    log(f"Exporting Camtrap DP package to {output_dir}...")

    # ── 1. Read deployment file ───────────────────────────────────────────────
    dep_rows = _read_deployment_file(project_dir)
    if not dep_rows:
        raise FileNotFoundError(
            "No deployment file found in project folder. "
            "Please create deployment.csv with columns: "
            "locationName, latitude, longitude, deploymentStart, deploymentEnd")

    log(f"  Found {len(dep_rows)} deployment(s)")

    # ── 2. Build deployments.csv ──────────────────────────────────────────────
    DEP_FIELDS = [
        "deploymentID", "locationID", "locationName",
        "latitude", "longitude", "coordinateUncertainty",
        "deploymentStart", "deploymentEnd",
        "setupBy", "cameraID", "cameraModel",
        "cameraDelay", "cameraHeight", "cameraDepth",
        "cameraTilt", "cameraHeading", "detectionDistance",
        "timestampIssues", "baitUse", "featureType",
        "habitat", "deploymentGroups", "deploymentTags",
        "deploymentComments",
    ]

    deployments = []
    dep_lookup  = {}  # locationName -> deploymentID

    for row in dep_rows:
        loc_name  = row.get("locationName","").strip()
        dep_start = _iso_datetime(row.get("deploymentStart",""), end_of_day=False)
        dep_id    = _make_id(loc_name, row.get("deploymentStart",""))
        loc_id    = _make_id(loc_name)

        dep_lookup[loc_name] = dep_id

        dep = {f: "" for f in DEP_FIELDS}
        dep.update({
            "deploymentID":    dep_id,
            "locationID":      loc_id,
            "locationName":    loc_name,
            "latitude":        row.get("latitude","").strip(),
            "longitude":       row.get("longitude","").strip(),
            "deploymentStart": dep_start,
            "deploymentEnd":   _iso_datetime(row.get("deploymentEnd",""), end_of_day=True),
            "setupBy":         row.get("setupBy","").strip(),
            "cameraModel":     row.get("cameraModel","").strip(),
            "habitat":         row.get("habitat","").strip(),
            "deploymentComments": row.get("deploymentComments","").strip(),
        })
        # Copy any extra fields from user into deploymentTags
        extra = {k: v for k, v in row.items()
                 if k not in dep and v.strip()}
        if extra:
            dep["deploymentTags"] = " | ".join(
                f"{k}:{v}" for k, v in extra.items())

        deployments.append(dep)

    with open(output_dir / "deployments.csv", "w",
              newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=DEP_FIELDS)
        w.writeheader()
        w.writerows(deployments)
    log(f"  Wrote deployments.csv ({len(deployments)} rows)")

    # ── 3. Read results_with_ids.csv ──────────────────────────────────────────
    results_path = project_dir / "results_with_ids.csv"
    if not results_path.exists():
        results_path = project_dir / "results.csv"
    if not results_path.exists():
        raise FileNotFoundError(
            "No results file found. Please run wildtag first.")

    with open(results_path, newline="", encoding="utf-8") as f:
        results = list(csv.DictReader(f))
    log(f"  Read {len(results)} detections from {results_path.name}")

    # ── 4. Build media.csv ────────────────────────────────────────────────────
    MEDIA_FIELDS = [
        "mediaID", "deploymentID", "captureMethod",
        "timestamp", "filePath", "filePublic",
        "fileName", "fileMediatype", "favorite", "mediaComments",
    ]

    # Group by image_id to avoid duplicate media rows
    seen_media = {}
    media_rows = []

    for r in results:
        img_id   = r.get("image_id","").strip()
        rel_path = r.get("relative_path","").strip()
        if not img_id or img_id in seen_media:
            continue

        # Match to deployment by folder/site name
        dep_id = ""
        parts  = Path(rel_path).parts
        if parts:
            site = parts[0]
            # Try exact match then fuzzy
            dep_id = dep_lookup.get(site, "")
            if not dep_id:
                for loc, did in dep_lookup.items():
                    if loc.lower() in site.lower() or site.lower() in loc.lower():
                        dep_id = did
                        break
        # Fallback: use first deployment
        if not dep_id and deployments:
            dep_id = deployments[0]["deploymentID"]

        # Timestamp from EXIF
        ts_raw = (r.get("DateTimeOriginal","") or
                  r.get("DateTime","")).strip()
        if ts_raw:
            try:
                dt = datetime.strptime(ts_raw, "%Y:%m:%d %H:%M:%S")
                ts = dt.strftime("%Y-%m-%dT%H:%M:%SZ")
            except ValueError:
                ts = ts_raw
        else:
            ts = ""

        media_row = {
            "mediaID":       img_id,
            "deploymentID":  dep_id,
            "captureMethod": "activityDetection",
            "timestamp":     ts,
            "filePath":      rel_path.replace("\\", "/"),
            "filePublic":    "true",
            "fileName":      Path(rel_path).name,
            "fileMediatype": "image/jpeg",
            "favorite":      "",
            "mediaComments": "",
        }
        media_rows.append(media_row)
        seen_media[img_id] = dep_id

    with open(output_dir / "media.csv", "w",
              newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MEDIA_FIELDS)
        w.writeheader()
        w.writerows(media_rows)
    log(f"  Wrote media.csv ({len(media_rows)} rows)")

    # ── 5. Build observations.csv ─────────────────────────────────────────────
    OBS_FIELDS = [
        "observationID", "deploymentID", "mediaID", "eventID",
        "eventStart", "eventEnd",
        "observationLevel", "observationType",
        "scientificName", "count",
        "lifeStage", "sex", "behavior", "individualID",
        "classificationMethod", "classifiedBy",
        "classificationTimestamp", "classificationProbability",
        "bboxX", "bboxY", "bboxWidth", "bboxHeight",
        "observationTags", "observationComments",
    ]

    # Species name -> scientific name mapping (common wildtag labels)
    COMMON_TO_SCIENTIFIC = {
        "roe_deer":           "Capreolus capreolus",
        "red_deer":           "Cervus elaphus",
        "fallow_deer":        "Dama dama",
        "sika_deer":          "Cervus nippon",
        "muntjac":            "Muntiacus reevesi",
        "chinese_water_deer": "Hydropotes inermis",
        "red_fox":            "Vulpes vulpes",
        "fox":                "Vulpes vulpes",
        "badger":             "Meles meles",
        "european_badger":    "Meles meles",
        "otter":              "Lutra lutra",
        "stoat":              "Mustela erminea",
        "weasel":             "Mustela nivalis",
        "polecat":            "Mustela putorius",
        "pine_marten":        "Martes martes",
        "mink":               "Neovison vison",
        "american_mink":      "Neovison vison",
        "wildcat":            "Felis silvestris",
        "rabbit":             "Oryctolagus cuniculus",
        "brown_hare":         "Lepus europaeus",
        "mountain_hare":      "Lepus timidus",
        "red_squirrel":       "Sciurus vulgaris",
        "grey_squirrel":      "Sciurus carolinensis",
        "squirrel":           "Sciurus sp.",
        "hedgehog":           "Erinaceus europaeus",
        "wild_boar":          "Sus scrofa",
        "european_roe_deer":  "Capreolus capreolus",
        "european_red_deer":  "Cervus elaphus",
    }

    # observationType mapping
    def _obs_type(label: str) -> str:
        l = label.lower()
        if l in ("empty", "blank"):      return "blank"
        if l == "human":                 return "human"
        if l == "vehicle":               return "vehicle"
        if l in ("low_confidence",
                 "unclassified",
                 "unknown"):             return "unknown"
        return "animal"

    now_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    observations = []

    for i, r in enumerate(results):
        det_id = r.get("detection_id","").strip()
        img_id = r.get("image_id","").strip()
        label  = (r.get("correct_label","").strip() or
                  r.get("label","").strip())

        dep_id = seen_media.get(img_id, "")
        if not dep_id and deployments:
            dep_id = deployments[0]["deploymentID"]

        ts_raw = (r.get("DateTimeOriginal","") or
                  r.get("DateTime","")).strip()
        if ts_raw:
            try:
                dt = datetime.strptime(ts_raw, "%Y:%m:%d %H:%M:%S")
                ts = dt.strftime("%Y-%m-%dT%H:%M:%SZ")
            except ValueError:
                ts = ts_raw
        else:
            ts = ""

        obs_type   = _obs_type(label)
        sci_name   = COMMON_TO_SCIENTIFIC.get(label, "")
        conf_str   = r.get("confidence","").strip()
        verified   = r.get("human_verified","").strip().lower() == "true"
        cls_method = "human" if verified else "machine"
        cls_by     = ("wildtag.ai human validator" if verified
                      else f"wildtag.ai / {r.get('cv_model','')}")

        # Bbox — convert from xyxy to xywh if normalised
        bbox_x = bbox_y = bbox_w = bbox_h = ""
        try:
            if r.get("bbox_normalised","") == "1":
                x0 = float(r.get("bbox_left",0) or 0)
                y0 = float(r.get("bbox_top",0) or 0)
                x1 = float(r.get("bbox_right",0) or 0)
                y1 = float(r.get("bbox_bottom",0) or 0)
                if x1 > x0 and y1 > y0:
                    bbox_x = str(round(x0, 6))
                    bbox_y = str(round(y0, 6))
                    bbox_w = str(round(x1 - x0, 6))
                    bbox_h = str(round(y1 - y0, 6))
        except Exception:
            pass

        obs = {f: "" for f in OBS_FIELDS}
        obs.update({
            "observationID":             det_id or f"obs{i+1}",
            "deploymentID":              dep_id,
            "mediaID":                   img_id,
            "eventStart":                ts,
            "eventEnd":                  ts,
            "observationLevel":          "media",
            "observationType":           obs_type,
            "scientificName":            sci_name,
            "count":                     "1" if obs_type == "animal" else "",
            "classificationMethod":      cls_method,
            "classifiedBy":              cls_by,
            "classificationTimestamp":   now_ts,
            "classificationProbability": conf_str if conf_str not in ("NA","") else "",
            "bboxX":                     bbox_x,
            "bboxY":                     bbox_y,
            "bboxWidth":                 bbox_w,
            "bboxHeight":                bbox_h,
        })
        observations.append(obs)

    with open(output_dir / "observations.csv", "w",
              newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=OBS_FIELDS)
        w.writeheader()
        w.writerows(observations)
    log(f"  Wrote observations.csv ({len(observations)} rows)")

    # ── 6. Write datapackage.json ─────────────────────────────────────────────
    pkg = {
        "name":    project_dir.name.lower().replace(" ", "_"),
        "profile": "https://raw.githubusercontent.com/tdwg/camtrap-dp/1.0/camtrap-dp-profile.json",
        "created": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "project": {
            "title":          project_dir.name,
            "samplingDesign": "systematic random",
            "captureMethod":  ["activityDetection"],
            "individualAnimals": False,
            "observationLevel": ["media"],
        },
        "resources": [
            {
                "name":     "deployments",
                "path":     "deployments.csv",
                "profile":  "tabular-data-resource",
                "format":   "csv",
                "encoding": "UTF-8",
                "schema":   "https://raw.githubusercontent.com/tdwg/camtrap-dp/1.0/deployments-table-schema.json",
            },
            {
                "name":     "media",
                "path":     "media.csv",
                "profile":  "tabular-data-resource",
                "format":   "csv",
                "encoding": "UTF-8",
                "schema":   "https://raw.githubusercontent.com/tdwg/camtrap-dp/1.0/media-table-schema.json",
            },
            {
                "name":     "observations",
                "path":     "observations.csv",
                "profile":  "tabular-data-resource",
                "format":   "csv",
                "encoding": "UTF-8",
                "schema":   "https://raw.githubusercontent.com/tdwg/camtrap-dp/1.0/observations-table-schema.json",
            },
        ],
    }

    with open(output_dir / "datapackage.json", "w",
              encoding="utf-8") as f:
        json.dump(pkg, f, indent=2)
    log(f"  Wrote datapackage.json")

    log(f"\nCamtrap DP export complete: {output_dir}")
    return output_dir
